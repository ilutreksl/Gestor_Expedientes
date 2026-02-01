"""
Módulo para estadísticas de resolución de expedientes completados
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog, Toplevel, Listbox, Scrollbar, MULTIPLE, END
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
matplotlib.use('Agg')  # Backend sin GUI
from app import connect_db
from lib.logger_config import get_logger

logger = get_logger()


def mostrar_estadisticas_resolucion(ventana_principal):
    """
    Muestra estadísticas de resolución de expedientes completados con filtros y exportación.
    
    Args:
        ventana_principal: Instancia de VentanaPrincipal con acceso a main_stats_frame y master
    """
    # Limpiar el frame principal
    ventana_principal.limpiar_marco_stats()
    
    # Título
    ctk.CTkLabel(
        ventana_principal.main_stats_frame, 
        text="📋 ESTADÍSTICAS DE RESOLUCIÓN - EXPEDIENTES COMPLETADOS", 
        font=ctk.CTkFont(size=18, weight="bold")
    ).pack(pady=20)
    
    # Obtener conexión a la base de datos
    conn, cursor = ventana_principal.master.conectar_db()
    if not conn:
        ctk.CTkLabel(
            ventana_principal.main_stats_frame, 
            text="Error al conectar con la base de datos.", 
            text_color="red"
        ).pack(pady=20)
        return
    
    cursor = conn.cursor()
    
    # Obtener clientes únicos para el filtro
    try:
        cursor.execute("""
            SELECT DISTINCT cliente 
            FROM rma_maestro 
            WHERE cliente IS NOT NULL AND cliente != ''
            AND (estado = 'Completado' OR fecha_gestion IS NOT NULL)
            ORDER BY cliente ASC
        """)
        clientes = ["Todos"] + [fila[0] for fila in cursor.fetchall()]
    except Exception as e:
        print(f"Error al obtener clientes: {e}")
        clientes = ["Todos"]
    
    # Variable para almacenar clientes seleccionados
    clientes_seleccionados = []
    
    # Frame de filtros
    filtros_frame = ctk.CTkFrame(ventana_principal.main_stats_frame)
    filtros_frame.pack(fill="x", padx=20, pady=10)
    
    # Fila 1: Cliente - Selector múltiple
    ctk.CTkLabel(filtros_frame, text="Cliente:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
    
    btn_seleccionar_clientes = ctk.CTkButton(
        filtros_frame,
        text="Seleccionar Clientes (Todos)",
        command=lambda: abrir_selector_clientes(clientes, clientes_seleccionados, btn_seleccionar_clientes, cargar_datos_wrapper),
        width=250
    )
    btn_seleccionar_clientes.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
    
    # Fila 2: Fechas
    ctk.CTkLabel(filtros_frame, text="Fecha Desde:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
    entry_fecha_desde = ctk.CTkEntry(filtros_frame, placeholder_text="DD/MM/AAAA", width=150)
    entry_fecha_desde.grid(row=1, column=1, padx=10, pady=5, sticky="w")
    
    ctk.CTkLabel(filtros_frame, text="Fecha Hasta:").grid(row=1, column=2, padx=10, pady=5, sticky="w")
    entry_fecha_hasta = ctk.CTkEntry(filtros_frame, placeholder_text="DD/MM/AAAA", width=150)
    entry_fecha_hasta.grid(row=1, column=3, padx=10, pady=5, sticky="ew")
    
    filtros_frame.grid_columnconfigure(1, weight=1)
    filtros_frame.grid_columnconfigure(3, weight=1)
    
    # Frame para KPIs (Panel superior)
    kpis_frame = ctk.CTkFrame(ventana_principal.main_stats_frame, fg_color="#f8f9fa")
    kpis_frame.pack(fill="x", padx=20, pady=10)
    
    # Grid de 4 columnas para KPIs
    kpis_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
    
    # KPI 1: Ticket Promedio
    kpi1_frame = ctk.CTkFrame(kpis_frame, fg_color="#e3f2fd", corner_radius=10)
    kpi1_frame.grid(row=0, column=0, padx=5, pady=10, sticky="ew")
    ctk.CTkLabel(kpi1_frame, text="💰 Ticket Promedio", font=ctk.CTkFont(size=10)).pack(pady=(5, 0))
    lbl_ticket_promedio = ctk.CTkLabel(kpi1_frame, text="0.00 €", font=ctk.CTkFont(size=16, weight="bold"), text_color="#1976d2")
    lbl_ticket_promedio.pack(pady=(0, 5))
    
    # KPI 2: % Abonos
    kpi2_frame = ctk.CTkFrame(kpis_frame, fg_color="#ffebee", corner_radius=10)
    kpi2_frame.grid(row=0, column=1, padx=5, pady=10, sticky="ew")
    ctk.CTkLabel(kpi2_frame, text="🔴 % Abonos", font=ctk.CTkFont(size=10)).pack(pady=(5, 0))
    lbl_porcentaje_abonos = ctk.CTkLabel(kpi2_frame, text="0%", font=ctk.CTkFont(size=16, weight="bold"), text_color="#c62828")
    lbl_porcentaje_abonos.pack(pady=(0, 5))
    
    # KPI 3: Coste Promedio/Artículo
    kpi3_frame = ctk.CTkFrame(kpis_frame, fg_color="#fff3e0", corner_radius=10)
    kpi3_frame.grid(row=0, column=2, padx=5, pady=10, sticky="ew")
    ctk.CTkLabel(kpi3_frame, text="📦 Coste/Artículo", font=ctk.CTkFont(size=10)).pack(pady=(5, 0))
    lbl_coste_articulo = ctk.CTkLabel(kpi3_frame, text="0.00 €", font=ctk.CTkFont(size=16, weight="bold"), text_color="#e65100")
    lbl_coste_articulo.pack(pady=(0, 5))
    
    # KPI 4: Cliente con más incidencias
    kpi4_frame = ctk.CTkFrame(kpis_frame, fg_color="#f3e5f5", corner_radius=10)
    kpi4_frame.grid(row=0, column=3, padx=5, pady=10, sticky="ew")
    ctk.CTkLabel(kpi4_frame, text="⭐ Cliente + Incidencias", font=ctk.CTkFont(size=10)).pack(pady=(5, 0))
    lbl_top_cliente = ctk.CTkLabel(kpi4_frame, text="-", font=ctk.CTkFont(size=12, weight="bold"), text_color="#6a1b9a")
    lbl_top_cliente.pack(pady=(0, 5))
    
    # Frame contenedor para tabla y gráfico
    contenido_frame = ctk.CTkFrame(ventana_principal.main_stats_frame)
    contenido_frame.pack(fill="both", expand=True, padx=20, pady=10)
    contenido_frame.grid_columnconfigure(0, weight=2)  # Tabla más ancha
    contenido_frame.grid_columnconfigure(1, weight=1)  # Gráfico
    contenido_frame.grid_rowconfigure(0, weight=1)
    
    # Frame para resultados (tabla)
    resultados_container = ctk.CTkFrame(contenido_frame)
    resultados_container.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
    
    # Frame para gráfico
    grafico_container = ctk.CTkFrame(contenido_frame)
    grafico_container.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
    
    # Frame para totales
    total_frame = ctk.CTkFrame(ventana_principal.main_stats_frame)
    total_frame.pack(fill="x", padx=20, pady=10)
    
    lbl_total_expedientes = ctk.CTkLabel(
        total_frame, 
        text="Total Expedientes: 0", 
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#3b82f6"
    )
    lbl_total_expedientes.pack(side="left", padx=20)
    
    lbl_total_coste = ctk.CTkLabel(
        total_frame, 
        text="Coste Total: 0.00 €", 
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#ef4444"
    )
    lbl_total_coste.pack(side="left", padx=20)
    
    def cargar_datos():
        """Carga los datos según los filtros aplicados."""
        # Limpiar resultados anteriores
        for widget in resultados_container.winfo_children():
            widget.destroy()
        
        # Obtener valores de filtros
        fecha_desde = entry_fecha_desde.get().strip()
        fecha_hasta = entry_fecha_hasta.get().strip()
        
        # Estados de artículos que indican fallo de producto
        ESTADOS_FALLO = [
            'NO FUNCIONA, ABONAR',
            'NO FUNCIONA ; NO ABONAR',
            'REPOSICION FALLO PRODUCTO',
            'REPOSICION ; ABONAR',
            'FALLO SOLDADURA ; ABONAR',
            'FALLO SOLDADURA ; NO ABONAR',
            'FALLO MODULO ; ABONAR'
        ]
        
        # Construir query con detección de fallos
        # Primero detectamos por expediente si tiene fallos, luego agrupamos
        query = """
            SELECT 
                exp.resultado_expediente,
                COUNT(*) as num_expedientes,
                SUM(exp.num_articulos) as num_articulos,
                SUM(exp.coste_total) as coste_total,
                exp.tiene_fallo
            FROM (
                SELECT 
                    m.resultado_expediente,
                    m.id,
                    COUNT(d.id) as num_articulos,
                    SUM(d.cantidad_entregada * COALESCE(d.precio_final, d.precio_unitario)) as coste_total,
                    MAX(CASE 
                        WHEN d.estado_producto IN ('NO FUNCIONA, ABONAR', 'NO FUNCIONA ; NO ABONAR', 
                                          'REPOSICION FALLO PRODUCTO', 'REPOSICION ; ABONAR',
                                          'FALLO SOLDADURA ; ABONAR', 'FALLO SOLDADURA ; NO ABONAR',
                                          'FALLO MODULO ; ABONAR')
                        THEN 1 
                        ELSE 0 
                    END) as tiene_fallo
                FROM rma_maestro m
                LEFT JOIN rma_detalles d ON m.id = d.rma_id
                WHERE (m.estado = 'Completado' OR m.fecha_gestion IS NOT NULL)
                AND m.resultado_expediente IS NOT NULL
                AND m.resultado_expediente != ''
        """
        
        params = []
        
        # Filtro múltiple de clientes
        if clientes_seleccionados and "Todos" not in clientes_seleccionados:
            placeholders = ",".join(["?" for _ in clientes_seleccionados])
            query += f" AND m.cliente IN ({placeholders})"
            params.extend(clientes_seleccionados)
        
        # Filtros de fecha (usando fecha_gestion como referencia para completados)
        if fecha_desde:
            try:
                fecha_obj = datetime.strptime(fecha_desde, "%d/%m/%Y")
                fecha_sql = fecha_obj.strftime("%Y-%m-%d")
                query += " AND m.fecha_gestion >= ?"
                params.append(fecha_sql)
            except ValueError:
                messagebox.showwarning("Fecha inválida", "El formato de 'Fecha Desde' debe ser DD/MM/AAAA")
                return
        
        if fecha_hasta:
            try:
                fecha_obj = datetime.strptime(fecha_hasta, "%d/%m/%Y")
                fecha_sql = fecha_obj.strftime("%Y-%m-%d")
                query += " AND m.fecha_gestion <= ?"
                params.append(fecha_sql)
            except ValueError:
                messagebox.showwarning("Fecha inválida", "El formato de 'Fecha Hasta' debe ser DD/MM/AAAA")
                return
        
        # Cerrar subconsulta y agrupar por resultado y tiene_fallo
        query += """
                GROUP BY m.id, m.resultado_expediente
            ) exp
            GROUP BY exp.resultado_expediente, exp.tiene_fallo
            ORDER BY num_expedientes DESC
        """
        
        # Ejecutar query
        try:
            cursor.execute(query, tuple(params))
            registros = cursor.fetchall()
            
            if not registros:
                ctk.CTkLabel(
                    resultados_container, 
                    text="No se encontraron expedientes completados con los filtros aplicados.", 
                    text_color="gray"
                ).pack(pady=20)
                lbl_total_expedientes.configure(text="Total Expedientes: 0")
                lbl_total_coste.configure(text="Coste Total: 0.00 €")
                cargar_datos.ultimos_registros = []
                return
            
            # Calcular totales generales primero
            total_expedientes = sum(int(reg[1]) if reg[1] else 0 for reg in registros)
            total_articulos = sum(int(reg[2]) if reg[2] else 0 for reg in registros)
            total_coste = sum(float(reg[3]) if reg[3] else 0.0 for reg in registros)
            
            # Calcular subtotales por categoría
            subtotales = {'abonos_fallo': 0, 'abonos': 0, 'no_abonos': 0, 'reposiciones': 0, 'otros': 0}
            subtotales_coste = {'abonos_fallo': 0.0, 'abonos': 0.0, 'no_abonos': 0.0, 'reposiciones': 0.0, 'otros': 0.0}
            
            for reg in registros:
                resultado = str(reg[0]).upper() if reg[0] else ""
                num_exp = int(reg[1]) if reg[1] else 0
                num_art = int(reg[2]) if reg[2] else 0
                coste = float(reg[3]) if reg[3] else 0.0
                tiene_fallo = int(reg[4]) if len(reg) > 4 and reg[4] else 0
                
                # Clasificar según resultado y si tiene artículos con fallo
                # Si tiene fallo Y es un abono, va a "abonos_fallo"
                if tiene_fallo and "ABONAR" in resultado and "NO ABONAR" not in resultado:
                    subtotales['abonos_fallo'] += num_exp
                    subtotales_coste['abonos_fallo'] += coste
                elif "ABONAR" in resultado and "NO ABONAR" not in resultado:
                    subtotales['abonos'] += num_exp
                    subtotales_coste['abonos'] += coste
                elif "NO ABONAR" in resultado:
                    subtotales['no_abonos'] += num_exp
                    subtotales_coste['no_abonos'] += coste
                elif "REPOSICION" in resultado or "REPUESTO" in resultado:
                    subtotales['reposiciones'] += num_exp
                    subtotales_coste['reposiciones'] += coste
                else:
                    subtotales['otros'] += num_exp
                    subtotales_coste['otros'] += coste
            
            # Calcular KPIs
            ticket_promedio = total_coste / total_expedientes if total_expedientes > 0 else 0
            # Porcentaje de abonos incluye ambos tipos: por fallo y en buen estado
            total_abonos = subtotales['abonos_fallo'] + subtotales['abonos']
            porcentaje_abonos = (total_abonos / total_expedientes * 100) if total_expedientes > 0 else 0
            coste_por_articulo = total_coste / total_articulos if total_articulos > 0 else 0
            
            # Obtener cliente con más incidencias
            query_top_cliente = """
                SELECT m.cliente, COUNT(DISTINCT m.id) as total
                FROM rma_maestro m
                WHERE (m.estado = 'Completado' OR m.fecha_gestion IS NOT NULL)
            """
            if clientes_seleccionados and "Todos" not in clientes_seleccionados:
                placeholders = ",".join(["?" for _ in clientes_seleccionados])
                query_top_cliente += f" AND m.cliente IN ({placeholders})"
            
            query_top_cliente += " GROUP BY m.cliente ORDER BY total DESC LIMIT 1"
            
            cursor.execute(query_top_cliente, tuple([c for c in clientes_seleccionados if c != "Todos"]) if clientes_seleccionados and "Todos" not in clientes_seleccionados else ())
            top_cliente_result = cursor.fetchone()
            top_cliente = top_cliente_result[0] if top_cliente_result else "-"
            
            # Actualizar KPIs
            lbl_ticket_promedio.configure(text=f"{ticket_promedio:,.2f} €")
            lbl_porcentaje_abonos.configure(text=f"{porcentaje_abonos:.1f}%")
            lbl_coste_articulo.configure(text=f"{coste_por_articulo:,.2f} €")
            lbl_top_cliente.configure(text=top_cliente[:20] + "..." if len(top_cliente) > 20 else top_cliente)
            
            # Variable para ordenación
            orden_actual = {'columna': None, 'ascendente': True}
            
            def ordenar_por_columna(columna_idx):
                """Ordena la tabla por la columna seleccionada."""
                if orden_actual['columna'] == columna_idx:
                    orden_actual['ascendente'] = not orden_actual['ascendente']
                else:
                    orden_actual['columna'] = columna_idx
                    orden_actual['ascendente'] = False
                
                # Reordenar registros
                registros_ordenados = sorted(
                    registros,
                    key=lambda x: x[columna_idx] if x[columna_idx] is not None else 0,
                    reverse=not orden_actual['ascendente']
                )
                
                # Redibujar tabla
                mostrar_tabla(registros_ordenados)
            
            def mostrar_tabla(registros_a_mostrar):
                """Muestra la tabla con los registros dados."""
                # Limpiar scroll_frame
                for widget in scroll_frame.winfo_children():
                    widget.destroy()
                
                # Encabezados con botones de ordenación
                header_font = ctk.CTkFont(weight="bold", size=12)
                headers = ["RESULTADO ⇅", "Nº EXPED. ⇅", "Nº ART. ⇅", "COSTE TOTAL ⇅", "% TOTAL"]
                
                for col, header in enumerate(headers):
                    if col < 4:  # Solo las primeras 4 son ordenables
                        btn = ctk.CTkButton(
                            scroll_frame,
                            text=header,
                            command=lambda c=col: ordenar_por_columna(c),
                            width=120,
                            height=30,
                            font=header_font
                        )
                        btn.grid(row=0, column=col, padx=10, pady=10, sticky="w")
                    else:
                        lbl = ctk.CTkLabel(scroll_frame, text=header, font=header_font)
                        lbl.grid(row=0, column=col, padx=10, pady=10, sticky="w")
                
                # Datos individuales
                row_idx = 1
                for reg in registros_a_mostrar:
                    resultado = reg[0]
                    num_exp = reg[1]
                    num_art = reg[2]
                    coste = reg[3]
                    tiene_fallo = int(reg[4]) if len(reg) > 4 and reg[4] else 0
                    
                    # Normalizar valores
                    try:
                        num_exp_int = int(num_exp) if num_exp else 0
                        num_art_int = int(num_art) if num_art else 0
                        coste_float = float(coste) if coste else 0.0
                    except (ValueError, TypeError):
                        num_exp_int = 0
                        num_art_int = 0
                        coste_float = 0.0
                    
                    # Calcular porcentaje
                    porcentaje = (num_exp_int / total_expedientes * 100) if total_expedientes > 0 else 0
                    
                    # Determinar color y texto según resultado
                    resultado_upper = str(resultado).upper() if resultado else ""
                    
                    # Modificar el texto según si tiene fallo o no
                    texto_resultado = str(resultado) if resultado else "SIN ESPECIFICAR"
                    
                    # Primero verificar si tiene artículos con fallo Y es un abono (rojo oscuro)
                    if tiene_fallo and "ABONAR" in resultado_upper and "NO ABONAR" not in resultado_upper:
                        color_resultado = "#dc2626"  # Rojo oscuro para abonos con fallo
                        texto_resultado = f"{resultado} (FALLO)"
                    elif "ABONAR" in resultado_upper and "NO ABONAR" not in resultado_upper:
                        color_resultado = "#f87171"  # Rojo claro para abonos OK
                        texto_resultado = f"{resultado} (OK)"
                    elif "NO ABONAR" in resultado_upper:
                        color_resultado = "#22c55e"
                    elif "REPOSICION" in resultado_upper or "REPUESTO" in resultado_upper:
                        color_resultado = "#f97316"
                        if tiene_fallo:
                            texto_resultado = f"{resultado} (FALLO)"
                        else:
                            texto_resultado = f"{resultado} (OK)"
                    else:
                        color_resultado = "#6b7280"
                    
                    # Columnas
                    ctk.CTkLabel(
                        scroll_frame,
                        text=texto_resultado,
                        text_color=color_resultado,
                        font=ctk.CTkFont(weight="bold", size=10)
                    ).grid(row=row_idx, column=0, padx=15, pady=3, sticky="w")
                    
                    ctk.CTkLabel(scroll_frame, text=str(num_exp_int), font=ctk.CTkFont(size=10)).grid(
                        row=row_idx, column=1, padx=15, pady=3, sticky="w")
                    
                    ctk.CTkLabel(scroll_frame, text=str(num_art_int), font=ctk.CTkFont(size=10)).grid(
                        row=row_idx, column=2, padx=15, pady=3, sticky="w")
                    
                    ctk.CTkLabel(
                        scroll_frame,
                        text=f"{coste_float:,.2f} €",
                        text_color="#ef4444" if coste_float > 0 else "gray",
                        font=ctk.CTkFont(size=10)
                    ).grid(row=row_idx, column=3, padx=15, pady=3, sticky="w")
                    
                    ctk.CTkLabel(
                        scroll_frame,
                        text=f"{porcentaje:.1f}%",
                        text_color="#3b82f6",
                        font=ctk.CTkFont(size=10, weight="bold")
                    ).grid(row=row_idx, column=4, padx=15, pady=3, sticky="w")
                    
                    row_idx += 1
            
            # Función para crear gráfico
            def crear_grafico():
                """Crea un gráfico de distribución de resoluciones."""
                # Limpiar contenedor del gráfico
                for widget in grafico_container.winfo_children():
                    widget.destroy()
                
                if not registros:
                    ctk.CTkLabel(
                        grafico_container,
                        text="📊\n\nNo hay datos\npara graficar",
                        font=ctk.CTkFont(size=14),
                        text_color="gray"
                    ).pack(expand=True)
                    return
                
                # Preparar datos para el gráfico
                labels = []
                sizes = []
                colors = []
                
                if subtotales['abonos_fallo'] > 0:
                    labels.append(f"Abonos Fallo\n{subtotales['abonos_fallo']} exp.")
                    sizes.append(subtotales['abonos_fallo'])
                    colors.append('#dc2626')  # Rojo más oscuro
                
                if subtotales['abonos'] > 0:
                    labels.append(f"Abonos OK\n{subtotales['abonos']} exp.")
                    sizes.append(subtotales['abonos'])
                    colors.append('#f87171')  # Rojo más claro
                
                if subtotales['no_abonos'] > 0:
                    labels.append(f"No Abonos\n{subtotales['no_abonos']} exp.")
                    sizes.append(subtotales['no_abonos'])
                    colors.append('#22c55e')
                
                # Separar reposiciones por fallo y OK
                reposiciones_fallo = 0
                reposiciones_ok = 0
                for reg in registros:
                    resultado_upper = str(reg[0]).upper() if reg[0] else ""
                    tiene_fallo = int(reg[4]) if len(reg) > 4 and reg[4] else 0
                    num_exp = int(reg[1]) if reg[1] else 0
                    if "REPOSICION" in resultado_upper or "REPUESTO" in resultado_upper:
                        if tiene_fallo:
                            reposiciones_fallo += num_exp
                        else:
                            reposiciones_ok += num_exp
                
                if reposiciones_fallo > 0:
                    labels.append(f"Reposiciones Fallo\n{reposiciones_fallo} exp.")
                    sizes.append(reposiciones_fallo)
                    colors.append('#f97316')
                
                if reposiciones_ok > 0:
                    labels.append(f"Reposiciones OK\n{reposiciones_ok} exp.")
                    sizes.append(reposiciones_ok)
                    colors.append('#fb923c')
                
                if subtotales['otros'] > 0:
                    labels.append(f"Otros\n{subtotales['otros']} exp.")
                    sizes.append(subtotales['otros'])
                    colors.append('#6b7280')
                
                # Crear figura
                fig, ax = plt.subplots(figsize=(5, 4), facecolor='white')
                
                # Crear gráfico de pastel
                wedges, texts, autotexts = ax.pie(
                    sizes,
                    labels=labels,
                    colors=colors,
                    autopct='%1.1f%%',
                    startangle=90,
                    textprops={'fontsize': 9}
                )
                
                # Estilo de los textos
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_weight('bold')
                    autotext.set_fontsize(10)
                
                ax.axis('equal')
                plt.title('Distribución de Resoluciones', fontsize=12, weight='bold', pad=15)
                plt.tight_layout()
                
                # Integrar en tkinter
                canvas = FigureCanvasTkAgg(fig, master=grafico_container)
                canvas.draw()
                canvas.get_tk_widget().pack(fill="both", expand=True, padx=5, pady=5)
            
            # Frame scrollable para la tabla
            scroll_frame = ctk.CTkScrollableFrame(resultados_container)
            scroll_frame.pack(fill="both", expand=True)
            
            # Mostrar tabla inicial
            mostrar_tabla(registros)
            
            # Crear gráfico inicial
            crear_grafico()
            
            # Actualizar totales (ya calculados al principio de cargar_datos)
            lbl_total_expedientes.configure(text=f"Total Expedientes: {total_expedientes}")
            lbl_total_coste.configure(text=f"Coste Total: {total_coste:,.2f} €")
            
            # Guardar registros para exportación
            cargar_datos.ultimos_registros = registros
            cargar_datos.totales = {
                'expedientes': total_expedientes,
                'coste': total_coste
            }
            
        except Exception as e:
            print(f"Error al cargar datos de resolución: {e}")
            messagebox.showerror("Error", f"Error al cargar los datos: {e}")
            cargar_datos.ultimos_registros = []
            cargar_datos.totales = {'expedientes': 0, 'coste': 0.0}
    
    # Inicializar variables para almacenar registros
    cargar_datos.ultimos_registros = []
    cargar_datos.totales = {'expedientes': 0, 'coste': 0.0}
    
    # Wrapper para poder llamar cargar_datos antes de definirla
    def cargar_datos_wrapper():
        cargar_datos()
    
    def exportar_a_excel():
        """Exporta los resultados actuales a Excel con formato mejorado y KPIs."""
        if not cargar_datos.ultimos_registros:
            messagebox.showwarning("Sin datos", "No hay datos para exportar. Primero aplique los filtros.")
            return
        
        try:
            # Abrir conexión para consultas adicionales
            conn_export = connect_db()
            if conn_export is None:
                messagebox.showerror("Error", "No se pudo conectar a la base de datos para exportar.")
                return
            cursor_export = conn_export.cursor()
            # Calcular totales y subtotales primero
            registros = cargar_datos.ultimos_registros
            total_expedientes = sum(int(reg[1]) if reg[1] else 0 for reg in registros)
            total_articulos = sum(int(reg[2]) if reg[2] else 0 for reg in registros)
            total_coste = sum(float(reg[3]) if reg[3] else 0.0 for reg in registros)
            
            # Calcular subtotales por categoría
            subtotales = {'abonos_fallo': 0, 'abonos': 0, 'no_abonos': 0, 'reposiciones': 0, 'otros': 0}
            subtotales_coste = {'abonos_fallo': 0.0, 'abonos': 0.0, 'no_abonos': 0.0, 'reposiciones': 0.0, 'otros': 0.0}
            
            for reg in registros:
                resultado = str(reg[0]).upper() if reg[0] else ""
                num_exp = int(reg[1]) if reg[1] else 0
                coste = float(reg[3]) if reg[3] else 0.0
                tiene_fallo = int(reg[4]) if len(reg) > 4 and reg[4] else 0
                
                # Clasificar según resultado y si tiene artículos con fallo
                if tiene_fallo and "ABONAR" in resultado and "NO ABONAR" not in resultado:
                    subtotales['abonos_fallo'] += num_exp
                    subtotales_coste['abonos_fallo'] += coste
                elif "ABONAR" in resultado and "NO ABONAR" not in resultado:
                    subtotales['abonos'] += num_exp
                    subtotales_coste['abonos'] += coste
                elif "NO ABONAR" in resultado:
                    subtotales['no_abonos'] += num_exp
                    subtotales_coste['no_abonos'] += coste
                elif "REPOSICION" in resultado or "REPUESTO" in resultado:
                    subtotales['reposiciones'] += num_exp
                    subtotales_coste['reposiciones'] += coste
                else:
                    subtotales['otros'] += num_exp
                    subtotales_coste['otros'] += coste
            
            # Preparar datos para DataFrame
            datos_export = []
            for reg in registros:
                resultado = reg[0]
                num_exp = reg[1]
                num_art = reg[2]
                coste = reg[3]
                tiene_fallo = int(reg[4]) if len(reg) > 4 and reg[4] else 0
                
                try:
                    num_exp_int = int(num_exp) if num_exp else 0
                    num_art_int = int(num_art) if num_art else 0
                    coste_float = float(coste) if coste else 0.0
                except (ValueError, TypeError):
                    num_exp_int = 0
                    num_art_int = 0
                    coste_float = 0.0
                
                porcentaje = (num_exp_int / total_expedientes * 100) if total_expedientes > 0 else 0
                
                # Determinar texto del resultado
                resultado_texto = str(resultado) if resultado else "SIN ESPECIFICAR"
                resultado_upper = resultado_texto.upper()
                
                if tiene_fallo and "ABONAR" in resultado_upper and "NO ABONAR" not in resultado_upper:
                    resultado_texto = f"{resultado} (FALLO)"
                elif "ABONAR" in resultado_upper and "NO ABONAR" not in resultado_upper:
                    resultado_texto = f"{resultado} (OK)"
                elif ("REPOSICION" in resultado_upper or "REPUESTO" in resultado_upper):
                    if tiene_fallo:
                        resultado_texto = f"{resultado} (FALLO)"
                    else:
                        resultado_texto = f"{resultado} (OK)"
                
                datos_export.append({
                    'Resultado': resultado_texto,
                    'Nº Expedientes': num_exp_int,
                    'Nº Artículos': num_art_int,
                    'Coste Total (€)': coste_float,
                    '% Total': round(porcentaje, 2)
                })
            
            # Añadir subtotales
            if subtotales['abonos_fallo'] > 0:
                datos_export.append({
                    'Resultado': '--- SUBTOTAL ABONOS FALLO ---',
                    'Nº Expedientes': subtotales['abonos_fallo'],
                    'Nº Artículos': '',
                    'Coste Total (€)': subtotales_coste['abonos_fallo'],
                    '% Total': round((subtotales['abonos_fallo'] / total_expedientes * 100), 2) if total_expedientes > 0 else 0
                })
            
            if subtotales['abonos'] > 0:
                datos_export.append({
                    'Resultado': '--- SUBTOTAL ABONOS OK ---',
                    'Nº Expedientes': subtotales['abonos'],
                    'Nº Artículos': '',
                    'Coste Total (€)': subtotales_coste['abonos'],
                    '% Total': round((subtotales['abonos'] / total_expedientes * 100), 2) if total_expedientes > 0 else 0
                })
            
            if subtotales['no_abonos'] > 0:
                datos_export.append({
                    'Resultado': '--- SUBTOTAL NO ABONOS ---',
                    'Nº Expedientes': subtotales['no_abonos'],
                    'Nº Artículos': '',
                    'Coste Total (€)': subtotales_coste['no_abonos'],
                    '% Total': round((subtotales['no_abonos'] / total_expedientes * 100), 2) if total_expedientes > 0 else 0
                })
            
            if subtotales['reposiciones'] > 0:
                datos_export.append({
                    'Resultado': '--- SUBTOTAL REPOSICIONES ---',
                    'Nº Expedientes': subtotales['reposiciones'],
                    'Nº Artículos': '',
                    'Coste Total (€)': subtotales_coste['reposiciones'],
                    '% Total': round((subtotales['reposiciones'] / total_expedientes * 100), 2) if total_expedientes > 0 else 0
                })
            
            # Añadir fila de totales
            datos_export.append({
                'Resultado': 'TOTAL GENERAL',
                'Nº Expedientes': total_expedientes,
                'Nº Artículos': total_articulos,
                'Coste Total (€)': total_coste,
                '% Total': 100.0
            })
            
            # Crear DataFrame
            df = pd.DataFrame(datos_export)
            
            # Solicitar ruta de guardado
            archivo = filedialog.asksaveasfilename(
                title="Guardar estadísticas de resolución",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                initialfile="estadisticas_resolucion_expedientes.xlsx"
            )
            
            if not archivo:
                return
            
            # Exportar a Excel con formato mejorado
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
                # Primero crear una hoja de KPIs
                kpis_data = []
                
                # Calcular KPIs
                ticket_promedio = total_coste / total_expedientes if total_expedientes > 0 else 0
                # Porcentaje de abonos incluye ambos tipos: por fallo y en buen estado
                total_abonos_excel = subtotales['abonos_fallo'] + subtotales['abonos']
                porcentaje_abonos = (total_abonos_excel / total_expedientes * 100) if total_expedientes > 0 else 0
                coste_por_articulo = total_coste / total_articulos if total_articulos > 0 else 0
                
                # Obtener cliente con más incidencias para Excel
                query_top_cliente_excel = """
                    SELECT m.cliente, COUNT(DISTINCT m.id) as total
                    FROM rma_maestro m
                    WHERE (m.estado = 'Completado' OR m.fecha_gestion IS NOT NULL)
                """
                if clientes_seleccionados and "Todos" not in clientes_seleccionados:
                    placeholders = ",".join(["?" for _ in clientes_seleccionados])
                    query_top_cliente_excel += f" AND m.cliente IN ({placeholders})"
                
                query_top_cliente_excel += " GROUP BY m.cliente ORDER BY total DESC LIMIT 1"
                
                cursor_export.execute(query_top_cliente_excel, tuple([c for c in clientes_seleccionados if c != "Todos"]) if clientes_seleccionados and "Todos" not in clientes_seleccionados else ())
                top_cliente_excel = cursor_export.fetchone()
                top_cliente_nombre = top_cliente_excel[0] if top_cliente_excel else "-"
                
                kpis_data.append({'Métrica': 'Ticket Promedio', 'Valor': f'{ticket_promedio:.2f} €'})
                kpis_data.append({'Métrica': '% Abonos', 'Valor': f'{porcentaje_abonos:.1f}%'})
                kpis_data.append({'Métrica': 'Coste por Artículo', 'Valor': f'{coste_por_articulo:.2f} €'})
                kpis_data.append({'Métrica': 'Cliente con más Incidencias', 'Valor': top_cliente_nombre})
                kpis_data.append({'Métrica': 'Total Expedientes', 'Valor': str(total_expedientes)})
                kpis_data.append({'Métrica': 'Total Artículos', 'Valor': str(total_articulos)})
                kpis_data.append({'Métrica': 'Coste Total', 'Valor': f'{total_coste:.2f} €'})
                
                df_kpis = pd.DataFrame(kpis_data)
                df_kpis.to_excel(writer, index=False, sheet_name='Resumen', startrow=1)
                
                # Hoja de datos detallados
                df.to_excel(writer, index=False, sheet_name='Datos Detallados')
                
                # Formatear hoja de KPIs
                workbook = writer.book
                ws_kpis = writer.sheets['Resumen']
                
                # Título de la hoja
                ws_kpis['A1'] = '📊 INDICADORES CLAVE (KPIs)'
                ws_kpis['A1'].font = Font(size=14, bold=True, color="1976D2")
                ws_kpis.merge_cells('A1:B1')
                
                # Formato de encabezados KPIs
                for cell in ws_kpis[2]:
                    cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Formato de valores KPIs
                for row in range(3, 9):
                    ws_kpis[f'A{row}'].font = Font(bold=True)
                    ws_kpis[f'B{row}'].font = Font(size=12, color="1976D2")
                    ws_kpis[f'B{row}'].alignment = Alignment(horizontal="right")
                
                # Ajustar anchos
                ws_kpis.column_dimensions['A'].width = 25
                ws_kpis.column_dimensions['B'].width = 20
                
                # Formatear hoja de datos
                worksheet = writer.sheets['Datos Detallados']
                
                # Ajustar ancho de columnas automáticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar formato a los encabezados
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Aplicar formato condicional: filas con "SUBTOTAL" o "TOTAL"
                subtotal_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                subtotal_font = Font(bold=True, color="1976D2")
                total_font = Font(bold=True, color="D84315")
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    resultado_valor = str(row[0].value).upper() if row[0].value else ""
                    
                    if "SUBTOTAL" in resultado_valor:
                        for cell in row:
                            cell.fill = subtotal_fill
                            cell.font = subtotal_font
                    elif "TOTAL" in resultado_valor:
                        for cell in row:
                            cell.fill = total_fill
                            cell.font = total_font
                    
                    # Formato condicional para costes altos
                    if row[3].value and isinstance(row[3].value, (int, float)):
                        if row[3].value > 1000:
                            row[3].font = Font(color="D84315", bold=True)  # Rojo
                        elif row[3].value > 500:
                            row[3].font = Font(color="F57C00", bold=True)  # Naranja
            
            conn_export.close()
            messagebox.showinfo("Éxito", f"Los datos se han exportado correctamente a:\n{archivo}\n\n✅ Incluye hoja de KPIs y datos detallados con formato mejorado.")
            
        except Exception as e:
            if 'conn_export' in locals():
                conn_export.close()
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{e}")
            print(f"Error en exportación: {e}")
    
    # Botones
    botones_frame = ctk.CTkFrame(filtros_frame)
    botones_frame.grid(row=2, column=0, columnspan=4, pady=10)
    
    btn_limpiar = ctk.CTkButton(
        botones_frame, 
        text="🗑️ Limpiar Fechas", 
        command=lambda: (entry_fecha_desde.delete(0, 'end'), entry_fecha_hasta.delete(0, 'end'), cargar_datos()),
        width=120
    )
    btn_limpiar.pack(side="left", padx=5)
    
    btn_aplicar = ctk.CTkButton(
        botones_frame, 
        text="🔍 Aplicar Filtros", 
        command=cargar_datos,
        width=140
    )
    btn_aplicar.pack(side="left", padx=5)
    
    btn_exportar = ctk.CTkButton(
        botones_frame,
        text="💾 Exportar a Excel",
        command=exportar_a_excel,
        width=140
    )
    btn_exportar.pack(side="left", padx=5)
    
    # Cargar datos iniciales
    cargar_datos()
    
    conn.close()


def abrir_selector_clientes(clientes_disponibles, clientes_seleccionados, boton, callback_actualizar):
    """Abre una ventana para seleccionar múltiples clientes."""
    
    # Crear ventana modal
    ventana_selector = Toplevel()
    ventana_selector.title("Seleccionar Clientes")
    ventana_selector.geometry("500x500")
    ventana_selector.transient()
    ventana_selector.grab_set()
    
    # Centrar ventana
    ventana_selector.update_idletasks()
    x = (ventana_selector.winfo_screenwidth() - 500) // 2
    y = (ventana_selector.winfo_screenheight() - 500) // 2
    ventana_selector.geometry(f"500x500+{x}+{y}")
    
    # Frame principal
    frame_principal = ctk.CTkFrame(ventana_selector)
    frame_principal.pack(fill="both", expand=True, padx=10, pady=10)
    
    ctk.CTkLabel(
        frame_principal,
        text="Seleccione los clientes a filtrar (múltiple):",
        font=ctk.CTkFont(size=12, weight="bold")
    ).pack(pady=10)
    
    # Frame con scrollbar para la lista
    frame_lista = ctk.CTkFrame(frame_principal)
    frame_lista.pack(fill="both", expand=True, padx=10, pady=10)
    
    scrollbar = Scrollbar(frame_lista)
    scrollbar.pack(side="right", fill="y")
    
    listbox = Listbox(frame_lista, selectmode=MULTIPLE, yscrollcommand=scrollbar.set, height=20)
    listbox.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=listbox.yview)
    
    # Añadir items a la lista
    for cliente in clientes_disponibles:
        listbox.insert(END, cliente)
        # Seleccionar si ya estaba seleccionado previamente
        if cliente in clientes_seleccionados or (not clientes_seleccionados and cliente == "Todos"):
            listbox.selection_set(clientes_disponibles.index(cliente))
    
    # Frame de botones
    frame_botones = ctk.CTkFrame(frame_principal)
    frame_botones.pack(fill="x", pady=10)
    
    def seleccionar_todos():
        listbox.selection_set(0, END)
    
    def deseleccionar_todos():
        listbox.selection_clear(0, END)
    
    def aplicar_seleccion():
        # Obtener elementos seleccionados
        indices = listbox.curselection()
        clientes_seleccionados.clear()
        
        if not indices:
            # Si no hay selección, seleccionar "Todos"
            clientes_seleccionados.append("Todos")
        else:
            for i in indices:
                clientes_seleccionados.append(listbox.get(i))
        
        # Actualizar texto del botón
        if "Todos" in clientes_seleccionados or not clientes_seleccionados:
            boton.configure(text="Seleccionar Clientes (Todos)")
        elif len(clientes_seleccionados) == 1:
            # Truncar si es muy largo
            texto = clientes_seleccionados[0]
            if len(texto) > 25:
                texto = texto[:22] + "..."
            boton.configure(text=f"Cliente: {texto}")
        else:
            boton.configure(text=f"Clientes seleccionados: {len(clientes_seleccionados)}")
        
        # Actualizar datos
        callback_actualizar()
        
        # Cerrar ventana
        ventana_selector.destroy()
    
    ctk.CTkButton(
        frame_botones,
        text="✓ Seleccionar Todos",
        command=seleccionar_todos,
        width=140
    ).pack(side="left", padx=5)
    
    ctk.CTkButton(
        frame_botones,
        text="✗ Deseleccionar Todos",
        command=deseleccionar_todos,
        width=140
    ).pack(side="left", padx=5)
    
    ctk.CTkButton(
        frame_botones,
        text="Aplicar",
        command=aplicar_seleccion,
        width=100
    ).pack(side="right", padx=5)
    
    ctk.CTkButton(
        frame_botones,
        text="Cancelar",
        command=ventana_selector.destroy,
        fg_color="#ef4444",
        width=100
    ).pack(side="right", padx=5)
