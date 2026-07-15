"""
Módulo para estadísticas de artículos con filtros por resultado de expediente y estado de artículo
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog, Toplevel, Listbox, Scrollbar, MULTIPLE, END
from collections import defaultdict
import pandas as pd
from lib.logger_config import get_logger

logger = get_logger()


def _extraer_referencia_base(ref):
    """Extrae la referencia base eliminando los sufijos -R y -R50M.
    Estos son los únicos sufijos que indican variante del mismo artículo.
    Cualquier otro segmento final (ej: -30, -40) forma parte del nombre base.

    Ejemplos:
        '24-1121'      → '24-1121'
        '24-1121-R'    → '24-1121'
        '24-1121-R50M' → '24-1121'
        '24-1717-30'   → '24-1717-30'
        '24-1717-30-R' → '24-1717-30'
    """
    if not ref:
        return ''
    if ref.endswith('-R50M'):
        return ref[:-5]
    if ref.endswith('-R'):
        return ref[:-2]
    return ref


def mostrar_estadisticas_articulos(ventana_principal):
    """
    Muestra estadísticas de artículos agrupadas por referencia base.
    Cada fila representa la suma de todos los estados y sub-variantes de una referencia.
    Al hacer clic en una referencia se muestra el desglose detallado.
    """
    ventana_principal.limpiar_marco_stats()

    ctk.CTkLabel(
        ventana_principal.main_stats_frame,
        text="📦 ESTADÍSTICAS DE ARTÍCULOS",
        font=ctk.CTkFont(size=18, weight="bold")
    ).pack(pady=20)

    conn, cursor = ventana_principal.master.conectar_db()
    if not conn:
        ctk.CTkLabel(
            ventana_principal.main_stats_frame,
            text="Error al conectar con la base de datos.",
            text_color="red"
        ).pack(pady=20)
        return

    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT DISTINCT estado_producto
            FROM rma_detalles
            WHERE estado_producto IS NOT NULL AND estado_producto != ''
            ORDER BY estado_producto ASC
        """)
        estados_articulos = ["Todos"] + [fila[0] for fila in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT resultado_expediente
            FROM rma_maestro
            WHERE resultado_expediente IS NOT NULL AND resultado_expediente != ''
            ORDER BY resultado_expediente ASC
        """)
        resultados_expedientes = ["Todos"] + [fila[0] for fila in cursor.fetchall()]

    except Exception as e:
        print(f"Error al obtener valores de filtros: {e}")
        estados_articulos = ["Todos"]
        resultados_expedientes = ["Todos"]

    # Frame de filtros
    filtros_frame = ctk.CTkFrame(ventana_principal.main_stats_frame)
    filtros_frame.pack(fill="x", padx=20, pady=10)

    estados_seleccionados = []

    ctk.CTkLabel(filtros_frame, text="Resultado del Expediente:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
    filtro_resultado = ctk.CTkOptionMenu(filtros_frame, values=resultados_expedientes)
    filtro_resultado.set("Todos")
    filtro_resultado.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

    ctk.CTkLabel(filtros_frame, text="Estado del Artículo:").grid(row=0, column=2, padx=10, pady=5, sticky="w")
    btn_seleccionar_estados = ctk.CTkButton(
        filtros_frame,
        text="Seleccionar Estados (Todos)",
        command=lambda: abrir_selector_estados(
            estados_articulos, estados_seleccionados, btn_seleccionar_estados, cargar_datos_wrapper
        )
    )
    btn_seleccionar_estados.grid(row=0, column=3, padx=10, pady=5, sticky="ew")

    ctk.CTkLabel(filtros_frame, text="Fecha Desde:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
    entry_fecha_desde = ctk.CTkEntry(filtros_frame, placeholder_text="DD/MM/AAAA")
    entry_fecha_desde.grid(row=1, column=1, padx=10, pady=5, sticky="ew")

    ctk.CTkLabel(filtros_frame, text="Fecha Hasta:").grid(row=1, column=2, padx=10, pady=5, sticky="w")
    entry_fecha_hasta = ctk.CTkEntry(filtros_frame, placeholder_text="DD/MM/AAAA")
    entry_fecha_hasta.grid(row=1, column=3, padx=10, pady=5, sticky="ew")

    ctk.CTkLabel(filtros_frame, text="Ordenar por:").grid(row=1, column=4, padx=10, pady=5, sticky="w")
    filtro_orden = ctk.CTkOptionMenu(
        filtros_frame,
        values=["Referencia", "Cantidad Total ↓", "Cantidad Total ↑", "Coste Total ↓", "Coste Total ↑"]
    )
    filtro_orden.set("Cantidad Total ↓")
    filtro_orden.grid(row=1, column=5, padx=10, pady=5, sticky="ew")

    filtros_frame.grid_columnconfigure(1, weight=1)
    filtros_frame.grid_columnconfigure(3, weight=1)
    filtros_frame.grid_columnconfigure(5, weight=1)

    resultados_container = ctk.CTkFrame(ventana_principal.main_stats_frame)
    resultados_container.pack(fill="both", expand=True, padx=20, pady=10)

    total_frame = ctk.CTkFrame(ventana_principal.main_stats_frame)
    total_frame.pack(fill="x", padx=20, pady=10)

    lbl_total = ctk.CTkLabel(
        total_frame,
        text="TOTAL: 0.00 €",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#22c55e"
    )
    lbl_total.pack(pady=10)

    def mostrar_detalle_referencia(base_ref, detalles):
        """Abre un popup con el desglose por referencia individual y estado."""
        win = ctk.CTkToplevel(ventana_principal.main_stats_frame.winfo_toplevel())
        win.title(f"Desglose: {base_ref}")
        win.geometry("920x520")
        win.transient(ventana_principal.main_stats_frame.winfo_toplevel())
        win.grab_set()

        win.update_idletasks()
        x = (win.winfo_screenwidth() - 920) // 2
        y = (win.winfo_screenheight() - 520) // 2
        win.geometry(f"920x520+{x}+{y}")

        ctk.CTkLabel(
            win,
            text=f"Desglose: {base_ref}",
            font=ctk.CTkFont(size=15, weight="bold")
        ).pack(pady=12)

        scroll = ctk.CTkScrollableFrame(win)
        scroll.pack(fill="both", expand=True, padx=12, pady=5)

        header_font = ctk.CTkFont(weight="bold", size=12)
        cols_det = ["REFERENCIA", "ESTADO", "CANTIDAD", "PRECIO UNIT.", "COSTE", "Nº EXPED."]
        for col, h in enumerate(cols_det):
            ctk.CTkLabel(scroll, text=h, font=header_font).grid(
                row=0, column=col, padx=10, pady=8, sticky="w"
            )

        sorted_detalles = sorted(detalles, key=lambda d: (d['referencia'], d['estado']))

        total_cant = 0.0
        total_coste = 0.0

        for i, det in enumerate(sorted_detalles, start=1):
            ctk.CTkLabel(scroll, text=det['referencia']).grid(row=i, column=0, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(scroll, text=det['estado']).grid(row=i, column=1, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(scroll, text=f"{det['cantidad']:.0f}",
                         font=ctk.CTkFont(weight="bold")).grid(row=i, column=2, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(scroll, text=f"{det['precio']:.2f} €").grid(row=i, column=3, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(scroll, text=f"{det['coste']:.2f} €",
                         text_color="#2563eb").grid(row=i, column=4, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(scroll, text=str(det['num_expedientes'])).grid(
                row=i, column=5, padx=10, pady=4, sticky="w"
            )
            total_cant += det['cantidad']
            total_coste += det['coste']

        fila_total = len(sorted_detalles) + 1
        bold_green = ctk.CTkFont(weight="bold", size=12)
        ctk.CTkLabel(scroll, text="TOTAL", font=bold_green, text_color="#22c55e").grid(
            row=fila_total, column=0, padx=10, pady=8, sticky="w"
        )
        ctk.CTkLabel(scroll, text=f"{total_cant:.0f}", font=bold_green).grid(
            row=fila_total, column=2, padx=10, pady=8, sticky="w"
        )
        ctk.CTkLabel(scroll, text=f"{total_coste:.2f} €", font=bold_green, text_color="#22c55e").grid(
            row=fila_total, column=4, padx=10, pady=8, sticky="w"
        )

        ctk.CTkButton(win, text="Cerrar", command=win.destroy, width=100).pack(pady=10)

    def cargar_datos():
        """Carga los datos, agrupa por referencia base y suma todos los estados."""
        from datetime import datetime

        for widget in resultados_container.winfo_children():
            widget.destroy()

        resultado_seleccionado = filtro_resultado.get()
        fecha_desde = entry_fecha_desde.get().strip()
        fecha_hasta = entry_fecha_hasta.get().strip()
        orden_seleccionado = filtro_orden.get()

        # Agrupamos por referencia y estado en SQL; la agrupación por referencia base la hacemos en Python
        query = """
            SELECT
                d.referencia_articulo,
                d.estado_producto,
                SUM(d.cantidad_entregada) AS cantidad_total,
                AVG(COALESCE(d.precio_final, d.precio_unitario)) AS precio_promedio,
                COUNT(DISTINCT m.codigo_rma) AS num_expedientes
            FROM rma_detalles d
            INNER JOIN rma_maestro m ON d.rma_id = m.id
            WHERE COALESCE(d.contabilizar, 1) = 1
        """

        params = []

        if resultado_seleccionado != "Todos":
            query += " AND m.resultado_expediente = ?"
            params.append(resultado_seleccionado)

        if estados_seleccionados and "Todos" not in estados_seleccionados:
            placeholders = ",".join(["?" for _ in estados_seleccionados])
            query += f" AND d.estado_producto IN ({placeholders})"
            params.extend(estados_seleccionados)

        if fecha_desde:
            try:
                fecha_obj = datetime.strptime(fecha_desde, "%d/%m/%Y")
                query += " AND m.fecha_emision >= ?"
                params.append(fecha_obj.strftime("%Y-%m-%d"))
            except ValueError:
                messagebox.showwarning("Fecha inválida", "El formato de 'Fecha Desde' debe ser DD/MM/AAAA")
                return

        if fecha_hasta:
            try:
                fecha_obj = datetime.strptime(fecha_hasta, "%d/%m/%Y")
                query += " AND m.fecha_emision <= ?"
                params.append(fecha_obj.strftime("%Y-%m-%d"))
            except ValueError:
                messagebox.showwarning("Fecha inválida", "El formato de 'Fecha Hasta' debe ser DD/MM/AAAA")
                return

        query += " GROUP BY d.referencia_articulo, d.estado_producto ORDER BY d.referencia_articulo ASC"

        try:
            cursor.execute(query, tuple(params))
            registros = cursor.fetchall()

            if not registros:
                ctk.CTkLabel(
                    resultados_container,
                    text="No se encontraron artículos con los filtros aplicados.",
                    text_color="gray"
                ).pack(pady=20)
                lbl_total.configure(text="TOTAL: 0.00 €")
                cargar_datos.ultimos_grupos = []
                return

            # Agrupar en Python por referencia base
            grupos = defaultdict(lambda: {'cantidad_total': 0.0, 'coste_total': 0.0, 'detalles': []})

            for ref, estado, cant, precio, num_exp in registros:
                base = _extraer_referencia_base(ref)
                cant_f = float(cant or 0)
                precio_f = float(precio or 0)
                coste_f = cant_f * precio_f

                grupos[base]['cantidad_total'] += cant_f
                grupos[base]['coste_total'] += coste_f
                grupos[base]['detalles'].append({
                    'referencia': ref or 'N/A',
                    'estado': estado or 'N/A',
                    'cantidad': cant_f,
                    'precio': precio_f,
                    'coste': coste_f,
                    'num_expedientes': int(num_exp or 0)
                })

            # Ordenar grupos
            if orden_seleccionado == "Referencia":
                sorted_grupos = sorted(grupos.items(), key=lambda x: x[0])
            elif orden_seleccionado == "Cantidad Total ↓":
                sorted_grupos = sorted(grupos.items(), key=lambda x: -x[1]['cantidad_total'])
            elif orden_seleccionado == "Cantidad Total ↑":
                sorted_grupos = sorted(grupos.items(), key=lambda x: x[1]['cantidad_total'])
            elif orden_seleccionado == "Coste Total ↓":
                sorted_grupos = sorted(grupos.items(), key=lambda x: -x[1]['coste_total'])
            elif orden_seleccionado == "Coste Total ↑":
                sorted_grupos = sorted(grupos.items(), key=lambda x: x[1]['coste_total'])
            else:
                sorted_grupos = sorted(grupos.items(), key=lambda x: -x[1]['cantidad_total'])

            # Frame scrollable
            scroll_frame = ctk.CTkScrollableFrame(resultados_container)
            scroll_frame.pack(fill="both", expand=True)

            # Encabezados
            header_font = ctk.CTkFont(weight="bold", size=12)
            headers_base = ["REFERENCIA BASE", "CANT. TOTAL", "PRECIO PROM.", "COSTE TOTAL", "VARIANTES / ESTADOS"]

            headers = []
            for col, h in enumerate(headers_base):
                if col == 1:
                    if orden_seleccionado == "Cantidad Total ↓":
                        headers.append(f"{h} ▼")
                    elif orden_seleccionado == "Cantidad Total ↑":
                        headers.append(f"{h} ▲")
                    else:
                        headers.append(h)
                elif col == 3:
                    if orden_seleccionado == "Coste Total ↓":
                        headers.append(f"{h} ▼")
                    elif orden_seleccionado == "Coste Total ↑":
                        headers.append(f"{h} ▲")
                    else:
                        headers.append(h)
                elif col == 0 and orden_seleccionado == "Referencia":
                    headers.append(f"{h} ▲")
                else:
                    headers.append(h)

            for col, h in enumerate(headers):
                ctk.CTkLabel(scroll_frame, text=h, font=header_font).grid(
                    row=0, column=col, padx=10, pady=10, sticky="w"
                )

            suma_total = 0.0
            grupos_para_export = []

            for i, (base_ref, datos) in enumerate(sorted_grupos, start=1):
                precio_prom = datos['coste_total'] / datos['cantidad_total'] if datos['cantidad_total'] > 0 else 0
                num_variantes = len(set(d['referencia'] for d in datos['detalles']))
                num_estados = len(set(d['estado'] for d in datos['detalles']))
                variantes_text = f"{num_variantes} ref. / {num_estados} estados"

                suma_total += datos['coste_total']

                # Referencia base como botón clickable (muestra el desglose)
                btn_ref = ctk.CTkButton(
                    scroll_frame,
                    text=base_ref if base_ref else "N/A",
                    command=lambda br=base_ref, d=datos['detalles']: mostrar_detalle_referencia(br, d),
                    fg_color="transparent",
                    text_color=("#1d4ed8", "#60a5fa"),
                    hover_color=("#dbeafe", "#1e3a5f"),
                    anchor="w",
                )
                btn_ref.grid(row=i, column=0, padx=10, pady=3, sticky="w")

                ctk.CTkLabel(
                    scroll_frame,
                    text=f"{datos['cantidad_total']:.0f}",
                    font=ctk.CTkFont(weight="bold")
                ).grid(row=i, column=1, padx=10, pady=3, sticky="w")

                ctk.CTkLabel(scroll_frame, text=f"{precio_prom:.2f} €").grid(
                    row=i, column=2, padx=10, pady=3, sticky="w"
                )

                ctk.CTkLabel(
                    scroll_frame,
                    text=f"{datos['coste_total']:.2f} €",
                    text_color="#2563eb",
                    font=ctk.CTkFont(weight="bold")
                ).grid(row=i, column=3, padx=10, pady=3, sticky="w")

                ctk.CTkLabel(scroll_frame, text=variantes_text, text_color="gray").grid(
                    row=i, column=4, padx=10, pady=3, sticky="w"
                )

                grupos_para_export.append({
                    'base_ref': base_ref,
                    'cantidad_total': datos['cantidad_total'],
                    'precio_prom': precio_prom,
                    'coste_total': datos['coste_total'],
                    'num_variantes': num_variantes,
                    'num_estados': num_estados,
                    'detalles': datos['detalles']
                })

            lbl_total.configure(text=f"TOTAL: {suma_total:,.2f} €")
            cargar_datos.ultimos_grupos = grupos_para_export

        except Exception as e:
            print(f"Error al cargar datos de artículos: {e}")
            messagebox.showerror("Error", f"Error al cargar los datos: {e}")
            cargar_datos.ultimos_grupos = []

    cargar_datos.ultimos_grupos = []

    def cargar_datos_wrapper():
        cargar_datos()

    def exportar_a_excel():
        """Exporta los resultados a Excel: hoja resumen y hoja con desglose completo."""
        if not cargar_datos.ultimos_grupos:
            messagebox.showwarning("Sin datos", "No hay datos para exportar. Primero aplique los filtros.")
            return

        try:
            # Hoja 1: resumen por referencia base
            datos_resumen = []
            datos_detalle = []

            for g in cargar_datos.ultimos_grupos:
                datos_resumen.append({
                    'Referencia Base': g['base_ref'],
                    'Cantidad Total': g['cantidad_total'],
                    'Precio Promedio (€)': round(g['precio_prom'], 2),
                    'Coste Total (€)': round(g['coste_total'], 2),
                    'Nº Variantes': g['num_variantes'],
                    'Nº Estados': g['num_estados'],
                })
                for det in sorted(g['detalles'], key=lambda d: (d['referencia'], d['estado'])):
                    datos_detalle.append({
                        'Referencia Base': g['base_ref'],
                        'Referencia': det['referencia'],
                        'Estado': det['estado'],
                        'Cantidad': det['cantidad'],
                        'Precio Unitario (€)': round(det['precio'], 2),
                        'Coste (€)': round(det['coste'], 2),
                        'Nº Expedientes': det['num_expedientes'],
                    })

            archivo = filedialog.asksaveasfilename(
                title="Guardar estadísticas de artículos",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                initialfile="estadisticas_articulos.xlsx"
            )
            if not archivo:
                return

            from openpyxl.styles import Font, PatternFill, Alignment

            with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
                df_resumen = pd.DataFrame(datos_resumen)
                df_detalle = pd.DataFrame(datos_detalle)

                df_resumen.to_excel(writer, index=False, sheet_name='Resumen')
                df_detalle.to_excel(writer, index=False, sheet_name='Desglose')

                for sheet_name in ['Resumen', 'Desglose']:
                    ws = writer.sheets[sheet_name]
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    for col in ws.columns:
                        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            messagebox.showinfo("Éxito", f"Datos exportados correctamente a:\n{archivo}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{e}")
            print(f"Error en exportación: {e}")

    # Botones de acción
    btn_limpiar = ctk.CTkButton(
        filtros_frame,
        text="🗑️ Limpiar Fechas",
        command=lambda: (entry_fecha_desde.delete(0, 'end'), entry_fecha_hasta.delete(0, 'end'), cargar_datos()),
        width=120
    )
    btn_limpiar.grid(row=2, column=0, padx=10, pady=10)

    btn_aplicar = ctk.CTkButton(
        filtros_frame,
        text="🔍 Aplicar Filtros",
        command=cargar_datos,
        width=140
    )
    btn_aplicar.grid(row=2, column=1, padx=10, pady=10)

    btn_exportar = ctk.CTkButton(
        filtros_frame,
        text="💾 Exportar a Excel",
        command=exportar_a_excel,
        width=140
    )
    btn_exportar.grid(row=2, column=2, columnspan=2, padx=10, pady=10)

    cargar_datos()

    conn.close()


def abrir_selector_estados(estados_disponibles, estados_seleccionados, boton, callback_actualizar):
    """Abre una ventana para seleccionar múltiples estados."""

    ventana_selector = Toplevel()
    ventana_selector.title("Seleccionar Estados de Artículo")
    ventana_selector.geometry("500x400")
    ventana_selector.transient()
    ventana_selector.grab_set()

    ventana_selector.update_idletasks()
    x = (ventana_selector.winfo_screenwidth() - 500) // 2
    y = (ventana_selector.winfo_screenheight() - 400) // 2
    ventana_selector.geometry(f"500x400+{x}+{y}")

    frame_principal = ctk.CTkFrame(ventana_selector)
    frame_principal.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        frame_principal,
        text="Seleccione los estados a filtrar (múltiple):",
        font=ctk.CTkFont(size=12, weight="bold")
    ).pack(pady=10)

    frame_lista = ctk.CTkFrame(frame_principal)
    frame_lista.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = Scrollbar(frame_lista)
    scrollbar.pack(side="right", fill="y")

    listbox = Listbox(frame_lista, selectmode=MULTIPLE, yscrollcommand=scrollbar.set, height=15)
    listbox.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=listbox.yview)

    for estado in estados_disponibles:
        listbox.insert(END, estado)
        if estado in estados_seleccionados or (not estados_seleccionados and estado == "Todos"):
            listbox.selection_set(estados_disponibles.index(estado))

    frame_botones = ctk.CTkFrame(frame_principal)
    frame_botones.pack(fill="x", pady=10)

    def seleccionar_todos():
        listbox.selection_set(0, END)

    def deseleccionar_todos():
        listbox.selection_clear(0, END)

    def aplicar_seleccion():
        indices = listbox.curselection()
        estados_seleccionados.clear()

        if not indices:
            estados_seleccionados.append("Todos")
        else:
            for i in indices:
                estados_seleccionados.append(listbox.get(i))

        if "Todos" in estados_seleccionados or not estados_seleccionados:
            boton.configure(text="Seleccionar Estados (Todos)")
        elif len(estados_seleccionados) == 1:
            texto = estados_seleccionados[0]
            if len(texto) > 25:
                texto = texto[:22] + "..."
            boton.configure(text=f"Estados: {texto}")
        else:
            boton.configure(text=f"Estados seleccionados: {len(estados_seleccionados)}")

        callback_actualizar()
        ventana_selector.destroy()

    ctk.CTkButton(frame_botones, text="✓ Seleccionar Todos", command=seleccionar_todos, width=140).pack(side="left", padx=5)
    ctk.CTkButton(frame_botones, text="✗ Deseleccionar Todos", command=deseleccionar_todos, width=140).pack(side="left", padx=5)
    ctk.CTkButton(frame_botones, text="Aplicar", command=aplicar_seleccion, width=100).pack(side="right", padx=5)
    ctk.CTkButton(
        frame_botones, text="Cancelar", command=ventana_selector.destroy,
        fg_color="#ef4444", width=100
    ).pack(side="right", padx=5)


