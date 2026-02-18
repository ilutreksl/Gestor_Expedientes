import customtkinter as ctk
import tkinter.messagebox as messagebox
import tkinter as tk
import tkinter.filedialog as filedialog
import sqlite3
import locale
import datetime
from lib.logger_config import get_logger

logger = get_logger()

def mostrar_rentabilidad_clientes(app):
    """Módulo externo que dibuja la estadística de rentabilidad por cliente.

    Parámetros:
    - app: instancia de VentanaPrincipal (para usar su frame `main_stats_frame` y `master.conectar_db()`)
    """
    # Limpiar marco principal
    try:
        app.limpiar_marco_stats()
    except Exception:
        pass

    frame = app.main_stats_frame
    if not frame:
        return

    ctk.CTkLabel(frame, text="RENTABILIDAD POR CLIENTE", font=ctk.CTkFont(size=18, weight="bold")).pack(pady=16)

    # Controles
    controles = ctk.CTkFrame(frame)
    controles.pack(fill="x", padx=16, pady=(0,10))
    controles.grid_columnconfigure((0,2,4), weight=0)
    controles.grid_columnconfigure((1,3,5), weight=1)

    # Resultado (No abonar / Abonar / Reposición)
    opciones_resultado = ["Todos"] + (getattr(app, 'OPCIONES', {}).get('Resultado_Expediente', []) if hasattr(app, 'OPCIONES') else [])
    ctk.CTkLabel(controles, text="Resultado:").grid(row=0, column=0, padx=6, pady=6, sticky="w")
    resultado_opt = ctk.CTkOptionMenu(controles, values=opciones_resultado)
    resultado_opt.grid(row=0, column=1, padx=6, pady=6, sticky="ew")
    resultado_opt.set(opciones_resultado[0])

    # Año
    ctk.CTkLabel(controles, text="Año:").grid(row=0, column=2, padx=6, pady=6, sticky="w")
    # Obtener años desde la BD — intento robusto: primero `strftime('%Y', ...)`,
    # si no devuelve resultados (formato dd/mm/yyyy), uso SUBSTR para extraer el año.
    años = ["Todos"]
    try:
        conn, cursor = app.master.conectar_db()
        if conn:
            rows = []
            try:
                cursor.execute("SELECT DISTINCT strftime('%Y', fecha_gestion) AS anio FROM rma_maestro WHERE fecha_gestion IS NOT NULL ORDER BY anio DESC")
                rows = [r[0] for r in cursor.fetchall() if r and r[0]]
            except Exception:
                rows = []

            if not rows:
                try:
                    cursor.execute("SELECT DISTINCT SUBSTR(fecha_gestion, -4, 4) AS anio FROM rma_maestro WHERE fecha_gestion IS NOT NULL ORDER BY anio DESC")
                    rows = [r[0] for r in cursor.fetchall() if r and r[0]]
                except Exception:
                    rows = []

            for anio in rows:
                años.append(anio)
            conn.close()
    except Exception:
        pass

    año_opt = ctk.CTkOptionMenu(controles, values=años)
    año_opt.grid(row=0, column=3, padx=6, pady=6, sticky="ew")
    año_opt.set(años[0] if años else "Todos")

    # Periodo (Trimestre / Semestre)
    ctk.CTkLabel(controles, text="Periodo: ").grid(row=1, column=0, padx=6, pady=6, sticky="w")
    periodo_tipo = ctk.CTkOptionMenu(controles, values=["Todos", "Trimestre", "Semestre"])
    periodo_tipo.grid(row=1, column=1, padx=6, pady=6, sticky="ew")
    periodo_tipo.set("Todos")

    periodo_valor = ctk.CTkOptionMenu(controles, values=["N/A"]) 
    periodo_valor.grid(row=1, column=3, padx=6, pady=6, sticky="ew")
    periodo_valor.set("N/A")

    def actualizar_valores_periodo(choice=None):
        t = periodo_tipo.get()
        if t == 'Trimestre':
            periodo_valor.configure(values=["Q1", "Q2", "Q3", "Q4"]) 
            periodo_valor.set("Q1")
        elif t == 'Semestre':
            periodo_valor.configure(values=["H1", "H2"]) 
            periodo_valor.set("H1")
        else:
            periodo_valor.configure(values=["N/A"]) 
            periodo_valor.set("N/A")

    # configurar el callback ahora que la función existe y asegurar valores iniciales
    try:
        periodo_tipo.configure(command=actualizar_valores_periodo)
    except Exception:
        # Si CTkOptionMenu no soporta configure de command, ignoramos
        pass
    actualizar_valores_periodo()

    # Filtro avanzado por cliente (texto)
    ctk.CTkLabel(controles, text="Filtro Cliente: ").grid(row=2, column=0, padx=6, pady=6, sticky="w")
    cliente_entry = ctk.CTkEntry(controles, placeholder_text="Parte del nombre del cliente...")
    cliente_entry.grid(row=2, column=1, columnspan=3, sticky="ew", padx=6, pady=6)

    # Botón aplicar filtros
    btn_frame = ctk.CTkFrame(controles)
    btn_frame.grid(row=3, column=0, columnspan=6, pady=8, sticky="ew")
    btn_frame.grid_columnconfigure(0, weight=1)

    resultados_frame = ctk.CTkFrame(frame)
    resultados_frame.pack(fill="both", expand=True, padx=16, pady=(0,10))

    def construir_where_and_params():
        where = ["1=1"]
        params = []
        
        logger.debug("=== Construyendo filtros de rentabilidad ===")
        
        # Resultado
        res = resultado_opt.get()
        logger.debug(f"Filtro Resultado: '{res}'")
        if res and res != 'Todos':
            where.append("lower(resultado_expediente) = ?")
            params.append(res.strip().lower())
            logger.debug(f"  -> Aplicado filtro resultado: {res.strip().lower()}")
        
        # Año (fecha_gestion en formato ISO: YYYY-MM-DD)
        anio = año_opt.get()
        logger.debug(f"Filtro Año: '{anio}'")
        if anio and anio != 'Todos' and anio.strip():
            where.append("fecha_gestion IS NOT NULL AND fecha_gestion != '' AND SUBSTR(fecha_gestion, 1, 4) = ?")
            params.append(anio.strip())
            logger.debug(f"  -> Aplicado filtro año: {anio.strip()}")
        
        # Periodo
        per_tipo = periodo_tipo.get()
        per_val = periodo_valor.get()
        logger.debug(f"Filtro Periodo: tipo='{per_tipo}', valor='{per_val}'")
        
        if per_tipo == 'Trimestre' and per_val in ('Q1','Q2','Q3','Q4'):
            mapping = {'Q1':('01','03'),'Q2':('04','06'),'Q3':('07','09'),'Q4':('10','12')}
            m1,m2 = mapping[per_val]
            where.append("fecha_gestion IS NOT NULL AND fecha_gestion != '' AND SUBSTR(fecha_gestion, 6, 2) BETWEEN ? AND ?")
            params.extend([m1,m2])
            logger.debug(f"  -> Aplicado filtro trimestre {per_val}: meses {m1}-{m2}")
        elif per_tipo == 'Semestre' and per_val in ('H1','H2'):
            mapping = {'H1':('01','06'),'H2':('07','12')}
            m1,m2 = mapping[per_val]
            where.append("fecha_gestion IS NOT NULL AND fecha_gestion != '' AND SUBSTR(fecha_gestion, 6, 2) BETWEEN ? AND ?")
            params.extend([m1,m2])
            logger.debug(f"  -> Aplicado filtro semestre {per_val}: meses {m1}-{m2}")
        
        # Cliente
        cliente_like = cliente_entry.get().strip()
        logger.debug(f"Filtro Cliente (texto): '{cliente_like}'")
        if cliente_like:
            where.append("cliente LIKE ?")
            params.append(f"%{cliente_like}%")
            logger.debug(f"  -> Aplicado filtro cliente con LIKE: %{cliente_like}%")
        
        where_final = " AND ".join(where)
        logger.debug(f"WHERE final: {where_final}")
        logger.debug(f"Parámetros: {params}")
        
        return where_final, params

    # caché de resultados para exportar
    results_cache = []

    def cargar_datos():
        logger.info("Cargando datos de rentabilidad por cliente")
        for w in resultados_frame.winfo_children():
            w.destroy()

        conn, cursor = app.master.conectar_db()
        if not conn:
            logger.error("No se pudo conectar a la base de datos")
            messagebox.showerror("Error", "No se pudo conectar a la base de datos.")
            return

        where_clause, params = construir_where_and_params()
        # Subconsulta que calcula el total por expediente a partir de rma_detalles (solo contabilizables)
        sub_calc = (
            "(SELECT rma_id, SUM(COALESCE(cantidad_entregada,0) * COALESCE(precio_final, precio_unitario,0)) as calc_total "
            "FROM rma_detalles WHERE COALESCE(contabilizar, 1) = 1 GROUP BY rma_id) AS calc"
        )

        # Seleccionamos por cliente: contamos expedientes y sumamos el valor definitivo,
        # usando precio_total_expediente si existe, sino el calculado a partir de detalles.
        sql = f"""
            SELECT m.cliente,
                   COUNT(m.id) AS total_expedientes,
                   SUM(COALESCE(m.precio_total_expediente, calc.calc_total, 0)) AS suma_total
            FROM rma_maestro m
            LEFT JOIN {sub_calc} ON calc.rma_id = m.id
            WHERE {where_clause}
            GROUP BY m.cliente
            ORDER BY suma_total DESC
            LIMIT 200
        """
        
        logger.debug(f"SQL ejecutada: {sql}")
        logger.debug(f"Parámetros: {params}")
        
        try:
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            logger.info(f"Se encontraron {len(rows)} clientes con los filtros aplicados")

            # Guardar resultados en caché para export
            try:
                results_cache.clear()
            except Exception:
                pass
            try:
                results_cache.extend(rows)
            except Exception:
                # en caso de que results_cache no sea mutable aún
                pass

            # Cuerpo con scrollbar — colocamos el encabezado y filas dentro de
            # un contenedor que usa una única grid, de forma que todas las filas
            # compartan las mismas columnas y sus anchuras.
            body = ctk.CTkScrollableFrame(resultados_frame, label_text='Resultados')
            body.pack(fill='both', expand=True, padx=8, pady=8)

            table = ctk.CTkFrame(body)
            table.pack(fill='both', expand=True, padx=4, pady=4)
            # Columnas: cliente (más ancha), total_expedientes, suma_total
            table.grid_columnconfigure(0, weight=4)
            table.grid_columnconfigure(1, weight=1)
            table.grid_columnconfigure(2, weight=1)

            # Encabezados — en la primera fila de la tabla
            ctk.CTkLabel(table, text='CLIENTE', font=ctk.CTkFont(weight='bold')).grid(row=0, column=0, sticky='w', padx=8, pady=4)
            ctk.CTkLabel(table, text='TOTAL EXPEDIENTES', font=ctk.CTkFont(weight='bold')).grid(row=0, column=1, sticky='e', padx=8, pady=4)
            ctk.CTkLabel(table, text='SUMA TOTAL (€)', font=ctk.CTkFont(weight='bold')).grid(row=0, column=2, sticky='e', padx=8, pady=4)

            try:
                for idx, r in enumerate(rows, start=1):
                    cliente = r[0] or '(Sin cliente)'
                    tot = r[1] or 0
                    suma = r[2] or 0.0

                    # Colocar directamente los widgets en la fila `idx` de la tabla
                    lbl = ctk.CTkLabel(table, text=cliente)
                    lbl.grid(row=idx, column=0, sticky='w', padx=8, pady=3)

                    lbl_tot = ctk.CTkLabel(table, text=str(tot))
                    lbl_tot.grid(row=idx, column=1, sticky='e', padx=8, pady=3)

                    try:
                        suma_text = locale.currency(float(suma), grouping=True, symbol=True)
                    except Exception:
                        suma_text = f"{float(suma):,.2f} €"
                    lbl_sum = ctk.CTkLabel(table, text=suma_text)
                    lbl_sum.grid(row=idx, column=2, sticky='e', padx=8, pady=3)

                    # Click para abrir ventana de detalle del cliente
                    def make_open(c_name):
                        def open_client(evt=None):
                            abrir_ventana_cliente(app, c_name, where_clause, params)
                        return open_client

                    lbl.configure(cursor='hand2')
                    lbl.bind('<Button-1>', make_open(cliente))
            except Exception as e:
                # Mostrar detalle del error para depuración en tiempo de ejecución
                logger.error(f"Error mostrando tabla de resultados: {e}", exc_info=True)
                try:
                    import traceback
                    tb = traceback.format_exc()
                    messagebox.showerror('Error', f'No se pudo mostrar la tabla: {e}\n\n{tb}')
                except Exception:
                    messagebox.showerror('Error', f'No se pudo mostrar la tabla: {e}')
                return

        except sqlite3.Error as e:
            logger.error(f"Error de BD al consultar rentabilidad: {e}", exc_info=True)
            messagebox.showerror('Error BD', f'Error al consultar: {e}')
        finally:
            conn.close()

    def abrir_ventana_cliente(app, cliente_nombre, where_clause, params):
        # Abrir nueva ventana con listados de expedientes del cliente
        logger.info(f"Abriendo ventana de detalle para cliente: {cliente_nombre}")
        win = ctk.CTkToplevel(app)
        win.title(f"Expedientes - {cliente_nombre}")
        win.geometry('900x600')

        cont = ctk.CTkFrame(win)
        cont.pack(fill='both', expand=True, padx=8, pady=8)

        # Cabecera
        ctk.CTkLabel(cont, text=f"Expedientes de: {cliente_nombre}", font=ctk.CTkFont(size=14, weight='bold')).pack(pady=6)

        tabla = ctk.CTkScrollableFrame(cont)
        tabla.pack(fill='both', expand=True, padx=6, pady=6)

        # Re-consultar expedientes filtrando por cliente y respetando otros filtros
        conn, cursor = app.master.conectar_db()
        if not conn:
            logger.error("No se pudo conectar a BD para detalles de cliente")
            messagebox.showerror('Error', 'No se pudo conectar a la base de datos.')
            win.destroy()
            return

        # Construimos WHERE similar pero forzando cliente
        where = [where_clause] if where_clause else ["1=1"]
        params_local = list(params)
        where.append("m.cliente = ?")
        params_local.append(cliente_nombre)

        # Calcular total contabilizable para cada expediente
        sql = f"""
            SELECT 
                m.id, 
                m.codigo_rma, 
                COALESCE(
                    (SELECT SUM(d.precio_final * d.cantidad_entregada)
                     FROM rma_detalles d
                     WHERE d.rma_id = m.id AND COALESCE(d.contabilizar, 1) = 1),
                    m.precio_total_expediente,
                    0
                ) as total_contabilizable,
                m.resultado_expediente, 
                m.fecha_gestion 
            FROM rma_maestro m
            WHERE {' AND '.join(where)} 
            ORDER BY m.fecha_gestion DESC
        """
        
        logger.debug(f"Consultando expedientes de {cliente_nombre}: {sql}")
        logger.debug(f"Parámetros: {params_local}")
        
        try:
            cursor.execute(sql, params_local)
            expedientes = cursor.fetchall()
            
            logger.info(f"Se encontraron {len(expedientes)} expedientes para {cliente_nombre}")

            # Encabezados
            hdr = ctk.CTkFrame(tabla)
            hdr.pack(fill='x')
            hdr.grid_columnconfigure(0, weight=2)
            hdr.grid_columnconfigure(1, weight=1)
            hdr.grid_columnconfigure(2, weight=1)
            hdr.grid_columnconfigure(3, weight=1)
            hdr.grid_columnconfigure(4, weight=0)
            ctk.CTkLabel(hdr, text='CÓDIGO RMA', font=ctk.CTkFont(weight='bold')).grid(row=0, column=0, sticky='w', padx=6)
            ctk.CTkLabel(hdr, text='IMPORTE (€)', font=ctk.CTkFont(weight='bold')).grid(row=0, column=1, sticky='e', padx=6)
            ctk.CTkLabel(hdr, text='RESULTADO', font=ctk.CTkFont(weight='bold')).grid(row=0, column=2, sticky='e', padx=6)
            ctk.CTkLabel(hdr, text='FECHA', font=ctk.CTkFont(weight='bold')).grid(row=0, column=3, sticky='e', padx=6)
            ctk.CTkLabel(hdr, text='', font=ctk.CTkFont(weight='bold')).grid(row=0, column=4, sticky='e', padx=6)

            for e in expedientes:
                rid, code, precio, resultado, fecha = e
                rowf = ctk.CTkFrame(tabla)
                rowf.grid_columnconfigure(0, weight=2)
                rowf.grid_columnconfigure(1, weight=1)
                rowf.grid_columnconfigure(2, weight=1)
                rowf.grid_columnconfigure(3, weight=1)
                rowf.grid_columnconfigure(4, weight=0)
                rowf.pack(fill='x', padx=6, pady=3)

                lbl_code = ctk.CTkLabel(rowf, text=code)
                lbl_code.grid(row=0, column=0, sticky='w')

                try:
                    precio_text = locale.currency(float(precio), grouping=True, symbol=True)
                except Exception:
                    precio_text = f"{float(precio or 0):,.2f} €"
                ctk.CTkLabel(rowf, text=precio_text).grid(row=0, column=1, sticky='e')
                ctk.CTkLabel(rowf, text=str(resultado or '')).grid(row=0, column=2, sticky='e')
                ctk.CTkLabel(rowf, text=str(fecha or '')).grid(row=0, column=3, sticky='e')

                # Abrir expediente con doble click en la fila (en lugar de botón)
                def _on_double(evt=None, _rid=rid, _code=code):
                    logger.info(f"Abriendo expediente {_code} (ID: {_rid})")
                    try:
                        app.mostrar_nuevo_rma(_rid)
                        try:
                            messagebox.showinfo('Expediente abierto', f'El expediente {_code} se ha abierto en la ventana principal.')
                        except Exception:
                            # No crítico; si showinfo falla, seguimos
                            pass
                        win.destroy()
                    except Exception as e:
                        logger.error(f"Error abriendo expediente {_code}: {e}", exc_info=True)
                        messagebox.showerror('Error', f'No se pudo abrir el expediente: {e}')

                # Hacer que la fila y sus hijos reaccionen al doble clic
                rowf.configure(cursor='hand2')
                rowf.bind('<Double-1>', _on_double)
                # También enlazar a los widgets hijos (por si capturan el evento)
                for _w in rowf.winfo_children():
                    try:
                        _w.bind('<Double-1>', _on_double)
                    except Exception:
                        pass

        except sqlite3.Error as e:
            logger.error(f"Error BD consultando expedientes de cliente: {e}", exc_info=True)
            messagebox.showerror('Error BD', f'Error al consultar expedientes: {e}')
        finally:
            conn.close()

    # Botón aplicar
    ctk.CTkButton(btn_frame, text='Aplicar Filtros', command=cargar_datos).pack(side='left')

    # Exportar resultados a Excel/CSV
    def export_results():
        logger.info("Iniciando exportación de resultados de rentabilidad")
        if not results_cache:
            logger.warning("No hay datos en caché para exportar")
            messagebox.showinfo('Exportar', 'No hay datos para exportar. Aplica filtros primero.')
            return

        # Pedir nombre de fichero
        fpath = filedialog.asksaveasfilename(
            defaultextension='.xlsx', 
            filetypes=[('Excel Workbook', '*.xlsx'), ('CSV', '*.csv'), ('All files', '*.*')],
            title='Guardar exportación como...'
        )
        if not fpath:
            logger.info("Exportación cancelada por el usuario")
            return

        logger.info(f"Exportando {len(results_cache)} registros a: {fpath}")

        # Intentar usar openpyxl si está disponible
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            use_xlsx = True
            logger.debug("openpyxl disponible, exportando con formato")
        except Exception as e:
            logger.warning(f"openpyxl no disponible, exportando sin formato: {e}")
            use_xlsx = False

        headers = ['CLIENTE', 'TOTAL_EXPEDIENTES', 'SUMA_TOTAL_€']

        try:
            if use_xlsx and fpath.lower().endswith('.xlsx'):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Rentabilidad por Cliente'
                
                # Escribir encabezados
                ws.append(headers)
                
                # Formatear encabezados
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Escribir datos
                for r in results_cache:
                    ws.append([r[0], int(r[1] or 0), float(r[2] or 0.0)])
                
                # Formatear datos
                border_style = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                for row_num in range(2, len(results_cache) + 2):
                    # Cliente
                    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left")
                    ws.cell(row=row_num, column=1).border = border_style
                    
                    # Total expedientes
                    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_num, column=2).border = border_style
                    
                    # Suma total (formato moneda)
                    cell_suma = ws.cell(row=row_num, column=3)
                    cell_suma.number_format = '#,##0.00 €'
                    cell_suma.alignment = Alignment(horizontal="right")
                    cell_suma.border = border_style
                
                # Ajustar anchos de columna
                ws.column_dimensions['A'].width = 40  # Cliente
                ws.column_dimensions['B'].width = 20  # Total expedientes
                ws.column_dimensions['C'].width = 20  # Suma total
                
                # Congelar primera fila (encabezados)
                ws.freeze_panes = 'A2'
                
                wb.save(fpath)
                logger.info(f"Archivo Excel guardado exitosamente: {fpath}")
            else:
                # guardar CSV
                logger.debug("Exportando como CSV")
                import csv
                with open(fpath, 'w', newline='', encoding='utf-8') as csvf:
                    writer = csv.writer(csvf, delimiter=';')
                    writer.writerow(headers)
                    for r in results_cache:
                        writer.writerow([r[0], r[1] or 0, r[2] or 0.0])
                logger.info(f"Archivo CSV guardado exitosamente: {fpath}")

            messagebox.showinfo('Exportar', f'Exportación completada:\n{fpath}')
        except Exception as e:
            logger.error(f"Error al exportar: {e}", exc_info=True)
            messagebox.showerror('Exportar', f'Error al exportar: {e}')

    ctk.CTkButton(btn_frame, text='Exportar a Excel', command=export_results).pack(side='left', padx=8)

    # Cargar inicialmente
    cargar_datos()
