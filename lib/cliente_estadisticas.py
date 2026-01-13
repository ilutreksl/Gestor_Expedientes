"""
Módulo para calcular estadísticas de clientes
"""
import customtkinter as ctk
from datetime import datetime
from tkinter import messagebox
import pandas as pd


def calcular_estadisticas_basicas_cliente(nombre_cliente, conn):
    """
    Calcula las estadísticas básicas de un cliente.
    
    Args:
        nombre_cliente: Nombre del cliente
        conn: Conexión a la base de datos
    
    Returns:
        dict: Diccionario con estadísticas básicas
    """
    cursor = conn.cursor()
    
    # Consulta para obtener estadísticas
    cursor.execute("""
        SELECT 
            COUNT(DISTINCT m.id) as total_expedientes,
            SUM(COALESCE(d.cantidad_entregada, 0) * COALESCE(d.precio_final, d.precio_unitario, 0)) as coste_total,
            COUNT(DISTINCT CASE WHEN m.estado = 'Completado' THEN m.id END) as expedientes_completados,
            COUNT(DISTINCT CASE WHEN m.resultado_expediente LIKE '%ABONAR%' AND m.resultado_expediente NOT LIKE '%NO ABONAR%' THEN m.id END) as total_abonos
        FROM rma_maestro m
        LEFT JOIN rma_detalles d ON m.id = d.rma_id
        WHERE m.cliente = ?
    """, (nombre_cliente,))
    
    resultado = cursor.fetchone()
    
    if resultado:
        total_exp, coste_total, completados, abonos = resultado
        
        # Convertir a tipos numéricos correctos
        total_exp = int(total_exp) if total_exp else 0
        coste_total = float(coste_total) if coste_total else 0.0
        completados = int(completados) if completados else 0
        abonos = int(abonos) if abonos else 0
        
        # Calcular porcentajes
        porcentaje_completados = (completados / total_exp * 100) if total_exp > 0 else 0
        porcentaje_abonos = (abonos / total_exp * 100) if total_exp > 0 else 0
        
        return {
            'total_expedientes': total_exp,
            'coste_total': coste_total,
            'expedientes_completados': completados,
            'porcentaje_completados': porcentaje_completados,
            'total_abonos': abonos,
            'porcentaje_abonos': porcentaje_abonos
        }
    
    return {
        'total_expedientes': 0,
        'coste_total': 0.0,
        'expedientes_completados': 0,
        'porcentaje_completados': 0,
        'total_abonos': 0,
        'porcentaje_abonos': 0
    }


def obtener_estadisticas_detalladas_cliente(nombre_cliente, conn, fecha_desde=None, fecha_hasta=None, filtro_estado=None):
    """
    Obtiene estadísticas detalladas de un cliente con filtros.
    
    Args:
        nombre_cliente: Nombre del cliente
        conn: Conexión a la base de datos
        fecha_desde: Fecha inicio filtro (formato YYYY-MM-DD)
        fecha_hasta: Fecha fin filtro (formato YYYY-MM-DD)
        filtro_estado: Estado a filtrar (Todos, Completado, En trámite, etc.)
    
    Returns:
        list: Lista de tuplas con datos detallados por expediente
    """
    cursor = conn.cursor()
    
    # Estados de fallo para clasificar
    ESTADOS_FALLO = [
        'NO FUNCIONA, ABONAR',
        'NO FUNCIONA ; NO ABONAR',
        'REPOSICION FALLO PRODUCTO',
        'REPOSICION ; ABONAR',
        'FALLO SOLDADURA ; ABONAR',
        'FALLO SOLDADURA ; NO ABONAR',
        'FALLO MODULO ; ABONAR'
    ]
    
    # Construir query con subconsulta para detectar fallos
    query = """
        SELECT 
            m.codigo_rma,
            m.fecha_emision,
            m.estado,
            m.resultado_expediente,
            COUNT(d.id) as num_articulos,
            SUM(d.cantidad_entregada * COALESCE(d.precio_final, d.precio_unitario)) as coste_expediente,
            MAX(CASE 
                WHEN d.estado_producto IN ('NO FUNCIONA, ABONAR', 'NO FUNCIONA ; NO ABONAR', 
                                  'REPOSICION FALLO PRODUCTO', 'REPOSICION ; ABONAR',
                                  'FALLO SOLDADURA ; ABONAR', 'FALLO SOLDADURA ; NO ABONAR',
                                  'FALLO MODULO ; ABONAR')
                THEN 1 
                ELSE 0 
            END) as tiene_fallo
        FROM rma_maestro m
        LEFT JOIN rma_detalles d ON m.id = d.rma_id
        WHERE m.cliente = ?
    """
    
    params = [nombre_cliente]
    
    # Aplicar filtros
    if fecha_desde:
        query += " AND m.fecha_emision >= ?"
        params.append(fecha_desde)
    
    if fecha_hasta:
        query += " AND m.fecha_emision <= ?"
        params.append(fecha_hasta)
    
    if filtro_estado and filtro_estado != "Todos":
        if filtro_estado == "NO ABONAR":
            query += " AND m.resultado_expediente LIKE '%NO ABONAR%'"
        elif filtro_estado == "ABONAR":
            query += """ AND (m.resultado_expediente LIKE '%ABONAR%' 
                        AND m.resultado_expediente NOT LIKE '%NO ABONAR%')"""
        elif filtro_estado == "ABONAR FALLO":
            # Abonos con fallos de producto - se filtrará después con tiene_fallo
            query += """ AND (m.resultado_expediente LIKE '%ABONAR%' 
                        AND m.resultado_expediente NOT LIKE '%NO ABONAR%')"""
        elif filtro_estado == "ABONAR OK":
            # Abonos sin fallos de producto - se filtrará después con tiene_fallo
            query += """ AND (m.resultado_expediente LIKE '%ABONAR%' 
                        AND m.resultado_expediente NOT LIKE '%NO ABONAR%')"""
        elif filtro_estado == "REPOSICION":
            query += " AND m.resultado_expediente LIKE '%REPOSICION%'"
    
    query += " GROUP BY m.id, m.codigo_rma, m.fecha_emision, m.estado, m.resultado_expediente"
    query += " ORDER BY m.fecha_emision DESC"
    
    cursor.execute(query, params)
    resultados = cursor.fetchall()
    
    # Clasificar abonos en fallo vs OK
    datos_procesados = []
    for row in resultados:
        codigo, fecha, estado, resultado, num_art, coste, tiene_fallo = row
        
        # Convertir tiene_fallo a entero para comparación correcta
        tiene_fallo = int(tiene_fallo) if tiene_fallo else 0
        
        # Debug: imprimir para ver qué está pasando
        # print(f"RMA: {codigo}, Resultado: {resultado}, Tiene_fallo: {tiene_fallo}")
        
        # Determinar clasificación
        clasificacion = resultado or "Sin especificar"
        resultado_upper = str(resultado).upper() if resultado else ""
        
        # Aplicar filtros específicos de ABONAR OK vs ABONAR FALLO
        if filtro_estado == "ABONAR FALLO":
            # Solo mostrar abonos con fallo (tiene_fallo = 1)
            if tiene_fallo == 0:
                continue
            clasificacion = f"{resultado} (FALLO)"
        elif filtro_estado == "ABONAR OK":
            # Solo mostrar abonos sin fallo (tiene_fallo = 0)
            if tiene_fallo != 0:
                continue
            clasificacion = f"{resultado} (OK)"
        elif tiene_fallo == 1 and "ABONAR" in resultado_upper and "NO ABONAR" not in resultado_upper:
            # Para otros filtros, añadir etiqueta (FALLO) o (OK)
            clasificacion = f"{resultado} (FALLO)"
        elif tiene_fallo == 0 and "ABONAR" in resultado_upper and "NO ABONAR" not in resultado_upper:
            clasificacion = f"{resultado} (OK)"
        
        datos_procesados.append((
            codigo,
            fecha,
            estado,
            clasificacion,
            int(num_art) if num_art else 0,
            float(coste) if coste else 0.0
        ))
    
    return datos_procesados


def exportar_estadisticas_cliente_excel(nombre_cliente, datos, ruta_archivo):
    """
    Exporta las estadísticas de cliente a Excel.
    
    Args:
        nombre_cliente: Nombre del cliente
        datos: Lista de tuplas con datos de expedientes
        ruta_archivo: Ruta donde guardar el archivo Excel
    
    Returns:
        bool: True si se exportó correctamente
    """
    try:
        # Crear DataFrame
        df = pd.DataFrame(datos, columns=[
            'Código RMA',
            'Fecha Emisión',
            'Estado',
            'Resultado/Clasificación',
            'Nº Artículos',
            'Coste (€)'
        ])
        
        # Calcular totales
        total_expedientes = len(df)
        total_coste = df['Coste (€)'].sum()
        total_articulos = df['Nº Artículos'].sum()
        
        # Agregar fila de totales
        df_totales = pd.DataFrame([{
            'Código RMA': '--- TOTALES ---',
            'Fecha Emisión': '',
            'Estado': '',
            'Resultado/Clasificación': f'{total_expedientes} expedientes',
            'Nº Artículos': total_articulos,
            'Coste (€)': total_coste
        }])
        
        df = pd.concat([df, df_totales], ignore_index=True)
        
        # Crear archivo Excel con formato
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Estadísticas_{nombre_cliente[:20]}', index=False)
            
            # Obtener workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets[f'Estadísticas_{nombre_cliente[:20]}']
            
            # Ajustar anchos de columna
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar formato a encabezados
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formato a fila de totales (última fila)
            totales_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            totales_font = Font(bold=True)
            
            last_row = len(df) + 1
            for cell in worksheet[last_row]:
                cell.fill = totales_fill
                cell.font = totales_font
        
        return True
        
    except Exception as e:
        print(f"Error exportando a Excel: {e}")
        return False


def crear_widget_estadisticas_basicas(parent_frame, stats):
    """
    Crea un widget visual con las estadísticas básicas del cliente.
    
    Args:
        parent_frame: Frame padre donde colocar el widget
        stats: Diccionario con estadísticas básicas
    
    Returns:
        Frame: Frame contenedor con las estadísticas
    """
    stats_frame = ctk.CTkFrame(parent_frame)
    stats_frame.pack(fill="x", pady=(5, 10))
    
    # Grid de 3 columnas
    stats_frame.grid_columnconfigure(0, weight=1)
    stats_frame.grid_columnconfigure(1, weight=1)
    stats_frame.grid_columnconfigure(2, weight=1)
    
    # Estadística 1: Total expedientes
    stat1_frame = ctk.CTkFrame(stats_frame, fg_color=("#3B8ED0", "#1F6AA5"))
    stat1_frame.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
    
    ctk.CTkLabel(stat1_frame, text="📦 Total Expedientes", 
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color="white").pack(pady=(5, 0))
    ctk.CTkLabel(stat1_frame, text=str(stats['total_expedientes']), 
                font=ctk.CTkFont(size=24, weight="bold"),
                text_color="white").pack()
    ctk.CTkLabel(stat1_frame, text=f"{stats['porcentaje_completados']:.1f}% Completados", 
                font=ctk.CTkFont(size=9),
                text_color="white").pack(pady=(0, 5))
    
    # Estadística 2: Coste total
    stat2_frame = ctk.CTkFrame(stats_frame, fg_color=("#2FA572", "#106A43"))
    stat2_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    
    ctk.CTkLabel(stat2_frame, text="💰 Coste Total", 
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color="white").pack(pady=(5, 0))
    ctk.CTkLabel(stat2_frame, text=f"{float(stats['coste_total']):.2f} €", 
                font=ctk.CTkFont(size=20, weight="bold"),
                text_color="white").pack()
    total_exp = int(stats['total_expedientes']) if stats['total_expedientes'] else 0
    coste_total = float(stats['coste_total']) if stats['coste_total'] else 0.0
    promedio = coste_total / total_exp if total_exp > 0 else 0
    ctk.CTkLabel(stat2_frame, text=f"Promedio: {promedio:.2f} €", 
                font=ctk.CTkFont(size=9),
                text_color="white").pack(pady=(0, 5))
    
    # Estadística 3: Abonos
    stat3_frame = ctk.CTkFrame(stats_frame, fg_color=("#E07A5F", "#C85A3F"))
    stat3_frame.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
    
    ctk.CTkLabel(stat3_frame, text="💸 Abonos", 
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color="white").pack(pady=(5, 0))
    ctk.CTkLabel(stat3_frame, text=str(stats['total_abonos']), 
                font=ctk.CTkFont(size=24, weight="bold"),
                text_color="white").pack()
    ctk.CTkLabel(stat3_frame, text=f"{stats['porcentaje_abonos']:.1f}% del total", 
                font=ctk.CTkFont(size=9),
                text_color="white").pack(pady=(0, 5))
    
    return stats_frame
