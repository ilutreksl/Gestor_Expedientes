"""
Módulo de estadísticas: Incidencias por Persona
Analiza el historial de expedientes para extraer "Pedido realizado por: x"
y muestra un ranking de personas con más incidencias.
"""

import customtkinter as ctk
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import sqlite3
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from lib.logger_config import get_logger

logger = get_logger()


def mostrar_incidencias_personas(app):
    """
    Módulo que muestra estadísticas de incidencias por persona basado en el
    historial de expedientes (campo descripcion_cambio en rma_historial).
    
    Parámetros:
    - app: instancia de VentanaPrincipal
    """
    logger.info("Abriendo módulo de Incidencias por Persona")
    
    # Limpiar marco principal
    try:
        app.limpiar_marco_stats()
    except Exception:
        pass

    frame = app.main_stats_frame
    if not frame:
        logger.error("No se encontró el marco principal de estadísticas")
        return

    ctk.CTkLabel(
        frame, 
        text="INCIDENCIAS POR PERSONA", 
        font=ctk.CTkFont(size=18, weight="bold")
    ).pack(pady=16)

    # === CONTROLES Y FILTROS ===
    controles = ctk.CTkFrame(frame)
    controles.pack(fill="x", padx=16, pady=(0, 10))
    controles.grid_columnconfigure((0, 2), weight=0)
    controles.grid_columnconfigure((1, 3), weight=1)

    # Fila 0: Fechas
    ctk.CTkLabel(controles, text="Fecha Inicial:").grid(row=0, column=0, padx=6, pady=6, sticky="w")
    fecha_inicio_entry = ctk.CTkEntry(controles, placeholder_text="DD/MM/YYYY")
    fecha_inicio_entry.grid(row=0, column=1, padx=6, pady=6, sticky="ew")

    ctk.CTkLabel(controles, text="Fecha Final:").grid(row=0, column=2, padx=6, pady=6, sticky="w")
    fecha_fin_entry = ctk.CTkEntry(controles, placeholder_text="DD/MM/YYYY")
    fecha_fin_entry.grid(row=0, column=3, padx=6, pady=6, sticky="ew")

    # Fila 1: Resultado de Expediente
    ctk.CTkLabel(controles, text="Resultado:").grid(row=1, column=0, padx=6, pady=6, sticky="w")
    opciones_resultado = ["Todos"] + (
        getattr(app, 'OPCIONES', {}).get('Resultado_Expediente', []) 
        if hasattr(app, 'OPCIONES') else []
    )
    resultado_opt = ctk.CTkOptionMenu(controles, values=opciones_resultado)
    resultado_opt.grid(row=1, column=1, padx=6, pady=6, sticky="ew")
    resultado_opt.set(opciones_resultado[0])

    # Checkbox: incluir expedientes no completados
    incluir_no_completados_var = ctk.BooleanVar(value=False)
    incluir_no_completados_chk = ctk.CTkCheckBox(
        controles,
        text="Incluir expedientes no completados",
        variable=incluir_no_completados_var,
        command=lambda: cargar_datos()
    )
    incluir_no_completados_chk.grid(row=1, column=2, columnspan=2, padx=6, pady=6, sticky="w")

    # Botones: Aplicar Filtros y Exportar
    btn_frame = ctk.CTkFrame(controles)
    btn_frame.grid(row=2, column=0, columnspan=4, pady=10)

    # Frame de resultados
    resultados_frame = ctk.CTkFrame(frame)
    resultados_frame.pack(fill="both", expand=True, padx=16, pady=(0, 10))

    # Cache de resultados para exportación
    results_cache = []

    def construir_where_and_params():
        """
        Construye las cláusulas WHERE y parámetros para filtrar los datos.
        Retorna: (where_clause, params)
        """
        where = ["1=1"]
        params = []

        logger.debug("=== Construyendo filtros de incidencias por persona ===")

        # Filtro de expedientes completos (por defecto, solo completos)
        incluir_no_completados = incluir_no_completados_var.get()
        logger.debug(f"Incluir expedientes no completados: {incluir_no_completados}")
        if not incluir_no_completados:
            where.append("(m.estado = 'Completado' OR (m.fecha_gestion IS NOT NULL AND m.fecha_gestion != ''))")

        # Filtro de fecha inicial
        fecha_ini = fecha_inicio_entry.get().strip()
        logger.debug(f"Filtro Fecha Inicio: '{fecha_ini}'")
        if fecha_ini:
            try:
                # Validar formato DD/MM/YYYY y convertir a ISO (YYYY-MM-DD)
                fecha_obj = datetime.strptime(fecha_ini, '%d/%m/%Y')
                fecha_iso = fecha_obj.strftime('%Y-%m-%d')
                where.append("date(h.fecha_cambio) >= ?")
                params.append(fecha_iso)
                logger.debug(f"  -> Aplicado filtro fecha inicio: {fecha_iso}")
            except ValueError:
                logger.warning(f"Fecha inicial con formato inválido: {fecha_ini}")
                messagebox.showerror("Error", f"Fecha inicial inválida: {fecha_ini}. Use DD/MM/YYYY")
                return None, None
        
        # Filtro de fecha final
        fecha_fin = fecha_fin_entry.get().strip()
        logger.debug(f"Filtro Fecha Fin: '{fecha_fin}'")
        if fecha_fin:
            try:
                fecha_obj = datetime.strptime(fecha_fin, '%d/%m/%Y')
                fecha_iso = fecha_obj.strftime('%Y-%m-%d')
                where.append("date(h.fecha_cambio) <= ?")
                params.append(fecha_iso)
                logger.debug(f"  -> Aplicado filtro fecha fin: {fecha_iso}")
            except ValueError:
                logger.warning(f"Fecha final con formato inválido: {fecha_fin}")
                messagebox.showerror("Error", f"Fecha final inválida: {fecha_fin}. Use DD/MM/YYYY")
                return None, None
        
        # Filtro de resultado de expediente
        resultado = resultado_opt.get()
        logger.debug(f"Filtro Resultado: '{resultado}'")
        if resultado and resultado != 'Todos':
            where.append("LOWER(m.resultado_expediente) = ?")
            params.append(resultado.strip().lower())
            logger.debug(f"  -> Aplicado filtro resultado: {resultado.strip().lower()}")
        
        where_final = " AND ".join(where)
        logger.debug(f"WHERE final: {where_final}")
        logger.debug(f"Parámetros: {params}")
        
        return where_final, params

    def cargar_datos():
        """Carga y muestra las estadísticas de incidencias por persona."""
        logger.info("Cargando datos de incidencias por persona")
        
        # Limpiar resultados anteriores
        for w in resultados_frame.winfo_children():
            w.destroy()

        conn, cursor = app.master.conectar_db()
        if not conn:
            logger.error("No se pudo conectar a la base de datos")
            messagebox.showerror("Error", "No se pudo conectar a la base de datos.")
            return

        where_clause, params = construir_where_and_params()
        if where_clause is None:  # Error en validación de fechas
            conn.close()
            return

        # Consulta SQL para extraer personas del historial
        # Patrón regex: "Pedido realizado por: Nombre"
        sql = f"""
            SELECT 
                h.descripcion_cambio,
                m.codigo_rma,
                m.resultado_expediente,
                m.fecha_gestion,
                m.id as rma_id
            FROM rma_historial h
            INNER JOIN rma_maestro m ON h.rma_id = m.id
            WHERE {where_clause}
                AND h.descripcion_cambio LIKE '%Pedido realizado por:%'
            ORDER BY h.fecha_cambio DESC
        """
        
        logger.debug(f"SQL ejecutada: {sql}")
        logger.debug(f"Parámetros: {params}")
        
        try:
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            logger.info(f"Se encontraron {len(rows)} registros de historial con 'Pedido realizado por:'")
            
            # Procesar resultados y extraer nombres
            personas_dict = {}  # {nombre: [expedientes]}
            # Patrón mejorado: captura nombres completos hasta paréntesis o final de línea
            patron = re.compile(r'Pedido realizado por:\s*([^(\n]+?)(?:\s*\(|$)', re.IGNORECASE)
            
            for row in rows:
                descripcion = row[0]
                codigo_rma = row[1]
                resultado = row[2]
                fecha = row[3]
                rma_id = row[4]
                
                # Buscar el nombre en la descripción
                match = patron.search(descripcion)
                if match:
                    nombre = match.group(1).strip()
                    if nombre:
                        if nombre not in personas_dict:
                            personas_dict[nombre] = []
                        personas_dict[nombre].append({
                            'codigo': codigo_rma,
                            'resultado': resultado,
                            'fecha': fecha,
                            'rma_id': rma_id
                        })
            
            logger.info(f"Se identificaron {len(personas_dict)} personas únicas")
            
            # Ordenar por cantidad de expedientes (descendente)
            personas_ordenadas = sorted(
                personas_dict.items(), 
                key=lambda x: len(x[1]), 
                reverse=True
            )
            
            # Guardar en cache para exportación
            results_cache.clear()
            results_cache.extend(personas_ordenadas)
            
            # Mostrar resultados en la interfaz
            if not personas_ordenadas:
                ctk.CTkLabel(
                    resultados_frame,
                    text="No se encontraron registros con 'Pedido realizado por:' en el historial",
                    font=ctk.CTkFont(size=14)
                ).pack(pady=20)
                conn.close()
                return
            
            # Crear tabla con scrollbar
            body = ctk.CTkScrollableFrame(resultados_frame, label_text='Resultados')
            body.pack(fill='both', expand=True, padx=8, pady=8)

            table = ctk.CTkFrame(body)
            table.pack(fill='both', expand=True, padx=4, pady=4)
            table.grid_columnconfigure(0, weight=3)  # Persona (más ancha)
            table.grid_columnconfigure(1, weight=1)  # Total Expedientes

            # Encabezados
            ctk.CTkLabel(
                table, 
                text='PERSONA', 
                font=ctk.CTkFont(weight='bold')
            ).grid(row=0, column=0, sticky='w', padx=8, pady=4)
            
            ctk.CTkLabel(
                table, 
                text='TOTAL EXPEDIENTES', 
                font=ctk.CTkFont(weight='bold')
            ).grid(row=0, column=1, sticky='e', padx=8, pady=4)

            # Filas de datos
            for idx, (persona, expedientes) in enumerate(personas_ordenadas, start=1):
                total_exp = len(expedientes)
                
                # Nombre de la persona (clickeable)
                btn_persona = ctk.CTkButton(
                    table,
                    text=persona,
                    anchor="w",
                    command=lambda p=persona, e=expedientes: abrir_detalle_persona(p, e)
                )
                btn_persona.grid(row=idx, column=0, sticky='ew', padx=8, pady=2)
                
                # Total de expedientes
                ctk.CTkLabel(table, text=str(total_exp)).grid(
                    row=idx, column=1, sticky='e', padx=8, pady=2
                )
            
            conn.close()
            logger.info("Datos cargados exitosamente en la interfaz")
            
        except sqlite3.Error as e:
            logger.error(f"Error de BD al consultar incidencias: {e}", exc_info=True)
            messagebox.showerror('Error BD', f'Error al consultar datos: {e}')
            conn.close()
        except Exception as e:
            logger.error(f"Error mostrando resultados: {e}", exc_info=True)
            messagebox.showerror('Error', f'Error al procesar resultados: {e}')
            conn.close()

    def abrir_detalle_persona(persona, expedientes):
        """
        Abre una ventana con el detalle de expedientes de una persona específica,
        incluyendo referencias de productos afectados.
        """
        logger.info(f"Abriendo detalle para persona: {persona}")
        
        # Crear ventana secundaria
        detalle_win = ctk.CTkToplevel(app)
        detalle_win.title(f"Detalle de {persona}")
        detalle_win.geometry("900x600")
        
        # Configurar ventana para que aparezca al frente
        detalle_win.transient(app.stats_window if hasattr(app, 'stats_window') else app)
        detalle_win.lift()
        detalle_win.focus_force()
        detalle_win.attributes('-topmost', True)
        detalle_win.after(100, lambda: detalle_win.attributes('-topmost', False))
        
        # Encabezado
        header_frame = ctk.CTkFrame(detalle_win)
        header_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(
            header_frame,
            text=f"Expedientes de: {persona}",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(side="left", padx=10)
        
        ctk.CTkLabel(
            header_frame,
            text=f"Total: {len(expedientes)} expedientes",
            font=ctk.CTkFont(size=14)
        ).pack(side="right", padx=10)
        
        # Tabla con expedientes y referencias
        tabla_frame = ctk.CTkScrollableFrame(detalle_win)
        tabla_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Encabezados
        headers = ["Código RMA", "Resultado", "Fecha", "Referencias Productos"]
        for col, header in enumerate(headers):
            ctk.CTkLabel(
                tabla_frame,
                text=header,
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=col, padx=5, pady=5, sticky="w")
        
        # Obtener referencias de productos para cada expediente
        conn, cursor = app.master.conectar_db()
        if not conn:
            logger.error("No se pudo conectar a la BD para obtener referencias")
            detalle_win.destroy()
            return
        
        def abrir_editor_expediente(codigo_rma):
            """Abre el editor de expediente al hacer clic en el código."""
            try:
                logger.info(f"Abriendo editor para expediente: {codigo_rma}")
                # Obtener rma_id del código
                conn_temp, cursor_temp = app.master.conectar_db()
                if not conn_temp:
                    logger.error("No se pudo conectar a la BD")
                    return
                cursor_temp.execute("SELECT id FROM rma_maestro WHERE Codigo_RMA = ?", (codigo_rma,))
                resultado = cursor_temp.fetchone()
                conn_temp.close()
                
                if resultado:
                    rma_id = resultado[0]
                    app._abrir_editor_rma(rma_id=rma_id)
                    logger.info(f"Editor abierto exitosamente para {codigo_rma}")
                else:
                    logger.warning(f"No se encontró expediente: {codigo_rma}")
                    messagebox.showwarning("No encontrado", f"No se encontró el expediente {codigo_rma}")
            except Exception as e:
                logger.error(f"Error al abrir editor de expediente: {e}", exc_info=True)
                messagebox.showerror("Error", f"No se pudo abrir el expediente:\n{e}")
        
        try:
            for idx, exp in enumerate(expedientes, start=1):
                # Obtener referencias de productos del expediente
                cursor.execute(
                    "SELECT referencia_articulo FROM rma_detalles WHERE rma_id = ?",
                    (exp['rma_id'],)
                )
                refs = cursor.fetchall()
                referencias_texto = ", ".join([r[0] for r in refs if r[0]]) if refs else "(Sin referencias)"
                
                # Código RMA (clickeable para editar)
                btn_codigo = ctk.CTkButton(
                    tabla_frame,
                    text=exp['codigo'],
                    width=120,
                    anchor="w",
                    command=lambda c=exp['codigo']: abrir_editor_expediente(c)
                )
                btn_codigo.grid(row=idx, column=0, padx=5, pady=2, sticky="w")
                
                # Resultado
                ctk.CTkLabel(tabla_frame, text=exp.get('resultado', 'N/A')).grid(
                    row=idx, column=1, padx=5, pady=2, sticky="w"
                )
                
                # Fecha
                fecha_display = exp.get('fecha', 'N/A')
                ctk.CTkLabel(tabla_frame, text=fecha_display).grid(
                    row=idx, column=2, padx=5, pady=2, sticky="w"
                )
                
                # Referencias
                ctk.CTkLabel(tabla_frame, text=referencias_texto).grid(
                    row=idx, column=3, padx=5, pady=2, sticky="w"
                )
            
            conn.close()
            logger.debug(f"Mostrados {len(expedientes)} expedientes para {persona}")
            
        except sqlite3.Error as e:
            logger.error(f"Error obteniendo referencias de productos: {e}", exc_info=True)
            messagebox.showerror("Error", f"Error al obtener referencias: {e}")
            conn.close()
            detalle_win.destroy()

    def exportar_excel():
        """Exporta los resultados a un archivo Excel formateado."""
        logger.info("Iniciando exportación a Excel")
        
        if not results_cache:
            logger.warning("No hay datos para exportar")
            messagebox.showwarning("Sin datos", "No hay datos para exportar. Primero aplique filtros.")
            return
        
        # Diálogo para guardar archivo
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar Excel"
        )
        
        if not archivo:
            logger.info("Exportación cancelada por el usuario")
            return
        
        try:
            logger.debug(f"Creando archivo Excel: {archivo}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Incidencias por Persona"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border_style = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Fila 1-2: Información de fechas
            fecha_ini = fecha_inicio_entry.get().strip() or "No especificada"
            fecha_fin = fecha_fin_entry.get().strip() or "No especificada"
            
            ws['A1'] = "INCIDENCIAS POR PERSONA - REPORTE"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A2'] = f"Período: {fecha_ini} - {fecha_fin}"
            ws['A2'].font = Font(size=11)
            
            ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A3'].font = Font(size=10, italic=True)
            
            # Obtener conexión para referencias de productos
            conn, cursor = app.master.conectar_db()
            if not conn:
                logger.error("No se pudo conectar a la BD para exportar")
                messagebox.showerror("Error", "No se pudo conectar a la base de datos")
                return
            
            # Fila actual
            current_row = 5
            
            # Por cada persona, crear tabla con sus expedientes
            for persona, expedientes in results_cache:
                logger.debug(f"Exportando datos de {persona}: {len(expedientes)} expedientes")
                
                # Encabezado de persona
                ws[f'A{current_row}'] = f"PERSONA: {persona}"
                ws[f'A{current_row}'].font = Font(bold=True, size=12, color="1F4E78")
                current_row += 1
                
                ws[f'A{current_row}'] = f"Total expedientes: {len(expedientes)}"
                ws[f'A{current_row}'].font = Font(italic=True, size=10)
                current_row += 1
                
                # Encabezados de tabla
                headers = ['Código RMA', 'Resultado', 'Fecha', 'Referencias Productos']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_style
                
                current_row += 1
                
                # Datos de expedientes
                for exp in expedientes:
                    # Obtener referencias de productos
                    cursor.execute(
                        "SELECT referencia_articulo FROM rma_detalles WHERE rma_id = ?",
                        (exp['rma_id'],)
                    )
                    refs = cursor.fetchall()
                    referencias = ", ".join([r[0] for r in refs if r[0]]) if refs else "(Sin referencias)"
                    
                    # Escribir datos
                    datos = [
                        exp['codigo'],
                        exp.get('resultado', 'N/A'),
                        exp.get('fecha', 'N/A'),
                        referencias
                    ]
                    
                    for col_idx, dato in enumerate(datos, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=dato)
                        cell.border = border_style
                        if col_idx == 1:  # Alineación izquierda para código
                            cell.alignment = Alignment(horizontal='left')
                        elif col_idx == 3:  # Fecha centrada
                            cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1
                
                # Espacio entre personas
                current_row += 2
            
            conn.close()
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 20  # Código RMA
            ws.column_dimensions['B'].width = 25  # Resultado
            ws.column_dimensions['C'].width = 15  # Fecha
            ws.column_dimensions['D'].width = 50  # Referencias
            
            # Congelar primeras filas
            ws.freeze_panes = 'A6'
            
            # Guardar archivo
            wb.save(archivo)
            logger.info(f"Excel exportado exitosamente: {archivo}")
            messagebox.showinfo("Éxito", f"Excel exportado correctamente:\n{archivo}")
            
        except Exception as e:
            logger.error(f"Error al exportar Excel: {e}", exc_info=True)
            messagebox.showerror("Error", f"Error al exportar: {e}")

    # Configurar botones
    ctk.CTkButton(btn_frame, text='Aplicar Filtros', command=cargar_datos).pack(side='left', padx=5)
    ctk.CTkButton(btn_frame, text='Exportar a Excel', command=exportar_excel).pack(side='left', padx=5)

    # Cargar datos iniciales (sin filtros)
    cargar_datos()
