"""Mixin extraido automaticamente de VentanaPrincipal (app.py).

Estas clases NO son instanciables por si solas: solo aportan metodos que se
combinan con VentanaPrincipal via herencia multiple. Dependen de atributos de
instancia (self.conn, self.username, self.tree_rmas, etc.) inicializados en
VentanaPrincipal.__init__.
"""
from lib.app_core import *  # noqa: F401,F403 - helpers/constantes/imports compartidos con app.py
from lib.app_core import _get_cached_query, invalidate_cache  # nombres "privados" que el wildcard import no trae

class RecepcionesMixin:
    def mostrar_recepciones_anticipadas(self):
        """Muestra estadísticas de recepciones anticipadas (recepción antes de autorización)."""
        from lib.recepciones_anticipadas import obtener_recepciones_anticipadas, obtener_expedientes_anticipados_por_cliente, buscar_clientes_anticipados, ordenar_resultados
        
        self.limpiar_marco_stats()
        
        # Título
        ctk.CTkLabel(self.main_stats_frame, 
                     text="⚠️ RECEPCIONES ANTICIPADAS", 
                     font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=20)
        
        ctk.CTkLabel(self.main_stats_frame,
                     text="Clientes que reciben productos antes de autorizarlos",
                     font=ctk.CTkFont(size=12),
                     text_color="gray"
        ).pack(pady=(0, 10))
        
        # Frame de controles (búsqueda y ordenamiento)
        controles_frame = ctk.CTkFrame(self.main_stats_frame)
        controles_frame.pack(fill="x", padx=20, pady=10)
        
        # Búsqueda
        ctk.CTkLabel(controles_frame, text="Buscar cliente:").pack(side="left", padx=5)
        entry_buscar = ctk.CTkEntry(controles_frame, placeholder_text="Nombre del cliente...", width=250)
        entry_buscar.pack(side="left", padx=5)
        
        # Ordenamiento
        ctk.CTkLabel(controles_frame, text="Ordenar por:").pack(side="left", padx=(20, 5))
        combo_orden = ctk.CTkOptionMenu(controles_frame, 
                                        values=["Cantidad", "Cliente", "Media días"],
                                        width=150)
        combo_orden.set("Cantidad")
        combo_orden.pack(side="left", padx=5)
        
        # Frame scrollable para resultados
        scroll_frame = ctk.CTkScrollableFrame(self.main_stats_frame, height=400)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        def actualizar_listado():
            # Limpiar resultados anteriores
            for widget in scroll_frame.winfo_children():
                widget.destroy()
            
            # Obtener datos
            conn, cursor = self.master.conectar_db()
            if not conn:
                ctk.CTkLabel(scroll_frame, text="Error al conectar con la base de datos.",
                            text_color="red").pack(pady=20)
                return
            
            try:
                termino = entry_buscar.get().strip()
                if termino:
                    resultados = buscar_clientes_anticipados(conn, termino)
                else:
                    resultados = obtener_recepciones_anticipadas(conn)
                
                conn.close()
                
                if not resultados:
                    ctk.CTkLabel(scroll_frame, 
                               text="No se encontraron recepciones anticipadas.",
                               text_color="gray").pack(pady=20)
                    return
                
                # Ordenar según criterio
                criterio_map = {
                    "Cantidad": "cantidad",
                    "Cliente": "cliente",
                    "Media días": "media_dias"
                }
                criterio = criterio_map.get(combo_orden.get(), "cantidad")
                resultados = ordenar_resultados(resultados, criterio)
                
                # Encabezados
                header_frame = ctk.CTkFrame(scroll_frame)
                header_frame.pack(fill="x", pady=(0, 10))
                header_frame.grid_columnconfigure(0, weight=3)
                header_frame.grid_columnconfigure(1, weight=1)
                header_frame.grid_columnconfigure(2, weight=1)
                
                header_font = ctk.CTkFont(weight="bold", size=12)
                ctk.CTkLabel(header_frame, text="CLIENTE", font=header_font, anchor="w").grid(row=0, column=0, padx=10, pady=5, sticky="w")
                ctk.CTkLabel(header_frame, text="EXPEDIENTES", font=header_font, anchor="center").grid(row=0, column=1, padx=10, pady=5)
                ctk.CTkLabel(header_frame, text="MEDIA DÍAS", font=header_font, anchor="center").grid(row=0, column=2, padx=10, pady=5)
                
                # Filas de datos
                for cliente, cantidad, media_dias in resultados:
                    row_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
                    row_frame.pack(fill="x", pady=2)
                    row_frame.grid_columnconfigure(0, weight=3)
                    row_frame.grid_columnconfigure(1, weight=1)
                    row_frame.grid_columnconfigure(2, weight=1)
                    
                    # Cliente (clickeable)
                    cliente_btn = ctk.CTkButton(row_frame, 
                                               text=cliente,
                                               anchor="w",
                                               command=lambda c=cliente: self.mostrar_detalle_recepciones_cliente(c))
                    cliente_btn.grid(row=0, column=0, padx=10, pady=2, sticky="ew")
                    
                    # Cantidad
                    ctk.CTkLabel(row_frame, text=str(cantidad), anchor="center").grid(row=0, column=1, padx=10, pady=2)
                    
                    # Media días
                    media_texto = f"{media_dias:.1f}" if media_dias is not None else "-"
                    color = "red" if media_dias and media_dias > 7 else "orange" if media_dias and media_dias > 3 else "green"
                    ctk.CTkLabel(row_frame, text=media_texto, text_color=color, anchor="center").grid(row=0, column=2, padx=10, pady=2)
                
            except Exception as e:
                print(f"Error al cargar recepciones anticipadas: {e}")
                ctk.CTkLabel(scroll_frame,
                            text=f"Error: {e}",
                            text_color="red").pack(pady=20)
        
        # Botones de acción
        btn_actualizar = ctk.CTkButton(controles_frame, text="🔍 Buscar", command=actualizar_listado, width=100)
        btn_actualizar.pack(side="left", padx=10)
        
        # Botones de exportación e impresión
        btn_frame_export = ctk.CTkFrame(self.main_stats_frame)
        btn_frame_export.pack(pady=10)
        
        def exportar_listado():
            # Obtener datos actuales
            conn, cursor = self.master.conectar_db()
            if not conn:
                return
            termino = entry_buscar.get().strip()
            if termino:
                resultados = buscar_clientes_anticipados(conn, termino)
            else:
                resultados = obtener_recepciones_anticipadas(conn)
            conn.close()
            
            if resultados:
                criterio_map = {"Cantidad": "cantidad", "Cliente": "cliente", "Media días": "media_dias"}
                criterio = criterio_map.get(combo_orden.get(), "cantidad")
                resultados = ordenar_resultados(resultados, criterio)
                self._exportar_listado_recepciones_excel(resultados)
        
        def imprimir_listado():
            # Obtener datos actuales
            conn, cursor = self.master.conectar_db()
            if not conn:
                return
            termino = entry_buscar.get().strip()
            if termino:
                resultados = buscar_clientes_anticipados(conn, termino)
            else:
                resultados = obtener_recepciones_anticipadas(conn)
            conn.close()
            
            if resultados:
                criterio_map = {"Cantidad": "cantidad", "Cliente": "cliente", "Media días": "media_dias"}
                criterio = criterio_map.get(combo_orden.get(), "cantidad")
                resultados = ordenar_resultados(resultados, criterio)
                self._imprimir_listado_recepciones_pdf(resultados)
        
        ctk.CTkButton(btn_frame_export, text="📄 Imprimir PDF", 
                     command=imprimir_listado, width=130).pack(side="left", padx=5)
        
        ctk.CTkButton(btn_frame_export, text="📊 Exportar Excel", 
                     command=exportar_listado, width=130).pack(side="left", padx=5)
        
        # Bind enter en búsqueda
        entry_buscar.bind("<Return>", lambda e: actualizar_listado())
        combo_orden.configure(command=lambda v: actualizar_listado())
        
        # Cargar datos iniciales
        actualizar_listado()

    def mostrar_detalle_recepciones_cliente(self, cliente):
        """Muestra ventana con detalle de expedientes con recepción anticipada de un cliente."""
        from lib.recepciones_anticipadas import obtener_expedientes_anticipados_por_cliente
        
        # Crear ventana
        ventana = ctk.CTkToplevel(self)
        ventana.title(f"Recepciones Anticipadas - {cliente}")
        ventana.geometry("900x600")
        ventana.transient(self)
        
        try:
            ventana.iconbitmap("Icono_Ilutrek.ico")
        except Exception:
            pass
        
        # Título
        ctk.CTkLabel(ventana,
                     text=f"📋 Expedientes con recepción anticipada",
                     font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=10)
        
        ctk.CTkLabel(ventana,
                     text=f"Cliente: {cliente}",
                     font=ctk.CTkFont(size=14),
                     text_color="gray"
        ).pack(pady=(0, 10))
        
        # Frame scrollable
        scroll_frame = ctk.CTkScrollableFrame(ventana)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Obtener datos
        conn, cursor = self.master.conectar_db()
        if not conn:
            ctk.CTkLabel(scroll_frame, text="Error al conectar con la base de datos.",
                        text_color="red").pack(pady=20)
            return
        
        try:
            expedientes = obtener_expedientes_anticipados_por_cliente(conn, cliente)
            conn.close()
            
            if not expedientes:
                ctk.CTkLabel(scroll_frame,
                           text="No se encontraron expedientes.",
                           text_color="gray").pack(pady=20)
                return
            
            # Encabezados
            header_frame = ctk.CTkFrame(scroll_frame)
            header_frame.pack(fill="x", pady=(0, 10))
            header_frame.grid_columnconfigure(0, weight=2)
            header_frame.grid_columnconfigure(1, weight=2)
            header_frame.grid_columnconfigure(2, weight=2)
            header_frame.grid_columnconfigure(3, weight=1)
            header_frame.grid_columnconfigure(4, weight=1)
            
            header_font = ctk.CTkFont(weight="bold", size=11)
            ctk.CTkLabel(header_frame, text="CÓDIGO RMA", font=header_font, anchor="w").grid(row=0, column=0, padx=10, pady=5, sticky="w")
            ctk.CTkLabel(header_frame, text="FECHA RECEPCIÓN", font=header_font, anchor="center").grid(row=0, column=1, padx=10, pady=5)
            ctk.CTkLabel(header_frame, text="FECHA AUTORIZACIÓN", font=header_font, anchor="center").grid(row=0, column=2, padx=10, pady=5)
            ctk.CTkLabel(header_frame, text="DÍAS", font=header_font, anchor="center").grid(row=0, column=3, padx=10, pady=5)
            ctk.CTkLabel(header_frame, text="ESTADO", font=header_font, anchor="center").grid(row=0, column=4, padx=10, pady=5)
            
            # Filas de expedientes
            for exp_id, codigo_rma, fecha_recep, fecha_auto, dias, estado in expedientes:
                row_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
                row_frame.pack(fill="x", pady=2)
                row_frame.grid_columnconfigure(0, weight=2)
                row_frame.grid_columnconfigure(1, weight=2)
                row_frame.grid_columnconfigure(2, weight=2)
                row_frame.grid_columnconfigure(3, weight=1)
                row_frame.grid_columnconfigure(4, weight=1)
                
                # Código (clickeable)
                codigo_btn = ctk.CTkButton(row_frame,
                                          text=codigo_rma,
                                          anchor="w",
                                          command=lambda eid=exp_id: [self._abrir_editor_rma(rma_id=eid), ventana.destroy()])
                codigo_btn.grid(row=0, column=0, padx=10, pady=2, sticky="ew")
                
                # Fechas
                ctk.CTkLabel(row_frame, text=fecha_recep or "-", anchor="center").grid(row=0, column=1, padx=10, pady=2)
                ctk.CTkLabel(row_frame, text=fecha_auto or "-", anchor="center").grid(row=0, column=2, padx=10, pady=2)
                
                # Días de adelanto
                dias_texto = f"{int(dias)}" if dias else "-"
                color = "red" if dias and dias > 7 else "orange" if dias and dias > 3 else "green"
                ctk.CTkLabel(row_frame, text=dias_texto, text_color=color, anchor="center", font=ctk.CTkFont(weight="bold")).grid(row=0, column=3, padx=10, pady=2)
                
                # Estado
                ctk.CTkLabel(row_frame, text=estado or "-", anchor="center", text_color="gray").grid(row=0, column=4, padx=10, pady=2)
        
        except Exception as e:
            print(f"Error al cargar detalle de recepciones: {e}")
            ctk.CTkLabel(scroll_frame,
                        text=f"Error: {e}",
                        text_color="red").pack(pady=20)
        
        # Botones de acción
        botones_frame = ctk.CTkFrame(ventana)
        botones_frame.pack(pady=10)
        
        ctk.CTkButton(botones_frame, text="📄 Imprimir PDF", 
                     command=lambda: self._imprimir_recepciones_anticipadas(cliente, expedientes),
                     width=120).pack(side="left", padx=5)
        
        ctk.CTkButton(botones_frame, text="📊 Exportar Excel", 
                     command=lambda: self._exportar_recepciones_excel(cliente, expedientes),
                     width=120).pack(side="left", padx=5)
        
        ctk.CTkButton(botones_frame, text="Cerrar", command=ventana.destroy, width=100).pack(side="left", padx=5)

    def limpiar_marco_stats(self):
        """Método auxiliar necesario para limpiar el marco principal antes de cargar una estadística."""
        for widget in self.main_stats_frame.winfo_children():
            widget.destroy()

    def _exportar_recepciones_excel(self, cliente, expedientes):
        """Exporta el listado de recepciones anticipadas a Excel con formato."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from tkinter import filedialog
        except ImportError:
            messagebox.showerror("Error", "OpenPyXL no está instalado. Instale openpyxl:\npip install openpyxl")
            return
        
        if not expedientes:
            messagebox.showwarning("Exportar", "No hay datos para exportar.")
            return
        
        # Solicitar ubicación para guardar
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Recepciones_Anticipadas_{cliente.replace(' ', '_')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            # Crear libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recepciones Anticipadas"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = f"RECEPCIONES ANTICIPADAS - {cliente}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Fecha de generación
            ws.merge_cells('A2:F2')
            fecha_cell = ws['A2']
            fecha_cell.value = f"Generado el: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
            fecha_cell.alignment = Alignment(horizontal="center")
            fecha_cell.font = Font(italic=True, size=10)
            
            # Encabezados
            headers = ["Código RMA", "Fecha Recepción", "Fecha Autorización", "Días Adelanto", "Estado", "Observaciones"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            for row_idx, (exp_id, codigo_rma, fecha_recep, fecha_auto, dias, estado) in enumerate(expedientes, start=5):
                ws.cell(row=row_idx, column=1, value=codigo_rma).border = border
                ws.cell(row=row_idx, column=2, value=fecha_recep or "-").border = border
                ws.cell(row=row_idx, column=3, value=fecha_auto or "-").border = border
                
                # Días con formato condicional
                dias_cell = ws.cell(row=row_idx, column=4, value=int(dias) if dias else 0)
                dias_cell.border = border
                dias_cell.alignment = Alignment(horizontal="center")
                
                if dias and dias > 7:
                    dias_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    dias_cell.font = Font(color="9C0006", bold=True)
                elif dias and dias > 3:
                    dias_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    dias_cell.font = Font(color="9C6500", bold=True)
                else:
                    dias_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    dias_cell.font = Font(color="006100", bold=True)
                
                ws.cell(row=row_idx, column=5, value=estado or "-").border = border
                ws.cell(row=row_idx, column=6, value="").border = border
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 30
            
            # Estadísticas al final
            last_row = len(expedientes) + 6
            ws.merge_cells(f'A{last_row}:B{last_row}')
            stats_cell = ws[f'A{last_row}']
            stats_cell.value = "Total expedientes:"
            stats_cell.font = Font(bold=True)
            stats_cell.alignment = Alignment(horizontal="right")
            
            ws[f'C{last_row}'] = len(expedientes)
            ws[f'C{last_row}'].font = Font(bold=True)
            ws[f'C{last_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Media de días
            if expedientes:
                media_dias = sum(dias for _, _, _, _, dias, _ in expedientes if dias) / len([d for _, _, _, _, d, _ in expedientes if d])
                ws.merge_cells(f'A{last_row+1}:B{last_row+1}')
                media_cell = ws[f'A{last_row+1}']
                media_cell.value = "Media días adelanto:"
                media_cell.font = Font(bold=True)
                media_cell.alignment = Alignment(horizontal="right")
                
                ws[f'C{last_row+1}'] = f"{media_dias:.1f}"
                ws[f'C{last_row+1}'].font = Font(bold=True)
                ws[f'C{last_row+1}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Guardar
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Archivo exportado correctamente:\n{filename}")
            
            # Abrir archivo
            import os
            os.startfile(filename)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{e}")

    def _exportar_listado_recepciones_excel(self, resultados):
        """Exporta el listado completo de clientes con recepciones anticipadas a Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from tkinter import filedialog
        except ImportError:
            messagebox.showerror("Error", "OpenPyXL no está instalado. Instale openpyxl:\npip install openpyxl")
            return
        
        if not resultados:
            messagebox.showwarning("Exportar", "No hay datos para exportar.")
            return
        
        # Solicitar ubicación para guardar
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Recepciones_Anticipadas_Listado.xlsx"
        )
        
        if not filename:
            return
        
        try:
            # Crear libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recepciones Anticipadas"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = "RECEPCIONES ANTICIPADAS - LISTADO DE CLIENTES"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Fecha de generación
            ws.merge_cells('A2:C2')
            fecha_cell = ws['A2']
            fecha_cell.value = f"Generado el: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
            fecha_cell.alignment = Alignment(horizontal="center")
            fecha_cell.font = Font(italic=True, size=10)
            
            # Encabezados
            headers = ["Cliente", "Nº Expedientes", "Media Días Adelanto"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            for row_idx, (cliente, cantidad, media_dias) in enumerate(resultados, start=5):
                ws.cell(row=row_idx, column=1, value=cliente).border = border
                ws.cell(row=row_idx, column=2, value=cantidad).border = border
                ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
                
                # Media días con formato condicional
                media_cell = ws.cell(row=row_idx, column=3, value=f"{media_dias:.1f}" if media_dias else "0")
                media_cell.border = border
                media_cell.alignment = Alignment(horizontal="center")
                
                if media_dias and media_dias > 7:
                    media_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    media_cell.font = Font(color="9C0006", bold=True)
                elif media_dias and media_dias > 3:
                    media_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    media_cell.font = Font(color="9C6500", bold=True)
                else:
                    media_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    media_cell.font = Font(color="006100", bold=True)
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 22
            
            # Estadísticas al final
            last_row = len(resultados) + 6
            ws[f'A{last_row}'] = "Total clientes:"
            ws[f'A{last_row}'].font = Font(bold=True)
            ws[f'A{last_row}'].alignment = Alignment(horizontal="right")
            
            ws[f'B{last_row}'] = len(resultados)
            ws[f'B{last_row}'].font = Font(bold=True)
            ws[f'B{last_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Total expedientes
            total_expedientes = sum(cantidad for _, cantidad, _ in resultados)
            ws[f'A{last_row+1}'] = "Total expedientes:"
            ws[f'A{last_row+1}'].font = Font(bold=True)
            ws[f'A{last_row+1}'].alignment = Alignment(horizontal="right")
            
            ws[f'B{last_row+1}'] = total_expedientes
            ws[f'B{last_row+1}'].font = Font(bold=True)
            ws[f'B{last_row+1}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Guardar
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Archivo exportado correctamente:\n{filename}")
            
            # Abrir archivo
            import os
            os.startfile(filename)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{e}")

    def _imprimir_listado_recepciones_pdf(self, resultados):
        """Genera un PDF con el listado completo de clientes con recepciones anticipadas."""
        from tkinter import filedialog
        import os
        
        if not resultados:
            messagebox.showwarning("Imprimir", "No hay datos para imprimir.")
            return
        
        # Solicitar ubicación para guardar PDF
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"Recepciones_Anticipadas_Listado.pdf"
        )
        
        if not filename:
            return
        
        try:
            # Importar reportlab
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER
            except ImportError:
                messagebox.showerror("Error", "ReportLab no está instalado. Instale reportlab:\npip install reportlab")
                return
            
            # Crear PDF
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("RECEPCIONES ANTICIPADAS - LISTADO DE CLIENTES", title_style))
            
            # Fecha
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(f"Generado el: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
            
            # Tabla de datos
            data = [['Cliente', 'Nº Expedientes', 'Media Días']]
            
            for cliente, cantidad, media_dias in resultados:
                data.append([
                    cliente,
                    str(cantidad),
                    f"{media_dias:.1f}" if media_dias else "0"
                ])
            
            # Crear tabla
            table = Table(data, colWidths=[10*cm, 3*cm, 3*cm])
            
            # Estilo de tabla
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')])
            ])
            
            # Colorear columna de media días según valor
            for i, (_, _, media_dias) in enumerate(resultados, start=1):
                if media_dias and media_dias > 7:
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFC7CE'))
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#9C0006'))
                elif media_dias and media_dias > 3:
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFEB9C'))
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#9C6500'))
                else:
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.HexColor('#C6EFCE'))
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#006100'))
            
            table.setStyle(table_style)
            elements.append(table)
            
            # Estadísticas
            elements.append(Spacer(1, 0.5*cm))
            total_expedientes = sum(cantidad for _, cantidad, _ in resultados)
            stats_text = f"<b>Total clientes:</b> {len(resultados)}  |  <b>Total expedientes:</b> {total_expedientes}"
            elements.append(Paragraph(stats_text, styles['Normal']))
            
            # Generar PDF
            doc.build(elements)
            messagebox.showinfo("Éxito", f"PDF generado correctamente:\n{filename}")
            
            # Abrir PDF
            os.startfile(filename)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar PDF:\n{e}")
            print(f"Error detallado: {e}")

    def _imprimir_recepciones_anticipadas(self, cliente, expedientes):
        """Genera un PDF con el listado de recepciones anticipadas para imprimir."""
        from tkinter import filedialog
        import os
        
        if not expedientes:
            messagebox.showwarning("Imprimir", "No hay datos para imprimir.")
            return
        
        # Solicitar ubicación para guardar PDF
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"Recepciones_Anticipadas_{cliente.replace(' ', '_')}.pdf"
        )
        
        if not filename:
            return
        
        try:
            # Importar reportlab
            try:
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
            except ImportError:
                messagebox.showerror("Error", "ReportLab no está instalado. Instale reportlab:\npip install reportlab")
                return
            
            # Crear PDF
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(f"RECEPCIONES ANTICIPADAS - {cliente}", title_style))
            
            # Fecha
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(f"Generado el: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
            
            # Tabla de datos
            data = [['Código RMA', 'Fecha Recepción', 'Fecha Autorización', 'Días', 'Estado']]
            
            for exp_id, codigo_rma, fecha_recep, fecha_auto, dias, estado in expedientes:
                data.append([
                    codigo_rma,
                    fecha_recep or "-",
                    fecha_auto or "-",
                    str(int(dias)) if dias else "0",
                    estado or "-"
                ])
            
            # Crear tabla
            table = Table(data, colWidths=[4*cm, 4*cm, 4*cm, 2*cm, 4*cm])
            
            # Estilo de tabla
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')])
            ])
            
            # Colorear columna de días según valor
            for i, row in enumerate(expedientes, start=1):
                dias = row[4]
                if dias and dias > 7:
                    table_style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFC7CE'))
                    table_style.add('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#9C0006'))
                elif dias and dias > 3:
                    table_style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFEB9C'))
                    table_style.add('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#9C6500'))
                else:
                    table_style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#C6EFCE'))
                    table_style.add('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#006100'))
            
            table.setStyle(table_style)
            elements.append(table)
            
            # Estadísticas
            elements.append(Spacer(1, 0.5*cm))
            stats_text = f"<b>Total expedientes:</b> {len(expedientes)}"
            if expedientes:
                media_dias = sum(dias for _, _, _, _, dias, _ in expedientes if dias) / len([d for _, _, _, _, d, _ in expedientes if d])
                stats_text += f"  |  <b>Media días adelanto:</b> {media_dias:.1f}"
            
            elements.append(Paragraph(stats_text, styles['Normal']))
            
            # Generar PDF
            doc.build(elements)
            messagebox.showinfo("Éxito", f"PDF generado correctamente:\n{filename}")
            
            # Abrir PDF
            os.startfile(filename)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar PDF:\n{e}")
            print(f"Error detallado: {e}")
