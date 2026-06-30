"""
Módulo para estadísticas de expedientes completados por quincena
Permite filtrar expedientes cerrados por quincena y año, con exportación a Excel
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from lib.logger_config import get_logger

logger = get_logger()


class ExpedientesQuincenaWindow:
    """Ventana para mostrar expedientes completados por quincena"""
    
    def __init__(self, parent, conectar_db_func, username):
        """
        Args:
            parent: Ventana padre
            conectar_db_func: Función para conectar a la base de datos
            username: Usuario actual
        """
        self.parent = parent
        self.conectar_db = conectar_db_func
        self.username = username
        self.expedientes_data = []
        self.orden_actual = {"columna": "fecha_gestion", "ascendente": False}
        
        logger.info(f"Iniciando ventana de expedientes por quincena - Usuario: {username}")
        
        self.crear_interfaz()
        self.cargar_datos()
    
    def crear_interfaz(self):
        """Crea la interfaz de la ventana"""
        # Frame principal
        main_frame = ctk.CTkFrame(self.parent)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Título
        titulo_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        titulo_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(
            titulo_frame,
            text="📅 EXPEDIENTES COMPLETADOS POR QUINCENA",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(side="left")
        
        # Frame de filtros
        filtros_frame = ctk.CTkFrame(main_frame)
        filtros_frame.pack(fill="x", pady=(0, 10))
        
        # Año
        ctk.CTkLabel(filtros_frame, text="Año:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=0, padx=5, pady=5, sticky="e"
        )
        
        año_actual = datetime.datetime.now().year
        self.combo_año = ctk.CTkComboBox(
            filtros_frame,
            values=[str(año) for año in range(año_actual - 5, año_actual + 2)],
            width=100
        )
        self.combo_año.set(str(año_actual))
        self.combo_año.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        # Mes
        ctk.CTkLabel(filtros_frame, text="Mes:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=2, padx=5, pady=5, sticky="e"
        )
        
        meses = [
            "01 - Enero", "02 - Febrero", "03 - Marzo", "04 - Abril",
            "05 - Mayo", "06 - Junio", "07 - Julio", "08 - Agosto",
            "09 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
        ]
        
        mes_actual = datetime.datetime.now().month
        self.combo_mes = ctk.CTkComboBox(filtros_frame, values=meses, width=150)
        self.combo_mes.set(meses[mes_actual - 1])
        self.combo_mes.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        
        # Quincena
        ctk.CTkLabel(filtros_frame, text="Quincena:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=4, padx=5, pady=5, sticky="e"
        )
        
        self.combo_quincena = ctk.CTkComboBox(
            filtros_frame,
            values=["Todas", "Q1 (Primera)", "Q2 (Segunda)"],
            width=150
        )
        self.combo_quincena.set("Todas")
        self.combo_quincena.grid(row=0, column=5, padx=5, pady=5, sticky="w")
        
        # Botones
        btn_frame = ctk.CTkFrame(filtros_frame, fg_color="transparent")
        btn_frame.grid(row=0, column=6, padx=20, pady=5, sticky="e")
        
        ctk.CTkButton(
            btn_frame,
            text="🔍 Buscar",
            command=self.cargar_datos,
            width=100
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="📊 Exportar Excel",
            command=self.exportar_excel,
            width=120
        ).pack(side="left", padx=5)
        
        # Frame de resultados con scrollbar
        resultados_container = ctk.CTkFrame(main_frame)
        resultados_container.pack(fill="both", expand=True)
        
        # Scrollable frame
        self.resultados_scroll = ctk.CTkScrollableFrame(resultados_container)
        self.resultados_scroll.pack(fill="both", expand=True)
        
        # Frame para la tabla
        self.tabla_frame = ctk.CTkFrame(self.resultados_scroll)
        self.tabla_frame.pack(fill="both", expand=True, padx=5, pady=5)
    
    def obtener_rango_fechas(self):
        """
        Calcula el formato de quincena según los filtros seleccionados
        
        Returns:
            tuple: (patron_quincena, descripcion) para filtrar en SQL
        """
        try:
            año = int(self.combo_año.get())
            mes = int(self.combo_mes.get().split(" - ")[0])
            quincena = self.combo_quincena.get()
            
            # Formato año de 2 dígitos
            año_corto = str(año)[2:]
            mes_str = f"{mes:02d}"
            
            if quincena == "Q1 (Primera)":
                patron = f"Q1-{mes_str}-{año_corto}"
                descripcion = f"Primera quincena de {self.combo_mes.get().split(' - ')[1]} {año}"
            elif quincena == "Q2 (Segunda)":
                patron = f"Q2-{mes_str}-{año_corto}"
                descripcion = f"Segunda quincena de {self.combo_mes.get().split(' - ')[1]} {año}"
            else:  # Todas
                patron = f"Q_-{mes_str}-{año_corto}"  # Patrón para ambas quincenas
                descripcion = f"{self.combo_mes.get().split(' - ')[1]} {año} (ambas quincenas)"
            
            logger.info(f"Patrón de quincena calculado: {patron}")
            return patron, descripcion
            
        except Exception as e:
            logger.error(f"Error al calcular patrón de quincena: {e}")
            messagebox.showerror("Error", f"Error al calcular quincena: {e}")
            return None, None
    
    def cargar_datos(self):
        """Carga los expedientes completados según los filtros"""
        patron_quincena, descripcion = self.obtener_rango_fechas()
        if not patron_quincena:
            return
        
        conn, cursor = self.conectar_db()
        if not conn:
            logger.error("No se pudo conectar a la base de datos")
            return
        
        try:
            # Consulta para obtener expedientes con fecha_para_factura que coincida
            # Si es "Todas", usar LIKE con patrón Q_-MM-YY (Q1 o Q2)
            # IMPORTANTE: Calcula el total contabilizable excluyendo artículos con contabilizar=0
            if "Q_" in patron_quincena:
                # Buscar ambas quincenas: reemplazar Q_ por Q% para wildcard
                patron_sql = patron_quincena.replace("Q_", "Q%")
                query = """
                    SELECT
                        m.codigo_rma,
                        m.cliente,
                        m.numero_documento_cliente,
                        m.fecha_emision,
                        m.fecha_recepcion,
                        m.fecha_autorizacion,
                        m.fecha_proceso,
                        m.fecha_gestion,
                        m.precio_total_expediente,
                        COALESCE(
                            (SELECT SUM(d.precio_final * d.cantidad_entregada)
                             FROM rma_detalles d
                             WHERE d.rma_id = m.id AND COALESCE(d.contabilizar, 1) = 1),
                            0
                        ) as total_contabilizable,
                        m.resultado_expediente,
                        m.fecha_para_factura,
                        COALESCE(m.numero_albaran_reposicion, '') as numero_albaran_reposicion,
                        COALESCE(m.numero_factura_abono, '') as numero_factura_abono,
                        m.id,
                        COALESCE(m.fecha_entregado_contabilidad, '') as fecha_entregado_contabilidad
                    FROM rma_maestro m
                    WHERE m.fecha_para_factura IS NOT NULL
                    AND m.fecha_para_factura != ''
                    AND m.fecha_para_factura != 'Seleccionar...'
                    AND m.fecha_para_factura LIKE ?
                    ORDER BY m.fecha_para_factura DESC, m.codigo_rma DESC
                """
                cursor.execute(query, (patron_sql,))
            else:
                # Buscar quincena específica
                query = """
                    SELECT
                        m.codigo_rma,
                        m.cliente,
                        m.numero_documento_cliente,
                        m.fecha_emision,
                        m.fecha_recepcion,
                        m.fecha_autorizacion,
                        m.fecha_proceso,
                        m.fecha_gestion,
                        m.precio_total_expediente,
                        COALESCE(
                            (SELECT SUM(d.precio_final * d.cantidad_entregada)
                             FROM rma_detalles d
                             WHERE d.rma_id = m.id AND COALESCE(d.contabilizar, 1) = 1),
                            0
                        ) as total_contabilizable,
                        m.resultado_expediente,
                        m.fecha_para_factura,
                        COALESCE(m.numero_albaran_reposicion, '') as numero_albaran_reposicion,
                        COALESCE(m.numero_factura_abono, '') as numero_factura_abono,
                        m.id,
                        COALESCE(m.fecha_entregado_contabilidad, '') as fecha_entregado_contabilidad
                    FROM rma_maestro m
                    WHERE m.fecha_para_factura = ?
                    ORDER BY m.codigo_rma DESC
                """
                cursor.execute(query, (patron_quincena,))
            
            self.expedientes_data = cursor.fetchall()
            
            conn.close()
            
            logger.info(f"Cargados {len(self.expedientes_data)} expedientes para {descripcion}")
            self.mostrar_tabla()
            
        except Exception as e:
            logger.error(f"Error al cargar expedientes: {e}")
            messagebox.showerror("Error", f"Error al cargar datos: {e}")
            if conn:
                conn.close()
    
    def ordenar_por(self, columna, indice):
        """
        Ordena los datos por la columna especificada
        
        Args:
            columna: Nombre de la columna
            indice: Índice de la columna en los datos
        """
        if self.orden_actual["columna"] == columna:
            # Invertir orden si es la misma columna
            self.orden_actual["ascendente"] = not self.orden_actual["ascendente"]
        else:
            # Nueva columna, orden ascendente por defecto
            self.orden_actual["columna"] = columna
            self.orden_actual["ascendente"] = True
        
        # Ordenar datos
        self.expedientes_data.sort(
            key=lambda x: x[indice] if x[indice] else "",
            reverse=not self.orden_actual["ascendente"]
        )
        
        logger.info(f"Datos ordenados por {columna} - Ascendente: {self.orden_actual['ascendente']}")
        self.mostrar_tabla()
    
    def mostrar_tabla(self):
        """Muestra la tabla con los expedientes"""
        # Limpiar tabla anterior
        for widget in self.tabla_frame.winfo_children():
            widget.destroy()
        
        if not self.expedientes_data:
            ctk.CTkLabel(
                self.tabla_frame,
                text="❌ No se encontraron expedientes en la quincena seleccionada",
                font=ctk.CTkFont(size=14),
                text_color="orange"
            ).pack(pady=20)
            return
        
        # Info de resultados
        info_frame = ctk.CTkFrame(self.tabla_frame, fg_color="transparent")
        info_frame.pack(fill="x", pady=(0, 10))
        
        total_expedientes = len(self.expedientes_data)
        # Usar total_contabilizable (índice 9) en lugar de precio_total_expediente (índice 8)
        total_importe = sum(float(exp[9]) if exp[9] else 0.0 for exp in self.expedientes_data)
        
        ctk.CTkLabel(
            info_frame,
            text=f"📋 Total: {total_expedientes} expedientes | 💰 Importe contabilizable: {total_importe:,.2f} €",
            font=ctk.CTkFont(size=13, weight="bold")
        ).pack(side="left")
        
        # Cabeceras - estructura: 0=codigo_rma, 1=cliente, 2=num_doc, 3-7=fechas,
        # 8=precio_total, 9=total_contabilizable, 10=resultado, 11=quincena,
        # 12=albaran_reposicion, 13=factura_abono, 14=id, 15=fecha_entregado_contabilidad
        columnas = [
            ("RMA", 0, 120),
            ("Cliente", 1, 200),
            ("Nº Doc.", 2, 120),
            ("F. Emisión", 3, 100),
            ("F. Recepción", 4, 100),
            ("F. Autorización", 5, 110),
            ("F. Proceso", 6, 100),
            ("F. Gestión", 7, 100),
            ("Importe", 9, 100),
            ("Resultado", 10, 150),
            ("Quincena", 11, 100)
        ]

        header_frame = ctk.CTkFrame(self.tabla_frame)
        header_frame.pack(fill="x", pady=(0, 5))

        # Cabecera del checkbox de contabilidad (sin texto, solo reserva espacio)
        ctk.CTkLabel(header_frame, text="✓", width=26,
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=(2, 1))

        for col_nombre, col_idx, col_ancho in columnas:
            btn = ctk.CTkButton(
                header_frame,
                text=col_nombre,
                command=lambda c=col_nombre, i=col_idx: self.ordenar_por(c, i),
                width=col_ancho,
                height=30
            )
            btn.pack(side="left", padx=1)

            # Indicador de orden
            if self.orden_actual["columna"] == col_nombre:
                indicador = "▼" if not self.orden_actual["ascendente"] else "▲"
                btn.configure(text=f"{col_nombre} {indicador}")

        # Datos
        for exp in self.expedientes_data:
            fila_frame = ctk.CTkFrame(self.tabla_frame)
            fila_frame.pack(fill="x", pady=1)

            # Checkbox de entregado a contabilidad
            rma_id_exp = exp[14] if len(exp) > 14 else None
            fecha_ctb = exp[15] if len(exp) > 15 else None
            esta_marcado = bool(fecha_ctb and str(fecha_ctb).strip())
            check_var = ctk.BooleanVar(value=esta_marcado)
            check = ctk.CTkCheckBox(
                fila_frame, text="", variable=check_var,
                checkbox_width=16, checkbox_height=16, width=26,
                command=lambda cv=check_var, rid=rma_id_exp, fa=fecha_ctb:
                    self._toggle_contabilidad(cv, rid, fa)
            )
            check.pack(side="left", padx=(2, 1))

            for i, (_, col_idx, col_ancho) in enumerate(columnas):
                valor = exp[col_idx] if exp[col_idx] else "-"

                # Formatear importe (índice 9 = total_contabilizable)
                if col_idx == 9 and valor != "-":
                    try:
                        valor = f"{float(valor):,.2f} €"
                    except:
                        pass

                # Color para campos especiales (por defecto None = color automático)
                color = None
                if col_idx == 7:  # F. Gestión
                    if valor != "-":
                        color = "#10b981"  # Verde si está gestionado
                    else:
                        color = "#f59e0b"  # Naranja si está pendiente
                elif col_idx == 11:  # Quincena
                    color = "#3b82f6"

                label = ctk.CTkLabel(
                    fila_frame,
                    text=str(valor),
                    width=col_ancho,
                    anchor="w",
                    font=ctk.CTkFont(size=11)
                )
                if color:
                    label.configure(text_color=color)
                label.pack(side="left", padx=1)
    
    def _toggle_contabilidad(self, check_var, rma_id, fecha_anterior):
        """Marca/desmarca el expediente como entregado a contabilidad, actualiza BD e historial."""
        if rma_id is None:
            logger.warning("_toggle_contabilidad: rma_id es None, operación ignorada")
            return

        marcado = check_var.get()
        nueva_fecha = datetime.date.today().isoformat() if marcado else None

        try:
            result = self.conectar_db()
            if isinstance(result, tuple):
                conn, cursor = result
            else:
                conn = result
                cursor = conn.cursor() if conn else None

            if not conn or not cursor:
                logger.error("_toggle_contabilidad: no se pudo conectar a la BD")
                return

            try:
                cursor.execute(
                    "UPDATE rma_maestro SET fecha_entregado_contabilidad = ? WHERE id = ?",
                    (nueva_fecha, rma_id)
                )

                # Registrar en historial con la misma estructura que el resto del sistema
                valor_anterior_str = str(fecha_anterior).strip() if fecha_anterior and str(fecha_anterior).strip() else "—"
                valor_nuevo_str = nueva_fecha if nueva_fecha else "—"
                descripcion = (
                    f"Campo 'Entregado a Contabilidad' modificado: "
                    f"'{valor_anterior_str}' -> '{valor_nuevo_str}'"
                )
                cursor.execute(
                    "INSERT INTO rma_historial (rma_id, fecha_cambio, usuario, descripcion_cambio) "
                    "VALUES (?, ?, ?, ?)",
                    (rma_id, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     self.username, descripcion)
                )

                conn.commit()
                logger.info(
                    f"Contabilidad {'marcada' if marcado else 'desmarcada'} "
                    f"para rma_id={rma_id}: {valor_anterior_str} -> {valor_nuevo_str}"
                )
            finally:
                conn.close()

            # Recargar la tabla para reflejar el estado actualizado
            self.cargar_datos()

        except Exception as e:
            logger.error(f"Error al actualizar entregado a contabilidad (rma_id={rma_id}): {e}")
            messagebox.showerror("Error", f"No se pudo actualizar el estado de contabilidad:\n{e}")

    def exportar_excel(self):
        """Exporta los datos a Excel. Cada expediente aparece en negrita con sus artículos indentados debajo."""
        if not self.expedientes_data:
            messagebox.showwarning("Sin datos", "No hay datos para exportar")
            return

        patron_quincena, descripcion = self.obtener_rango_fechas()
        nombre_sugerido = f"Expedientes_{patron_quincena}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_sugerido
        )

        if not filepath:
            return

        try:
            logger.info(f"Exportando {len(self.expedientes_data)} expedientes a Excel: {filepath}")

            wb = Workbook()
            ws = wb.active
            ws.title = "Expedientes Quincena"

            # --- Estilos ---
            header_font      = Font(bold=True, size=11, color="FFFFFF")
            exp_font         = Font(bold=True, size=10)
            art_font         = Font(size=10, italic=True)
            header_fill      = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            art_fill         = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            header_align     = Alignment(horizontal="center", vertical="center")
            left_align       = Alignment(horizontal="left",   vertical="center")

            border_side      = Side(style='thin', color="BBBBBB")
            border           = Border(left=border_side, right=border_side,
                                      top=border_side, bottom=border_side)

            # --- Título ---
            # 11 columnas originales + 1 columna nueva: Alb.Repos./Fact.Abono
            num_cols = 12
            ws.merge_cells(f'A1:{chr(64 + num_cols)}1')
            c = ws['A1']
            c.value = f"EXPEDIENTES - {descripcion.upper()}"
            c.font = Font(bold=True, size=14)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # --- Resumen ---
            ws.merge_cells(f'A2:{chr(64 + num_cols)}2')
            total_importe = sum(float(exp[9]) if exp[9] else 0.0 for exp in self.expedientes_data)
            c = ws['A2']
            c.value = (f"Total: {len(self.expedientes_data)} expedientes | "
                       f"Importe contabilizable: {total_importe:,.2f} € "
                       f"(excluye artículos no contabilizables)")
            c.alignment = Alignment(horizontal="center")

            # --- Cabeceras (mismas 11 columnas que la vista) ---
            headers = [
                "RMA", "Cliente", "Nº Documento",
                "F. Emisión", "F. Recepción", "F. Autorización",
                "F. Proceso", "F. Gestión", "Importe", "Resultado", "Quincena",
                "Alb. Repos. / Fact. Abono"
            ]
            for col_i, header in enumerate(headers, start=1):
                c = ws.cell(row=4, column=col_i)
                c.value = header
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = border

            # --- Cargar artículos desde BD ---
            # conectar_db devuelve (conn, cursor) — desempaquetar correctamente
            codigos_a_id    = {}
            articulos_por_exp = {}
            try:
                result = self.conectar_db()
                if isinstance(result, tuple):
                    conn_art, cur_art = result
                else:
                    conn_art = result
                    cur_art  = conn_art.cursor()

                if conn_art and cur_art:
                    codigos_rma = [exp[0] for exp in self.expedientes_data if exp[0]]
                    if codigos_rma:
                        ph = ",".join(["?" for _ in codigos_rma])
                        cur_art.execute(
                            f"SELECT id, codigo_rma FROM rma_maestro WHERE codigo_rma IN ({ph})",
                            codigos_rma
                        )
                        for row in cur_art.fetchall():
                            codigos_a_id[row[1]] = row[0]

                    if codigos_a_id:
                        ids_list = list(codigos_a_id.values())
                        ph2 = ",".join(["?" for _ in ids_list])
                        # SELECT * para no fallar si numero_albaran/numero_order aún no existen en Turso
                        cur_art.execute(
                            f"SELECT * FROM rma_detalles WHERE rma_id IN ({ph2}) ORDER BY rma_id",
                            ids_list
                        )
                        cols_art = [d[0] for d in cur_art.description]
                        for raw_row in cur_art.fetchall():
                            art_dict = dict(zip(cols_art, raw_row))
                            art_dict.setdefault('numero_albaran', '')
                            art_dict.setdefault('numero_order', '')
                            art_dict.setdefault('estado_producto', '')
                            art_dict.setdefault('precio_final', 0)
                            art_dict.setdefault('cantidad_entregada', 0)
                            if art_dict['numero_albaran'] is None:
                                art_dict['numero_albaran'] = ''
                            rid = art_dict['rma_id']
                            if rid not in articulos_por_exp:
                                articulos_por_exp[rid] = []
                            articulos_por_exp[rid].append(art_dict)

                    conn_art.close()
            except Exception as e:
                logger.warning(f"No se pudieron cargar artículos para export: {e}")

            # --- Mapeo índices del expediente → columnas Excel ---
            # exp: 0=codigo_rma, 1=cliente, 2=num_doc, 3-7=fechas,
            #      8=precio_total, 9=total_contabilizable, 10=resultado, 11=quincena
            # índices: 0-7 datos, 9=importe contabilizable, 10=resultado, 11=quincena
            # 12=albaran_reposicion, 13=factura_abono → se combinan en la col 12 del Excel
            columnas_indices = [0, 1, 2, 3, 4, 5, 6, 7, 9, 10, 11]

            # --- Escribir filas ---
            row_idx = 5
            for exp in self.expedientes_data:
                codigo_rma = exp[0]

                # Fila del expediente (negrita, sin relleno)
                for excel_col, data_col in enumerate(columnas_indices, start=1):
                    c = ws.cell(row=row_idx, column=excel_col)
                    valor = exp[data_col] if exp[data_col] else "-"
                    if data_col == 9 and valor != "-":
                        try:
                            c.value = float(valor)
                            c.number_format = '#,##0.00 "€"'
                        except Exception:
                            c.value = valor
                    else:
                        c.value = valor
                    c.font   = exp_font
                    c.border = border
                    c.alignment = left_align

                # Columna 12: Nº Albarán Reposición / Nº Factura Abono (combinados)
                alb_repos = str(exp[12]).strip() if len(exp) > 12 and exp[12] else ''
                fact_abono = str(exp[13]).strip() if len(exp) > 13 and exp[13] else ''
                if alb_repos and fact_abono:
                    valor_combinado = f"{alb_repos} - {fact_abono}"
                elif alb_repos:
                    valor_combinado = alb_repos
                elif fact_abono:
                    valor_combinado = fact_abono
                else:
                    valor_combinado = ""
                c12 = ws.cell(row=row_idx, column=12)
                c12.value = valor_combinado
                c12.font = exp_font
                c12.border = border
                c12.alignment = left_align

                row_idx += 1

                # Filas de artículos indentadas (mismas columnas, fondo gris)
                rma_id_exp = codigos_a_id.get(codigo_rma)
                if rma_id_exp and rma_id_exp in articulos_por_exp:
                    for art in articulos_por_exp[rma_id_exp]:
                        # Saltar artículos no contabilizables
                        try:
                            if int(art.get('contabilizar', 1) if art.get('contabilizar') is not None else 1) == 0:
                                continue
                        except (ValueError, TypeError):
                            pass

                        ref_art     = art.get('referencia_articulo', '') or ''
                        num_albaran = art.get('numero_albaran', '') or ''
                        estado_art  = art.get('estado_producto', '') or ''
                        try:
                            importe_art = float(art.get('precio_final', 0) or 0) * float(art.get('cantidad_entregada', 0) or 0)
                        except Exception:
                            importe_art = 0.0

                        # Rellenar columnas 1-12 con fondo gris
                        for c_i in range(1, num_cols + 1):
                            c = ws.cell(row=row_idx, column=c_i)
                            c.fill   = art_fill
                            c.border = border
                            c.font   = art_font
                            c.alignment = left_align

                        # Col 1 → Referencia con indentación visual
                        ws.cell(row=row_idx, column=1).value = f"    ↳ {ref_art or ''}"
                        # Col 2 → Nº Albarán
                        ws.cell(row=row_idx, column=2).value = f"Alb: {num_albaran}" if num_albaran else ""
                        # Col 4 → Estado artículo
                        ws.cell(row=row_idx, column=4).value = estado_art or ""
                        # Col 6 → Precio unitario (reutiliza "F. Proceso", vacía en artículos)
                        try:
                            precio_unit = float(art.get('precio_unitario', 0) or 0)
                        except Exception:
                            precio_unit = 0.0
                        c_pu = ws.cell(row=row_idx, column=6)
                        c_pu.value = precio_unit
                        c_pu.number_format = '#,##0.00 "€"'
                        c_pu.fill = art_fill
                        c_pu.font = art_font
                        c_pu.border = border
                        c_pu.alignment = left_align
                        # Col 7 → Unidades (reutiliza "F. Gestión", vacía en artículos)
                        try:
                            unidades = float(art.get('cantidad_entregada', 0) or 0)
                        except Exception:
                            unidades = 0.0
                        c_ud = ws.cell(row=row_idx, column=7)
                        c_ud.value = unidades
                        c_ud.fill = art_fill
                        c_ud.font = art_font
                        c_ud.border = border
                        c_ud.alignment = left_align
                        # Col 9 → Importe artículo
                        c_imp = ws.cell(row=row_idx, column=9)
                        c_imp.value  = importe_art
                        c_imp.number_format = '#,##0.00 "€"'
                        c_imp.fill = art_fill
                        c_imp.font = art_font
                        c_imp.border = border
                        c_imp.alignment = left_align
                        row_idx += 1

            # --- Ajustar anchos de columna ---
            anchos = [15, 30, 15, 12, 12, 14, 12, 12, 12, 20, 12, 28]
            for col_i, ancho in enumerate(anchos, start=1):
                ws.column_dimensions[chr(64 + col_i)].width = ancho

            # --- Guardar ---
            wb.save(filepath)
            logger.info(f"Excel exportado exitosamente: {filepath}")
            messagebox.showinfo("Éxito", f"Archivo exportado correctamente:\n{filepath}")

        except Exception as e:
            logger.error(f"Error al exportar a Excel: {e}")
            messagebox.showerror("Error", f"Error al exportar: {e}")
def mostrar_expedientes_quincena(parent_frame, conectar_db_func, username):
    """
    Función principal para mostrar la ventana de expedientes por quincena
    
    Args:
        parent_frame: Frame padre donde se mostrará el contenido
        conectar_db_func: Función para conectar a la base de datos
        username: Usuario actual
    """
    # Limpiar frame padre
    for widget in parent_frame.winfo_children():
        widget.destroy()
    
    # Crear ventana
    ExpedientesQuincenaWindow(parent_frame, conectar_db_func, username)
