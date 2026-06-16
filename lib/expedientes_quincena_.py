"""
Módulo para estadísticas de expedientes completados por quincena
Permite filtrar expedientes cerrados por quincena y año, con exportación a Excel
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from lib.logger_config import get_logger

logger = get_logger()


class ExpedientesQuincenaWindow:
    """Ventana para mostrar expedientes completados por quincena"""
    
    def __init__(self, parent, conectar_db_func, username):
        """
        Args:
            parent: Ventana padre
            conectar_db_func: Función para conectar a la base de datos
            username: Usuario actual
        """
        self.parent = parent
        self.conectar_db = conectar_db_func
        self.username = username
        self.expedientes_data = []
        self.orden_actual = {"columna": "fecha_gestion", "ascendente": False}
        
        logger.info(f"Iniciando ventana de expedientes por quincena - Usuario: {username}")
        
        self.crear_interfaz()
        self.cargar_datos()
    
    def crear_interfaz(self):
        """Crea la interfaz de la ventana"""
        # Frame principal
        main_frame = ctk.CTkFrame(self.parent)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Título
        titulo_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        titulo_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(
            titulo_frame,
            text="📅 EXPEDIENTES COMPLETADOS POR QUINCENA",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(side="left")
        
        # Frame de filtros
        filtros_frame = ctk.CTkFrame(main_frame)
        filtros_frame.pack(fill="x", pady=(0, 10))
        
        # Año
        ctk.CTkLabel(filtros_frame, text="Año:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=0, padx=5, pady=5, sticky="e"
        )
        
        año_actual = datetime.datetime.now().year
        self.combo_año = ctk.CTkComboBox(
            filtros_frame,
            values=[str(año) for año in range(año_actual - 5, año_actual + 2)],
            width=100
        )
        self.combo_año.set(str(año_actual))
        self.combo_año.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        # Mes
        ctk.CTkLabel(filtros_frame, text="Mes:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=2, padx=5, pady=5, sticky="e"
        )
        
        meses = [
            "01 - Enero", "02 - Febrero", "03 - Marzo", "04 - Abril",
            "05 - Mayo", "06 - Junio", "07 - Julio", "08 - Agosto",
            "09 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
        ]
        
        mes_actual = datetime.datetime.now().month
        self.combo_mes = ctk.CTkComboBox(filtros_frame, values=meses, width=150)
        self.combo_mes.set(meses[mes_actual - 1])
        self.combo_mes.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        
        # Quincena
        ctk.CTkLabel(filtros_frame, text="Quincena:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=4, padx=5, pady=5, sticky="e"
        )
        
        self.combo_quincena = ctk.CTkComboBox(
            filtros_frame,
            values=["Todas", "Q1 (Primera)", "Q2 (Segunda)"],
            width=150
        )
        self.combo_quincena.set("Todas")
        self.combo_quincena.grid(row=0, column=5, padx=5, pady=5, sticky="w")
        
        # Botones
        btn_frame = ctk.CTkFrame(filtros_frame, fg_color="transparent")
        btn_frame.grid(row=0, column=6, padx=20, pady=5, sticky="e")
        
        ctk.CTkButton(
            btn_frame,
            text="🔍 Buscar",
            command=self.cargar_datos,
            width=100
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="📊 Exportar Excel",
            command=self.exportar_excel,
            width=120
        ).pack(side="left", padx=5)
        
        # Frame de resultados con scrollbar
        resultados_container = ctk.CTkFrame(main_frame)
        resultados_container.pack(fill="both", expand=True)
        
        # Scrollable frame
        self.resultados_scroll = ctk.CTkScrollableFrame(resultados_container)
        self.resultados_scroll.pack(fill="both", expand=True)
        
        # Frame para la tabla
        self.tabla_frame = ctk.CTkFrame(self.resultados_scroll)
        self.tabla_frame.pack(fill="both", expand=True, padx=5, pady=5)
    
    def obtener_rango_fechas(self):
        """
        Calcula el formato de quincena según los filtros seleccionados
        
        Returns:
            tuple: (patron_quincena, descripcion) para filtrar en SQL
        """
        try:
            año = int(self.combo_año.get())
            mes = int(self.combo_mes.get().split(" - ")[0])
            quincena = self.combo_quincena.get()
            
            # Formato año de 2 dígitos
            año_corto = str(año)[2:]
            mes_str = f"{mes:02d}"
            
            if quincena == "Q1 (Primera)":
                patron = f"Q1-{mes_str}-{año_corto}"
                descripcion = f"Primera quincena de {self.combo_mes.get().split(' - ')[1]} {año}"
            elif quincena == "Q2 (Segunda)":
                patron = f"Q2-{mes_str}-{año_corto}"
                descripcion = f"Segunda quincena de {self.combo_mes.get().split(' - ')[1]} {año}"
            else:  # Todas
                patron = f"Q_-{mes_str}-{año_corto}"  # Patrón para ambas quincenas
                descripcion = f"{self.combo_mes.get().split(' - ')[1]} {año} (ambas quincenas)"
            
            logger.info(f"Patrón de quincena calculado: {patron}")
            return patron, descripcion
            
        except Exception as e:
            logger.error(f"Error al calcular patrón de quincena: {e}")
            messagebox.showerror("Error", f"Error al calcular quincena: {e}")
            return None, None
    
    def cargar_datos(self):
        """Carga los expedientes completados según los filtros"""
        patron_quincena, descripcion = self.obtener_rango_fechas()
        if not patron_quincena:
            return
        
        conn, cursor = self.conectar_db()
        if not conn:
            logger.error("No se pudo conectar a la base de datos")
            return
        
        try:
            # Consulta para obtener expedientes con fecha_para_factura que coincida
            # Si es "Todas", usar LIKE con patrón Q_-MM-YY (Q1 o Q2)
            # IMPORTANTE: Calcula el total contabilizable excluyendo artículos con contabilizar=0
            if "Q_" in patron_quincena:
                # Buscar ambas quincenas: reemplazar Q_ por Q% para wildcard
                patron_sql = patron_quincena.replace("Q_", "Q%")
                query = """
                    SELECT 
                        m.codigo_rma,
                        m.cliente,
                        m.numero_documento_cliente,
                        m.fecha_emision,
                        m.fecha_recepcion,
                        m.fecha_autorizacion,
                        m.fecha_proceso,
                        m.fecha_gestion,
                        m.precio_total_expediente,
                        COALESCE(
                            (SELECT SUM(d.precio_final * d.cantidad_entregada)
                             FROM rma_detalles d
                             WHERE d.rma_id = m.id AND COALESCE(d.contabilizar, 1) = 1),
                            0
                        ) as total_contabilizable,
                        m.resultado_expediente,
                        m.fecha_para_factura
                    FROM rma_maestro m
                    WHERE m.fecha_para_factura IS NOT NULL 
                    AND m.fecha_para_factura != ''
                    AND m.fecha_para_factura != 'Seleccionar...'
                    AND m.fecha_para_factura LIKE ?
                    ORDER BY m.fecha_para_factura DESC, m.codigo_rma DESC
                """
                cursor.execute(query, (patron_sql,))
            else:
                # Buscar quincena específica
                query = """
                    SELECT 
                        m.codigo_rma,
                        m.cliente,
                        m.numero_documento_cliente,
                        m.fecha_emision,
                        m.fecha_recepcion,
                        m.fecha_autorizacion,
                        m.fecha_proceso,
                        m.fecha_gestion,
                        m.precio_total_expediente,
                        COALESCE(
                            (SELECT SUM(d.precio_final * d.cantidad_entregada)
                             FROM rma_detalles d
                             WHERE d.rma_id = m.id AND COALESCE(d.contabilizar, 1) = 1),
                            0
                        ) as total_contabilizable,
                        m.resultado_expediente,
                        m.fecha_para_factura
                    FROM rma_maestro m
                    WHERE m.fecha_para_factura = ?
                    ORDER BY m.codigo_rma DESC
                """
                cursor.execute(query, (patron_quincena,))
            
            self.expedientes_data = cursor.fetchall()
            
            conn.close()
            
            logger.info(f"Cargados {len(self.expedientes_data)} expedientes para {descripcion}")
            self.mostrar_tabla()
            
        except Exception as e:
            logger.error(f"Error al cargar expedientes: {e}")
            messagebox.showerror("Error", f"Error al cargar datos: {e}")
            if conn:
                conn.close()
    
    def ordenar_por(self, columna, indice):
        """
        Ordena los datos por la columna especificada
        
        Args:
            columna: Nombre de la columna
            indice: Índice de la columna en los datos
        """
        if self.orden_actual["columna"] == columna:
            # Invertir orden si es la misma columna
            self.orden_actual["ascendente"] = not self.orden_actual["ascendente"]
        else:
            # Nueva columna, orden ascendente por defecto
            self.orden_actual["columna"] = columna
            self.orden_actual["ascendente"] = True
        
        # Ordenar datos
        self.expedientes_data.sort(
            key=lambda x: x[indice] if x[indice] else "",
            reverse=not self.orden_actual["ascendente"]
        )
        
        logger.info(f"Datos ordenados por {columna} - Ascendente: {self.orden_actual['ascendente']}")
        self.mostrar_tabla()
    
    def mostrar_tabla(self):
        """Muestra la tabla con los expedientes"""
        # Limpiar tabla anterior
        for widget in self.tabla_frame.winfo_children():
            widget.destroy()
        
        if not self.expedientes_data:
            ctk.CTkLabel(
                self.tabla_frame,
                text="❌ No se encontraron expedientes en la quincena seleccionada",
                font=ctk.CTkFont(size=14),
                text_color="orange"
            ).pack(pady=20)
            return
        
        # Info de resultados
        info_frame = ctk.CTkFrame(self.tabla_frame, fg_color="transparent")
        info_frame.pack(fill="x", pady=(0, 10))
        
        total_expedientes = len(self.expedientes_data)
        # Usar total_contabilizable (índice 9) en lugar de precio_total_expediente (índice 8)
        total_importe = sum(float(exp[9]) if exp[9] else 0.0 for exp in self.expedientes_data)
        
        ctk.CTkLabel(
            info_frame,
            text=f"📋 Total: {total_expedientes} expedientes | 💰 Importe contabilizable: {total_importe:,.2f} €",
            font=ctk.CTkFont(size=13, weight="bold")
        ).pack(side="left")
        
        # Cabeceras - estructura: 0=codigo_rma, 1=cliente, 2=num_doc, 3-7=fechas, 
        # 8=precio_total, 9=total_contabilizable, 10=resultado, 11=quincena
        columnas = [
            ("RMA", 0, 120),
            ("Cliente", 1, 200),
            ("Nº Doc.", 2, 120),
            ("F. Emisión", 3, 100),
            ("F. Recepción", 4, 100),
            ("F. Autorización", 5, 110),
            ("F. Proceso", 6, 100),
            ("F. Gestión", 7, 100),
            ("Importe", 9, 100),  # Cambio a índice 9 (total_contabilizable)
            ("Resultado", 10, 150),  # Cambio a índice 10
            ("Quincena", 11, 100)  # Cambio a índice 11
        ]
        
        header_frame = ctk.CTkFrame(self.tabla_frame)
        header_frame.pack(fill="x", pady=(0, 5))
        
        for col_nombre, col_idx, col_ancho in columnas:
            btn = ctk.CTkButton(
                header_frame,
                text=col_nombre,
                command=lambda c=col_nombre, i=col_idx: self.ordenar_por(c, i),
                width=col_ancho,
                height=30
            )
            btn.pack(side="left", padx=1)
            
            # Indicador de orden
            if self.orden_actual["columna"] == col_nombre:
                indicador = "▼" if not self.orden_actual["ascendente"] else "▲"
                btn.configure(text=f"{col_nombre} {indicador}")
        
        # Datos
        for exp in self.expedientes_data:
            fila_frame = ctk.CTkFrame(self.tabla_frame)
            fila_frame.pack(fill="x", pady=1)
            
            for i, (_, col_idx, col_ancho) in enumerate(columnas):
                valor = exp[col_idx] if exp[col_idx] else "-"
                
                # Formatear importe (índice 9 = total_contabilizable)
                if col_idx == 9 and valor != "-":
                    try:
                        valor = f"{float(valor):,.2f} €"
                    except:
                        pass
                
                # Color para campos especiales (por defecto None = color automático)
                color = None
                if col_idx == 7:  # F. Gestión
                    if valor != "-":
                        color = "#10b981"  # Verde si está gestionado
                    else:
                        color = "#f59e0b"  # Naranja si está pendiente
                elif col_idx == 11:  # Quincena (ahora en índice 11)
                    color = "#3b82f6"
                
                label = ctk.CTkLabel(
                    fila_frame,
                    text=str(valor),
                    width=col_ancho,
                    anchor="w",
                    font=ctk.CTkFont(size=11)
                )
                if color:
                    label.configure(text_color=color)
                label.pack(side="left", padx=1)
    
    def exportar_excel(self):
        """Exporta los datos a Excel con formato"""
        if not self.expedientes_data:
            messagebox.showwarning("Sin datos", "No hay datos para exportar")
            return
        
        # Diálogo para guardar archivo
        patron_quincena, descripcion = self.obtener_rango_fechas()
        nombre_sugerido = f"Expedientes_{patron_quincena}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_sugerido
        )
        
        if not filepath:
            return
        
        try:
            logger.info(f"Exportando {len(self.expedientes_data)} expedientes a Excel: {filepath}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Expedientes Quincena"
            
            # Estilos
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            border_side = Side(style='thin', color="000000")
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            
            # Título
            ws.merge_cells('A1:L1')
            cell = ws['A1']
            cell.value = f"EXPEDIENTES - {descripcion.upper()}"
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Info general - Usar total_contabilizable (índice 9)
            ws.merge_cells('A2:K2')
            total_importe = sum(float(exp[9]) if exp[9] else 0.0 for exp in self.expedientes_data)
            cell = ws['A2']
            cell.value = f"Total: {len(self.expedientes_data)} expedientes | Importe contabilizable: {total_importe:,.2f} € (excluye artículos no contabilizables)"
            cell.alignment = Alignment(horizontal="center")
            
            # Cabeceras
            headers = [
                "RMA", "Cliente", "Nº Documento", "F. Emisión", "F. Recepción",
                "F. Autorización", "F. Proceso", "F. Gestión", "Importe", "Resultado", "Quincena"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border
            
            # Datos - Mapeo de índices (excluir precio_total_expediente, usar total_contabilizable)
            # Columnas Excel: RMA(0), Cliente(1), Doc(2), Fechas(3-7), Importe(9), Resultado(10), Quincena(11)
            columnas_indices = [0, 1, 2, 3, 4, 5, 6, 7, 9, 10, 11]  # Saltar índice 8 (precio_total_expediente)
            
            for row_idx, exp in enumerate(self.expedientes_data, start=5):
                for excel_col_idx, data_col_idx in enumerate(columnas_indices, start=1):
                    cell = ws.cell(row=row_idx, column=excel_col_idx)
                    valor = exp[data_col_idx] if exp[data_col_idx] else "-"
                    
                    # Formatear importe (cuando data_col_idx == 9, es el total_contabilizable)
                    if data_col_idx == 9 and valor != "-":
                        try:
                            cell.value = float(valor)
                            cell.number_format = '#,##0.00 "€"'
                        except:
                            cell.value = valor
                    else:
                        cell.value = valor
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajustar anchos de columna
            anchos = [15, 30, 15, 12, 12, 14, 12, 12, 12, 20, 12]
            for col, ancho in enumerate(anchos, start=1):
                ws.column_dimensions[chr(64 + col)].width = ancho
            
            # Guardar
            wb.save(filepath)
            logger.info(f"Excel exportado exitosamente: {filepath}")
            messagebox.showinfo("Éxito", f"Archivo exportado correctamente:\n{filepath}")
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {e}")
            messagebox.showerror("Error", f"Error al exportar: {e}")


def mostrar_expedientes_quincena(parent_frame, conectar_db_func, username):
    """
    Función principal para mostrar la ventana de expedientes por quincena
    
    Args:
        parent_frame: Frame padre donde se mostrará el contenido
        conectar_db_func: Función para conectar a la base de datos
        username: Usuario actual
    """
    # Limpiar frame padre
    for widget in parent_frame.winfo_children():
        widget.destroy()
    
    # Crear ventana
    ExpedientesQuincenaWindow(parent_frame, conectar_db_func, username)
