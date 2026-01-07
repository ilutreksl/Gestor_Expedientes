"""
Módulo para estadísticas anuales de expedientes
Muestra métricas generales del año: expedientes creados, cerrados, importes, productos y clientes con más incidencias
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
import sqlite3
from lib.logger_config import get_logger

logger = get_logger()


# Estados que se consideran como "mal estado"
ESTADOS_MAL_ESTADO = [
    "NO FUNCIONA; ABONAR",
    "NO FUNCIONA; NO ABONAR",
    "REPOSICION FALLO PRODUCTO",
    "FALLO SOLDADURA; ABONAR",
    "FALLO SOLDADURA; NO ABONAR",
    "FALLO MODULO; ABONAR"
]


def obtener_estadisticas_anuales(conn, año):
    """
    Calcula todas las estadísticas del año especificado
    
    Args:
        conn: Conexión a la base de datos
        año: Año a analizar
        
    Returns:
        dict: Diccionario con todas las estadísticas
    """
    cursor = conn.cursor()
    
    # Rango de fechas del año
    fecha_inicio = f"{año}-01-01"
    fecha_fin = f"{año}-12-31"
    
    stats = {}
    
    try:
        # 1. Total expedientes creados en el año
        cursor.execute("""
            SELECT COUNT(*) 
            FROM rma_maestro 
            WHERE fecha_emision BETWEEN ? AND ?
        """, (fecha_inicio, fecha_fin))
        stats['total_creados'] = int(cursor.fetchone()[0] or 0)
        
        # 2. Expedientes abiertos (tienen fecha_emision pero NO tienen fecha_gestion)
        cursor.execute("""
            SELECT COUNT(*) 
            FROM rma_maestro 
            WHERE fecha_emision BETWEEN ? AND ?
            AND fecha_emision IS NOT NULL 
            AND fecha_emision != ''
            AND (fecha_gestion IS NULL OR fecha_gestion = '')
        """, (fecha_inicio, fecha_fin))
        stats['abiertos'] = int(cursor.fetchone()[0] or 0)
        
        # 3. Expedientes pendientes de autorizar (tienen fecha_emision pero NO tienen fecha_autorizacion)
        cursor.execute("""
            SELECT COUNT(*) 
            FROM rma_maestro 
            WHERE fecha_emision BETWEEN ? AND ?
            AND fecha_emision IS NOT NULL 
            AND fecha_emision != ''
            AND (fecha_autorizacion IS NULL OR fecha_autorizacion = '')
        """, (fecha_inicio, fecha_fin))
        stats['pendientes_autorizar'] = int(cursor.fetchone()[0] or 0)
        
        # 4. Expedientes cerrados (tienen fecha_gestion)
        cursor.execute("""
            SELECT COUNT(*) 
            FROM rma_maestro 
            WHERE fecha_emision BETWEEN ? AND ?
            AND fecha_gestion IS NOT NULL 
            AND fecha_gestion != ''
        """, (fecha_inicio, fecha_fin))
        stats['cerrados'] = int(cursor.fetchone()[0] or 0)
        
        # 5. Total euros en expedientes cerrados del año
        cursor.execute("""
            SELECT COALESCE(SUM(CAST(precio_total_expediente AS REAL)), 0)
            FROM rma_maestro 
            WHERE fecha_emision BETWEEN ? AND ?
            AND fecha_gestion IS NOT NULL 
            AND fecha_gestion != ''
            AND precio_total_expediente IS NOT NULL
            AND precio_total_expediente != ''
        """, (fecha_inicio, fecha_fin))
        stats['total_euros_cerrados'] = float(cursor.fetchone()[0] or 0.0)
        
        # 6. Total euros en productos en "mal estado"
        placeholders = ','.join('?' * len(ESTADOS_MAL_ESTADO))
        cursor.execute(f"""
            SELECT COALESCE(SUM(CAST(rma_detalles.precio_unitario AS REAL)), 0)
            FROM rma_detalles
            INNER JOIN rma_maestro ON rma_detalles.rma_id = rma_maestro.id
            WHERE rma_maestro.fecha_emision BETWEEN ? AND ?
            AND rma_detalles.estado_producto IN ({placeholders})
            AND rma_detalles.precio_unitario IS NOT NULL
            AND rma_detalles.precio_unitario != ''
        """, (fecha_inicio, fecha_fin, *ESTADOS_MAL_ESTADO))
        stats['total_euros_mal_estado'] = float(cursor.fetchone()[0] or 0.0)
        
        # 7. Producto con más incidencias (referencia más frecuente)
        cursor.execute("""
            SELECT rma_detalles.referencia_articulo, COUNT(*) as cantidad
            FROM rma_detalles
            INNER JOIN rma_maestro ON rma_detalles.rma_id = rma_maestro.id
            WHERE rma_maestro.fecha_emision BETWEEN ? AND ?
            AND rma_detalles.referencia_articulo IS NOT NULL
            AND rma_detalles.referencia_articulo != ''
            GROUP BY rma_detalles.referencia_articulo
            ORDER BY cantidad DESC
            LIMIT 1
        """, (fecha_inicio, fecha_fin))
        producto_top = cursor.fetchone()
        if producto_top:
            stats['producto_top'] = producto_top[0]
            stats['producto_top_cantidad'] = int(producto_top[1])
        else:
            stats['producto_top'] = "N/A"
            stats['producto_top_cantidad'] = 0
        
        # 8. Cliente con más incidencias (cliente con más expedientes)
        cursor.execute("""
            SELECT cliente, COUNT(*) as cantidad
            FROM rma_maestro
            WHERE fecha_emision BETWEEN ? AND ?
            AND cliente IS NOT NULL
            AND cliente != ''
            GROUP BY cliente
            ORDER BY cantidad DESC
            LIMIT 1
        """, (fecha_inicio, fecha_fin))
        cliente_top = cursor.fetchone()
        if cliente_top:
            stats['cliente_top'] = cliente_top[0]
            stats['cliente_top_cantidad'] = int(cliente_top[1])
        else:
            stats['cliente_top'] = "N/A"
            stats['cliente_top_cantidad'] = 0
        
        # 9. Tasa de cierre (% de expedientes cerrados)
        if stats['total_creados'] > 0:
            stats['tasa_cierre'] = (stats['cerrados'] / stats['total_creados']) * 100
        else:
            stats['tasa_cierre'] = 0.0
        
        # 10. Tiempo promedio de tramitación (días entre fecha_emision y fecha_gestion)
        cursor.execute("""
            SELECT AVG(julianday(fecha_gestion) - julianday(fecha_emision)) as dias_promedio
            FROM rma_maestro
            WHERE fecha_emision BETWEEN ? AND ?
            AND fecha_emision IS NOT NULL AND fecha_emision != ''
            AND fecha_gestion IS NOT NULL AND fecha_gestion != ''
        """, (fecha_inicio, fecha_fin))
        dias_prom = cursor.fetchone()[0]
        stats['dias_tramitacion_promedio'] = round(dias_prom, 1) if dias_prom else 0.0
        
        # 11. Total de artículos procesados
        cursor.execute("""
            SELECT COUNT(*)
            FROM rma_detalles
            INNER JOIN rma_maestro ON rma_detalles.rma_id = rma_maestro.id
            WHERE rma_maestro.fecha_emision BETWEEN ? AND ?
        """, (fecha_inicio, fecha_fin))
        stats['total_articulos'] = int(cursor.fetchone()[0] or 0)
        
        # 12. Resultado más común en expedientes cerrados
        cursor.execute("""
            SELECT resultado_expediente, COUNT(*) as cantidad
            FROM rma_maestro
            WHERE fecha_emision BETWEEN ? AND ?
            AND fecha_gestion IS NOT NULL AND fecha_gestion != ''
            AND resultado_expediente IS NOT NULL AND resultado_expediente != ''
            GROUP BY resultado_expediente
            ORDER BY cantidad DESC
            LIMIT 1
        """, (fecha_inicio, fecha_fin))
        resultado_top = cursor.fetchone()
        if resultado_top:
            stats['resultado_top'] = resultado_top[0]
            stats['resultado_top_cantidad'] = int(resultado_top[1])
        else:
            stats['resultado_top'] = "N/A"
            stats['resultado_top_cantidad'] = 0
        
        return stats
        
    except Exception as e:
        print(f"Error al calcular estadísticas anuales: {e}")
        import traceback
        traceback.print_exc()
        return None


def mostrar_estadisticas_anuales(ventana_principal):
    """
    Muestra la interfaz de estadísticas anuales
    
    Args:
        ventana_principal: Instancia de VentanaPrincipal con acceso a main_stats_frame
    """
    # Limpiar el frame principal
    ventana_principal.limpiar_marco_stats()
    
    # Título
    ctk.CTkLabel(
        ventana_principal.main_stats_frame, 
        text="📊 ESTADÍSTICAS ANUALES - RESUMEN DEL AÑO", 
        font=ctk.CTkFont(size=18, weight="bold")
    ).pack(pady=20)
    
    # Frame de selección de año
    selector_frame = ctk.CTkFrame(ventana_principal.main_stats_frame)
    selector_frame.pack(fill="x", padx=20, pady=10)
    
    # Obtener año actual
    año_actual = datetime.now().year
    años_disponibles = [str(año) for año in range(2020, año_actual + 2)]
    
    # Modo: Individual o Comparativo
    modo_var = ctk.StringVar(value="individual")
    
    ctk.CTkLabel(
        selector_frame, 
        text="Modo:", 
        font=ctk.CTkFont(size=14)
    ).pack(side="left", padx=(10, 5))
    
    ctk.CTkRadioButton(
        selector_frame,
        text="Año Individual",
        variable=modo_var,
        value="individual"
    ).pack(side="left", padx=5)
    
    ctk.CTkRadioButton(
        selector_frame,
        text="Comparativa Multi-Año",
        variable=modo_var,
        value="comparativo"
    ).pack(side="left", padx=5)
    
    # Selector de año individual
    año_var = ctk.StringVar(value=str(año_actual))
    
    año_label = ctk.CTkLabel(
        selector_frame, 
        text="Año:", 
        font=ctk.CTkFont(size=14)
    )
    año_label.pack(side="left", padx=(20, 5))
    
    año_selector = ctk.CTkComboBox(
        selector_frame,
        values=años_disponibles,
        variable=año_var,
        width=120
    )
    año_selector.pack(side="left", padx=5)
    
    # Selectores para modo comparativo
    años_comparar_label = ctk.CTkLabel(
        selector_frame,
        text="Años a comparar:",
        font=ctk.CTkFont(size=12)
    )
    
    # Frame para checkboxes de años
    años_frame = ctk.CTkFrame(selector_frame)
    años_checkboxes = {}
    años_vars = {}
    
    # Crear checkboxes para cada año (en horizontal)
    for i, año in enumerate(años_disponibles):
        var = ctk.IntVar(value=1 if año == str(año_actual) else 0)
        años_vars[año] = var
        cb = ctk.CTkCheckBox(
            años_frame,
            text=año,
            variable=var,
            width=60
        )
        cb.grid(row=0, column=i, padx=2, pady=2)
        años_checkboxes[año] = cb
    
    # Frame para mostrar las estadísticas
    stats_display_frame = ctk.CTkScrollableFrame(ventana_principal.main_stats_frame)
    stats_display_frame.pack(fill="both", expand=True, padx=20, pady=10)
    
    def cambiar_modo():
        """Cambia entre modo individual y comparativo"""
        modo = modo_var.get()
        
        if modo == "individual":
            # Ocultar checkboxes, mostrar selector individual
            años_comparar_label.pack_forget()
            años_frame.pack_forget()
            btn_exportar.pack_forget()
            año_label.pack(side="left", padx=(20, 5))
            año_selector.pack(side="left", padx=5)
            btn_exportar_individual.pack(side="left", padx=5)
        else:
            # Ocultar selector individual, mostrar checkboxes
            año_label.pack_forget()
            año_selector.pack_forget()
            btn_exportar_individual.pack_forget()
            años_comparar_label.pack(side="left", padx=(20, 5))
            años_frame.pack(side="left", padx=5)
            btn_exportar.pack(side="left", padx=5)
    
    def cargar_estadisticas():
        """Carga y muestra las estadísticas según el modo seleccionado"""
        modo = modo_var.get()
        
        if modo == "individual":
            cargar_estadisticas_individual()
        else:
            cargar_estadisticas_comparativas()
    
    def cargar_estadisticas_individual():
        """Carga y muestra las estadísticas del año seleccionado"""
        # Limpiar display
        for widget in stats_display_frame.winfo_children():
            widget.destroy()
        
        año = int(año_var.get())
        
        # Obtener conexión a la BD
        conn, cursor = ventana_principal.master.conectar_db()
        if not conn:
            ctk.CTkLabel(
                stats_display_frame, 
                text="Error al conectar con la base de datos.", 
                text_color="red"
            ).pack(pady=20)
            return
        
        # Calcular estadísticas
        stats = obtener_estadisticas_anuales(conn, año)
        conn.close()
        
        if not stats:
            ctk.CTkLabel(
                stats_display_frame, 
                text="Error al calcular estadísticas.", 
                text_color="red"
            ).pack(pady=20)
            return
        
        # Mostrar título del año
        ctk.CTkLabel(
            stats_display_frame,
            text=f"Resumen del Año {año}",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=(0, 20))
        
        # Grid de KPIs principales
        kpis_frame = ctk.CTkFrame(stats_display_frame, fg_color="transparent")
        kpis_frame.pack(fill="x", pady=10)
        kpis_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # KPI 1: Total Creados
        crear_kpi(
            kpis_frame, 0, 
            "📝 Expedientes Creados", 
            str(stats['total_creados']),
            "#e3f2fd"
        )
        
        # KPI 2: Expedientes Abiertos
        crear_kpi(
            kpis_frame, 1, 
            "📂 Abiertos", 
            str(stats['abiertos']),
            "#fff3e0"
        )
        
        # KPI 3: Pendientes de Autorizar
        crear_kpi(
            kpis_frame, 2, 
            "⏳ Pendientes Autorización", 
            str(stats['pendientes_autorizar']),
            "#fce4ec"
        )
        
        # KPI 4: Cerrados
        crear_kpi(
            kpis_frame, 3, 
            "✅ Cerrados", 
            str(stats['cerrados']),
            "#e8f5e9"
        )
        
        # Segunda fila de KPIs
        kpis_frame2 = ctk.CTkFrame(stats_display_frame, fg_color="transparent")
        kpis_frame2.pack(fill="x", pady=10)
        kpis_frame2.grid_columnconfigure((0, 1, 2), weight=1)
        
        # KPI 5: Total Euros Cerrados
        crear_kpi(
            kpis_frame2, 0, 
            "💰 Total € Cerrados", 
            f"{stats['total_euros_cerrados']:,.2f} €",
            "#f3e5f5"
        )
        
        # KPI 6: Total Euros Mal Estado
        crear_kpi(
            kpis_frame2, 1, 
            "⚠️ € en Mal Estado", 
            f"{stats['total_euros_mal_estado']:,.2f} €",
            "#ffebee"
        )
        
        # KPI 7: Tasa de Cierre
        crear_kpi(
            kpis_frame2, 2, 
            "📈 Tasa de Cierre", 
            f"{stats['tasa_cierre']:.1f}%",
            "#e0f2f1"
        )
        
        # Separador
        ctk.CTkFrame(stats_display_frame, height=2, fg_color="#cccccc").pack(fill="x", pady=20)
        
        # Sección: Top Incidencias
        ctk.CTkLabel(
            stats_display_frame,
            text="🔝 Top Incidencias",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(10, 15))
        
        top_frame = ctk.CTkFrame(stats_display_frame)
        top_frame.pack(fill="x", padx=10, pady=5)
        top_frame.grid_columnconfigure((0, 1), weight=1)
        
        # Producto con más incidencias
        producto_frame = ctk.CTkFrame(top_frame, fg_color="#fff9c4", corner_radius=10)
        producto_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        ctk.CTkLabel(
            producto_frame,
            text="🔧 Producto con Más Incidencias",
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(pady=(10, 5))
        
        ctk.CTkLabel(
            producto_frame,
            text=stats['producto_top'],
            font=ctk.CTkFont(size=14)
        ).pack(pady=5)
        
        ctk.CTkLabel(
            producto_frame,
            text=f"{stats['producto_top_cantidad']} incidencias",
            font=ctk.CTkFont(size=11),
            text_color="gray"
        ).pack(pady=(0, 10))
        
        # Cliente con más incidencias
        cliente_frame = ctk.CTkFrame(top_frame, fg_color="#e1f5fe", corner_radius=10)
        cliente_frame.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        
        ctk.CTkLabel(
            cliente_frame,
            text="👤 Cliente con Más Incidencias",
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(pady=(10, 5))
        
        ctk.CTkLabel(
            cliente_frame,
            text=stats['cliente_top'],
            font=ctk.CTkFont(size=14)
        ).pack(pady=5)
        
        ctk.CTkLabel(
            cliente_frame,
            text=f"{stats['cliente_top_cantidad']} expedientes",
            font=ctk.CTkFont(size=11),
            text_color="gray"
        ).pack(pady=(0, 10))
        
        # Separador
        ctk.CTkFrame(stats_display_frame, height=2, fg_color="#cccccc").pack(fill="x", pady=20)
        
        # Sección: Otras Métricas
        ctk.CTkLabel(
            stats_display_frame,
            text="📋 Otras Métricas",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(10, 15))
        
        otras_frame = ctk.CTkFrame(stats_display_frame)
        otras_frame.pack(fill="x", padx=10, pady=5)
        
        # Métrica: Total Artículos
        crear_metrica_fila(
            otras_frame,
            "📦 Total de Artículos Procesados:",
            str(stats['total_articulos'])
        )
        
        # Métrica: Tiempo Promedio
        crear_metrica_fila(
            otras_frame,
            "⏱️ Tiempo Promedio de Tramitación:",
            f"{stats['dias_tramitacion_promedio']} días"
        )
        
        # Métrica: Resultado más común
        crear_metrica_fila(
            otras_frame,
            "📊 Resultado Más Común:",
            f"{stats['resultado_top']} ({stats['resultado_top_cantidad']} veces)"
        )
    
    def cargar_estadisticas_comparativas():
        """Carga y muestra las estadísticas comparativas de múltiples años"""
        # Limpiar display
        for widget in stats_display_frame.winfo_children():
            widget.destroy()
        
        # Obtener años seleccionados desde checkboxes
        años_seleccionados = [int(año) for año, var in años_vars.items() if var.get() == 1]
        
        if not años_seleccionados:
            ctk.CTkLabel(
                stats_display_frame,
                text="Por favor, seleccione al menos un año para comparar",
                text_color="orange"
            ).pack(pady=20)
            return
        
        años_seleccionados.sort()
        
        # Obtener conexión a la BD
        conn, cursor = ventana_principal.master.conectar_db()
        if not conn:
            ctk.CTkLabel(
                stats_display_frame,
                text="Error al conectar con la base de datos.",
                text_color="red"
            ).pack(pady=20)
            return
        
        # Calcular estadísticas para cada año
        stats_por_año = {}
        for año in años_seleccionados:
            stats = obtener_estadisticas_anuales(conn, año)
            if stats:
                stats_por_año[año] = stats
        
        conn.close()
        
        if not stats_por_año:
            ctk.CTkLabel(
                stats_display_frame,
                text="Error al calcular estadísticas.",
                text_color="red"
            ).pack(pady=20)
            return
        
        # Título
        años_texto = ", ".join(str(a) for a in años_seleccionados)
        ctk.CTkLabel(
            stats_display_frame,
            text=f"Comparativa de Años: {años_texto}",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=(0, 20))
        
        # Crear tabla comparativa
        tabla_frame = ctk.CTkFrame(stats_display_frame)
        tabla_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Configurar grid
        num_años = len(años_seleccionados)
        tabla_frame.grid_columnconfigure(0, weight=1, minsize=250)
        for i in range(num_años):
            tabla_frame.grid_columnconfigure(i + 1, weight=1, minsize=120)
        
        # Cabecera
        ctk.CTkLabel(
            tabla_frame,
            text="Métrica",
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#1976D2",
            corner_radius=5,
            text_color="white"
        ).grid(row=0, column=0, sticky="ew", padx=2, pady=2)
        
        for i, año in enumerate(años_seleccionados):
            ctk.CTkLabel(
                tabla_frame,
                text=str(año),
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color="#1976D2",
                corner_radius=5,
                text_color="white"
            ).grid(row=0, column=i + 1, sticky="ew", padx=2, pady=2)
        
        # Datos
        metricas = [
            ("📝 Expedientes Creados", "total_creados", "int"),
            ("📂 Abiertos", "abiertos", "int"),
            ("⏳ Pendientes Autorización", "pendientes_autorizar", "int"),
            ("✅ Cerrados", "cerrados", "int"),
            ("📈 Tasa de Cierre (%)", "tasa_cierre", "percent"),
            ("💰 Total € Cerrados", "total_euros_cerrados", "euro"),
            ("⚠️ € Mal Estado", "total_euros_mal_estado", "euro"),
            ("📦 Total Artículos", "total_articulos", "int"),
            ("⏱️ Días Tramitación Promedio", "dias_tramitacion_promedio", "float"),
            ("🔧 Producto Top", "producto_top", "text"),
            ("  └─ Incidencias", "producto_top_cantidad", "int"),
            ("👤 Cliente Top", "cliente_top", "text"),
            ("  └─ Expedientes", "cliente_top_cantidad", "int"),
            ("📊 Resultado Más Común", "resultado_top", "text"),
            ("  └─ Veces", "resultado_top_cantidad", "int"),
        ]
        
        fila = 1
        for label, key, tipo in metricas:
            # Color alternado para filas
            bg_color = "#f5f5f5" if fila % 2 == 0 else "transparent"
            
            # Etiqueta de métrica
            ctk.CTkLabel(
                tabla_frame,
                text=label,
                font=ctk.CTkFont(size=11),
                anchor="w",
                fg_color=bg_color
            ).grid(row=fila, column=0, sticky="ew", padx=5, pady=2)
            
            # Valores por año
            for i, año in enumerate(años_seleccionados):
                valor = stats_por_año[año].get(key, "N/A")
                
                # Formatear según tipo
                if tipo == "int":
                    texto = str(valor)
                elif tipo == "float":
                    texto = f"{valor:.1f}"
                elif tipo == "percent":
                    texto = f"{valor:.1f}%"
                elif tipo == "euro":
                    texto = f"{valor:,.2f} €"
                else:  # text
                    texto = str(valor)
                
                ctk.CTkLabel(
                    tabla_frame,
                    text=texto,
                    font=ctk.CTkFont(size=11),
                    anchor="center",
                    fg_color=bg_color
                ).grid(row=fila, column=i + 1, sticky="ew", padx=2, pady=2)
            
            fila += 1
    
    def exportar_individual_excel():
        """Exporta las estadísticas del año individual a Excel"""
        año = int(año_var.get())
        
        # Obtener conexión a la BD
        conn, cursor = ventana_principal.master.conectar_db()
        if not conn:
            messagebox.showerror("Error", "No se pudo conectar a la base de datos")
            return
        
        # Calcular estadísticas
        stats = obtener_estadisticas_anuales(conn, año)
        conn.close()
        
        if not stats:
            messagebox.showerror("Error", "No se pudieron calcular las estadísticas")
            return
        
        # Solicitar archivo de destino
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"estadisticas_anuales_{año}.xlsx"
        )
        
        if not archivo:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Año {año}"
            
            # Estilos
            titulo_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            titulo_font = Font(bold=True, color="FFFFFF", size=14)
            header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            header_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:C1')
            cell = ws['A1']
            cell.value = f"ESTADÍSTICAS ANUALES - AÑO {año}"
            cell.fill = titulo_fill
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Cabecera
            fila = 3
            ws['A3'] = "Métrica"
            ws['B3'] = "Valor"
            ws['C3'] = "Descripción"
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}3']
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            # Datos
            fila = 4
            datos = [
                ("Expedientes Creados", stats['total_creados'], "int", "Total de expedientes creados en el año"),
                ("Expedientes Abiertos", stats['abiertos'], "int", "Expedientes en proceso"),
                ("Pendientes de Autorización", stats['pendientes_autorizar'], "int", "Expedientes pendientes de autorizar"),
                ("Expedientes Cerrados", stats['cerrados'], "int", "Expedientes completados"),
                ("Tasa de Cierre", stats['tasa_cierre'], "percent", "Porcentaje de expedientes cerrados"),
                ("Total € Cerrados", stats['total_euros_cerrados'], "euro", "Suma total de expedientes cerrados"),
                ("Total € Mal Estado", stats['total_euros_mal_estado'], "euro", "Suma de productos en mal estado"),
                ("Total Artículos", stats['total_articulos'], "int", "Artículos procesados en el año"),
                ("Días Tramitación Promedio", stats['dias_tramitacion_promedio'], "float", "Tiempo promedio de tramitación"),
                ("Producto con Más Incidencias", stats['producto_top'], "text", f"{stats['producto_top_cantidad']} incidencias"),
                ("Cliente con Más Incidencias", stats['cliente_top'], "text", f"{stats['cliente_top_cantidad']} expedientes"),
                ("Resultado Más Común", stats['resultado_top'], "text", f"{stats['resultado_top_cantidad']} veces"),
            ]
            
            for metrica, valor, tipo, descripcion in datos:
                ws[f'A{fila}'] = metrica
                ws[f'C{fila}'] = descripcion
                
                # Formatear valor según tipo
                cell_valor = ws[f'B{fila}']
                if tipo == "int":
                    cell_valor.value = valor
                    cell_valor.number_format = '#,##0'
                elif tipo == "float":
                    cell_valor.value = valor
                    cell_valor.number_format = '0.0'
                elif tipo == "percent":
                    cell_valor.value = valor / 100
                    cell_valor.number_format = '0.0%'
                elif tipo == "euro":
                    cell_valor.value = valor
                    cell_valor.number_format = '#,##0.00 €'
                else:
                    cell_valor.value = valor
                
                # Aplicar bordes
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{fila}'].border = border
                
                fila += 1
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            
            # Guardar
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Estadísticas exportadas correctamente a:\n{archivo}")
            
        except ImportError:
            messagebox.showerror("Error", "No se pudo importar openpyxl. Instale el paquete con: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{e}")
            import traceback
            traceback.print_exc()
    
    def exportar_comparativa_excel():
        """Exporta la comparativa de años a Excel"""
        # Obtener años seleccionados
        años_seleccionados = [int(año) for año, var in años_vars.items() if var.get() == 1]
        
        if not años_seleccionados:
            messagebox.showwarning("Advertencia", "Por favor, seleccione al menos un año para exportar")
            return
        
        años_seleccionados.sort()
        
        # Obtener conexión a la BD
        conn, cursor = ventana_principal.master.conectar_db()
        if not conn:
            messagebox.showerror("Error", "No se pudo conectar a la base de datos")
            return
        
        # Calcular estadísticas para cada año
        stats_por_año = {}
        for año in años_seleccionados:
            stats = obtener_estadisticas_anuales(conn, año)
            if stats:
                stats_por_año[año] = stats
        
        conn.close()
        
        if not stats_por_año:
            messagebox.showerror("Error", "No se pudieron calcular las estadísticas")
            return
        
        # Solicitar archivo de destino
        años_texto = "_".join(str(a) for a in años_seleccionados)
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"comparativa_anual_{años_texto}.xlsx"
        )
        
        if not archivo:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparativa Anual"
            
            # Estilos
            titulo_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            titulo_font = Font(bold=True, color="FFFFFF", size=14)
            header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            header_font = Font(bold=True, size=11)
            alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            num_cols = len(años_seleccionados) + 1
            ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
            cell = ws['A1']
            años_texto_titulo = ", ".join(str(a) for a in años_seleccionados)
            cell.value = f"COMPARATIVA ESTADÍSTICAS ANUALES - AÑOS: {años_texto_titulo}"
            cell.fill = titulo_fill
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Cabecera
            ws['A3'] = "Métrica"
            ws['A3'].fill = header_fill
            ws['A3'].font = header_font
            ws['A3'].border = border
            
            for i, año in enumerate(años_seleccionados):
                col = get_column_letter(i + 2)
                cell = ws[f'{col}3']
                cell.value = str(año)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            metricas = [
                ("Expedientes Creados", "total_creados", "int"),
                ("Abiertos", "abiertos", "int"),
                ("Pendientes Autorización", "pendientes_autorizar", "int"),
                ("Cerrados", "cerrados", "int"),
                ("Tasa de Cierre (%)", "tasa_cierre", "percent"),
                ("Total € Cerrados", "total_euros_cerrados", "euro"),
                ("€ Mal Estado", "total_euros_mal_estado", "euro"),
                ("Total Artículos", "total_articulos", "int"),
                ("Días Tramitación Promedio", "dias_tramitacion_promedio", "float"),
                ("Producto Top", "producto_top", "text"),
                ("  Incidencias Producto", "producto_top_cantidad", "int"),
                ("Cliente Top", "cliente_top", "text"),
                ("  Expedientes Cliente", "cliente_top_cantidad", "int"),
                ("Resultado Más Común", "resultado_top", "text"),
                ("  Veces Resultado", "resultado_top_cantidad", "int"),
            ]
            
            fila = 4
            for label, key, tipo in metricas:
                # Color alternado
                fill = alt_fill if fila % 2 == 0 else None
                
                # Etiqueta
                cell = ws[f'A{fila}']
                cell.value = label
                cell.border = border
                if fill:
                    cell.fill = fill
                
                # Valores por año
                for i, año in enumerate(años_seleccionados):
                    col = get_column_letter(i + 2)
                    cell = ws[f'{col}{fila}']
                    valor = stats_por_año[año].get(key, "N/A")
                    
                    # Formatear según tipo
                    if tipo == "int":
                        cell.value = valor if isinstance(valor, int) else 0
                        cell.number_format = '#,##0'
                    elif tipo == "float":
                        cell.value = valor if isinstance(valor, (int, float)) else 0
                        cell.number_format = '0.0'
                    elif tipo == "percent":
                        cell.value = valor / 100 if isinstance(valor, (int, float)) else 0
                        cell.number_format = '0.0%'
                    elif tipo == "euro":
                        cell.value = valor if isinstance(valor, (int, float)) else 0
                        cell.number_format = '#,##0.00 €'
                    else:
                        cell.value = str(valor)
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if fill:
                        cell.fill = fill
                
                fila += 1
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 30
            for i in range(len(años_seleccionados)):
                col = get_column_letter(i + 2)
                ws.column_dimensions[col].width = 15
            
            # Guardar
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Comparativa exportada correctamente a:\n{archivo}")
            
        except ImportError:
            messagebox.showerror("Error", "No se pudo importar openpyxl. Instale el paquete con: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{e}")
            import traceback
            traceback.print_exc()
    
    # Botón para exportar individual
    btn_exportar_individual = ctk.CTkButton(
        selector_frame,
        text="📥 Exportar Excel",
        command=exportar_individual_excel,
        fg_color="#4CAF50",
        hover_color="#45a049",
        width=140
    )
    
    # Botón para exportar comparativa
    btn_exportar = ctk.CTkButton(
        selector_frame,
        text="📥 Exportar Excel",
        command=exportar_comparativa_excel,
        fg_color="#4CAF50",
        hover_color="#45a049",
        width=140
    )
    
    # Botón para cargar estadísticas
    btn_cargar = ctk.CTkButton(
        selector_frame,
        text="📊 Cargar Estadísticas",
        command=cargar_estadisticas,
        fg_color="#2196F3",
        hover_color="#1976D2",
        width=180
    )
    btn_cargar.pack(side="left", padx=10)
    
    # Vincular cambio de modo
    modo_var.trace_add("write", lambda *args: cambiar_modo())
    
    # Cargar automáticamente el año actual
    cargar_estadisticas()


def crear_kpi(parent, column, titulo, valor, color):
    """
    Crea un widget KPI estilizado
    
    Args:
        parent: Frame padre
        column: Columna en el grid
        titulo: Título del KPI
        valor: Valor a mostrar
        color: Color de fondo
    """
    kpi_frame = ctk.CTkFrame(parent, fg_color=color, corner_radius=10)
    kpi_frame.grid(row=0, column=column, padx=5, pady=5, sticky="ew")
    
    ctk.CTkLabel(
        kpi_frame,
        text=titulo,
        font=ctk.CTkFont(size=11)
    ).pack(pady=(10, 5))
    
    ctk.CTkLabel(
        kpi_frame,
        text=valor,
        font=ctk.CTkFont(size=20, weight="bold")
    ).pack(pady=(0, 10))


def crear_metrica_fila(parent, etiqueta, valor):
    """
    Crea una fila de métrica con etiqueta y valor
    
    Args:
        parent: Frame padre
        etiqueta: Texto de la etiqueta
        valor: Valor a mostrar
    """
    fila_frame = ctk.CTkFrame(parent, fg_color="transparent")
    fila_frame.pack(fill="x", pady=5, padx=10)
    
    ctk.CTkLabel(
        fila_frame,
        text=etiqueta,
        font=ctk.CTkFont(size=12),
        anchor="w"
    ).pack(side="left", padx=5)
    
    ctk.CTkLabel(
        fila_frame,
        text=valor,
        font=ctk.CTkFont(size=12, weight="bold"),
        anchor="e"
    ).pack(side="right", padx=5)
